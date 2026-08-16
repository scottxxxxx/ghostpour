"""Document template registry (phase 2 pilot: Smartsheet-style Gantt).

The design lesson from the live prompt-only experiment (2026-07-12): the
model followed the spec where it got hex codes and improvised everywhere
else, took ~8 minutes, and will drift run to run. So the LLM never draws.
A template turn asks the model for STRUCTURED JSON ONLY (one cheap text
turn, no sandbox), and a deterministic renderer here draws the identical
file every time — seconds, pennies, byte-stable styling.

Registry entries pair an extraction prompt with a renderer; the two
version together, which is why the schema lives HERE and never in a
client prompt. Triage: the intent classifier's ask matches template
hints; the offer proposes the template; a confirm routes to this lane;
anything custom falls through to ad-hoc sandbox generation, which stays
the never-locked-in fallback.
"""

from __future__ import annotations

import hashlib
import json
import logging
import re
from datetime import date, datetime, timedelta, timezone
from io import BytesIO

logger = logging.getLogger("ghostpour.doc_templates")

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# What earns a row, shared by both extraction prompts.
#
# "Extract every task and milestone discussed" was the whole instruction,
# and against project memory it is badly under-specified. The memory block
# the model reads is GROUPED UNDER MEETING HEADINGS, so "discussed" invited
# it to enumerate the meetings themselves: a real export came back with 29
# rows of which 22 were meeting titles ("Weekly Status Sync on CKS and ABM
# Items"), each a single day long because a meeting happens on one date,
# chained one per working day into a critical path made of somebody's
# calendar. The progress curve then measured the share of MEETINGS whose
# subject sounded finished, and fell from 50% to 38% when later meetings
# reopened topics, which is not a project going backwards.
#
# Scott, 2026-08-13: meetings should not become tasks, only commitments
# should. Naming the container explicitly is what makes that instruction
# land, because the model is not being careless, it is reading a document
# whose top-level structure is meetings and being told to extract
# everything in it.
_WHAT_EARNS_A_ROW = (
    "What earns a row: a task is work somebody does. It has an actor and "
    "an outcome and it can still be unfinished. A meeting is not work. "
    "The memory you are reading is grouped under meeting headings, and "
    "those headings are containers for what was said, never rows in the "
    "plan: never emit a task or a milestone for a meeting, standup, "
    "sync, review or call happening, and never name a row after one. "
    "Extract what people committed to INSIDE those meetings: the "
    "deliverables, the action items, the fixes and the decisions that "
    "carry dates. A commitment that has only a due date, with no span of "
    "work stated, is a milestone rather than a task. Extract every such "
    "task and milestone discussed, and nothing else."
)

_GANTT_SCHEMA_PROMPT = (
    "Extract this project's plan from the conversation and meeting content "
    "as JSON ONLY, no prose, no code fences. Schema: {\"project\": str, "
    "\"meeting_date\": \"YYYY-MM-DD\"|null, "
    "\"tasks\": [{\"id\": int, \"name\": str, \"type\": \"phase\"|\"task\"|"
    "\"milestone\", \"parent_id\": int|null, \"owner\": str|null, "
    "\"status\": \"complete\"|\"in_progress\"|\"on_hold\"|\"not_started\"|"
    "\"blocked\", \"start\": \"YYYY-MM-DD\", \"end\": \"YYYY-MM-DD\", "
    "\"depends_on\": [int], \"percent_complete\": int|null}]}. "
    "Rules: phases have parent_id null; every task and milestone "
    "carries the id of a phase in this same list, and when the "
    "meetings suggest no grouping at all, emit one phase named for "
    "the project and put everything under it; tasks and "
    "milestones carry the id of their phase; milestones have start equal to "
    "end; dates must be consistent with dependencies (a task never starts "
    "before its predecessor ends); owner is the person's name as spoken; "
    "meeting_date is the date of the meeting this plan comes from as stated "
    "in the content (the most recent one when several), null when no date "
    "is evident. percent_complete is STRICTLY what a person stated in the "
    "content (\"about 80 percent\" is 80): when nobody stated a value, use "
    "null; never estimate, and never infer a percent from status. "
    "Compose project and name strings without em or en dashes, and never "
    "imitate a dash with a spaced hyphen (word - word): a hyphen may "
    "appear only inside a hyphenated word. Where a dash would fit, use "
    "a colon or comma instead, even when the source content uses one. "
    + _WHAT_EARNS_A_ROW +
    " Output only the JSON object."
)

_GANTT_DETAILED_SCHEMA_PROMPT = (
    "Extract this project's plan from the conversation and meeting content "
    "as JSON ONLY, no prose, no code fences. Schema: {\"project\": str, "
    "\"meeting_date\": \"YYYY-MM-DD\"|null, "
    "\"tasks\": [{\"id\": int, \"name\": str, \"type\": \"phase\"|\"task\"|"
    "\"milestone\", \"parent_id\": int|null, \"owner\": str|null, "
    "\"status\": \"complete\"|\"in_progress\"|\"on_hold\"|\"not_started\"|"
    "\"blocked\", \"start\": \"YYYY-MM-DD\", \"end\": \"YYYY-MM-DD\", "
    "\"depends_on\": [int], \"percent_complete\": int|null, "
    "\"effort\": str|null, \"conditional\": str|null, "
    "\"commitments\": [{\"date\": \"YYYY-MM-DD\", "
    "\"as_of\": \"YYYY-MM-DD\"|null, \"quote\": str, \"speaker\": str|null, "
    "\"reason\": str|null}], "
    "\"evidence\": [{\"field\": str, \"quote\": str, "
    "\"speaker\": str|null, \"meeting_date\": \"YYYY-MM-DD\"|null}]}]}. "
    "Rules: phases have parent_id null; every task and milestone "
    "carries the id of a phase in this same list, and when the "
    "meetings suggest no grouping at all, emit one phase named for "
    "the project and put everything under it; tasks "
    "and milestones carry the id of their phase; milestones have start "
    "equal to end; dates must be consistent with dependencies (a task "
    "never starts before its predecessor ends); owner is the person's name "
    "as spoken; meeting_date is the date of the meeting this plan comes "
    "from as stated in the content (the most recent one when several), "
    "null when no date is evident. percent_complete and effort are "
    "STRICTLY what a person stated in the content (\"about 80 percent\" "
    "is 80; \"two days of work\" is \"2 days\"): when nobody stated a "
    "value, use null. Never estimate, and never infer a percent from "
    "status. commitments is the due date this task was given each time "
    "somebody stated one, oldest first: a task whose date was stated once "
    "has exactly one entry, and a date that was restated unchanged is not "
    "a new entry. Each entry carries the date stated, the meeting date it "
    "was said in when the content dates its meetings (else null), the "
    "verbatim line, the speaker when identifiable, and a reason ONLY when "
    "the person gave one (\"their sandbox creds took a week\"), else null. "
    "Never invent a commitment for a date nobody stated. conditional is "
    "the condition itself when a date was stated conditionally (from "
    "\"release August 3 if beta is quiet\", conditional is \"if beta is "
    "quiet\"), and null when the date was stated flatly; never mark a date "
    "conditional unless somebody stated the condition. evidence lists "
    "short verbatim quotes from the content that "
    "support extracted values (dates, status, percent_complete, effort, "
    "owner), with field naming which value each quote supports and "
    "meeting_date naming the meeting it was said in when the content "
    "dates its meetings; include "
    "speaker when identifiable; omit evidence you do not have rather than "
    "paraphrasing. Compose project, name, and effort strings without em "
    "or en dashes, and never imitate a dash with a spaced hyphen "
    "(word - word): a hyphen may appear only inside a hyphenated word. "
    "Where a dash would fit, use a colon or comma instead, even when "
    "the source content uses one. Evidence quotes are the one exception and "
    "stay verbatim, exactly as spoken. "
    + _WHAT_EARNS_A_ROW +
    " Output only the JSON object."
)

# palette lifted from the reference artifact (ABM_Gantt_Smartsheet_Style)
_C = {
    "bar": "A8B9C9", "summary": "6E7B8A", "project": "3D4653",
    "weekend": "F3F3F3", "today": "FFF6DE", "risk": "E0341E",
    "conditional": "FBF0D5",
    "risk_done": "9A1B12", "risk_rest": "F1948A",
    "grid": "E9E9E9", "white": "FFFFFF",
    "status": {"complete": "1F4E9C", "in_progress": "2E9E4F",
               "on_hold": "F5A623", "not_started": "E0341E",
               "blocked": "E0341E"},
    "chips": ["1F4E9C", "2E9E4F", "F5A623", "D35400", "7B4EA3", "2C7A7B"],
}


def _d(s):
    return datetime.strptime(str(s)[:10], "%Y-%m-%d").date()


def _points_are_milestones(data: dict) -> dict:
    """Retype zero-duration tasks as milestones.

    A task is a span of work and a milestone is a point in time. The
    schema already says as much ("milestones have start equal to end"), so
    a `task` whose start equals its end is not work, it is an event.

    The structural half of the meetings-are-not-tasks fix. The extraction
    prompt now tells the model that a meeting happening is not a task, but
    a prompt constrains and does not prove, and what it guards against is
    expensive: 22 of 29 rows in a live export were meeting titles, every
    bar one day long, and the progress curve then weighed them as if they
    were work.

    Retyping rather than dropping loses nothing a reader can see. The row
    keeps its name, date, owner and status and draws as a point on the
    timeline. What changes is the arithmetic: phases and milestones carry
    no weight in any of the progress maths (`_total_days`, `_earned_days`
    and `_weighted_progress` all skip them), so an extraction that hands
    us points instead of spans can no longer manufacture a percentage out
    of them. A plan made ENTIRELY of points therefore charts no progress
    curve at all, which is the honest outcome: nothing in it states how
    long any work takes.

    Deliberately not a name heuristic. Whether a string looks like a
    meeting title is a guess; whether a row has a duration is a fact.
    """
    tasks = data.get("tasks") or []
    if not tasks:
        return data
    out, retyped = [], []
    for t in tasks:
        if t.get("type") == "task":
            try:
                if _d(t["start"]) == _d(t["end"]):
                    t = {**t, "type": "milestone"}
                    retyped.append(str(t.get("name") or "")[:60])
            except (KeyError, ValueError, TypeError):
                pass
        out.append(t)
    if not retyped:
        return data
    # Loud on purpose: this fires when the extraction handed us events
    # where work was asked for, and we want that visible rather than
    # quietly corrected.
    logger.info("gantt_zero_duration_tasks_retyped count=%d of=%d names=%r",
                len(retyped), len(tasks), retyped[:5])
    return {**data, "tasks": out}


def _dep_code(pred: dict, succ: dict) -> str:
    """Two-letter dependency nomenclature, DERIVED from the extracted
    dates — never asked of the model (Scott 2026-07-15: the model would
    invent the minority types; dates it already committed to are
    arithmetic). Starts align -> SS, ends align -> FF, everything else —
    including anything ambiguous — defaults to FS. The user can always
    correct a cell."""
    if _d(succ["start"]) == _d(pred["start"]):
        return "SS"
    if _d(succ["end"]) == _d(pred["end"]):
        return "FF"
    return "FS"


_STATUS_LABELS = {
    "complete": "Complete", "in_progress": "In Progress",
    "on_hold": "On Hold", "not_started": "Not Started", "blocked": "Blocked",
}


def render_gantt(data: dict, *, today: date | None = None,
                 history: list[dict] | None = None) -> bytes:
    """Deterministic Smartsheet-style Gantt from extracted plan JSON.
    `history` is accepted for renderer-signature parity (the template
    lane passes it to every renderer) and unused by the simple style."""
    wb, _ = _build_gantt_wb(
        _split_undated(_with_workdays(_points_are_milestones(data))),
        today=today)
    return _serialize_wb(wb)


# --- Working-day scheduling ------------------------------------------------
# Weekends were shaded on the grid (WEEKDAY()>5) but the date arithmetic was
# calendar days, so the sheet contradicted itself the moment anyone edited:
# pushing a predecessor could land a dependent's Start on a Saturday and then
# draw its bar straight across the weekend it had just greyed out. Dates now
# snap to working days and offsets are counted in working days, so a push
# lands on a weekday and a duration means what a reader assumes it means.
#
# Weekends only. Holidays are locale-specific and the meeting never states
# them, so inventing a calendar would break the only-what-the-meeting-knows
# rule; WORKDAY() takes an optional holiday range if we ever get a real one.

def _next_workday(d: date) -> date:
    while d.weekday() >= 5:
        d += timedelta(days=1)
    return d


def _prev_workday(d: date) -> date:
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


def _snap_span(s: date, e: date) -> tuple[date, date]:
    """Pull a span inward onto working days: a weekend start moves to the
    next Monday, a weekend end back to the previous Friday. A span that
    lives entirely on one weekend collapses to a single working day rather
    than inverting."""
    s2, e2 = _next_workday(s), _prev_workday(e)
    if e2 < s2:
        e2 = s2
    return s2, e2


def _workday_offset(a: date, b: date) -> int:
    """Signed working-day offset k such that Excel's WORKDAY(a, k) == b,
    for a and b on working days. Counts working days in (a, b] going
    forward, or negates the count in [b, a) going back."""
    if a == b:
        return 0
    step = 1 if b > a else -1
    lo, hi = (a, b) if b > a else (b, a)
    n, cur = 0, lo
    while cur < hi:
        cur += timedelta(days=1)
        if cur.weekday() < 5:
            n += 1
    return n * step


def _workday_formula(expr: str, k: int) -> str:
    """WORKDAY() moves by working days. k == 0 stays a bare reference: the
    initial value must equal the extracted date byte for byte, and
    WORKDAY(x, 0) is not a guaranteed identity for every viewer."""
    return expr if k == 0 else f"WORKDAY({expr},{k})"


def _with_workdays(data: dict) -> dict:
    """Copy of `data` whose task spans sit on working days. Applied once at
    the renderer entry point so the view, the Slip sheet and the Receipts
    sheet all read the same dates."""
    tasks = []
    for t in (data.get("tasks") or []):
        t = dict(t)
        try:
            s, e = _snap_span(_d(t["start"]), _d(t["end"]))
        except (KeyError, ValueError, TypeError):
            tasks.append(t)
            continue
        t["start"], t["end"] = s.isoformat(), e.isoformat()
        tasks.append(t)
    return {**data, "tasks": tasks}


def _split_undated(data: dict) -> dict:
    """Date coverage honesty (PM review 2026-08-11): an activity the
    meetings never dated cannot sit on a timeline, and before this pass it
    took the whole render down instead (the axis min() raised on its
    dates). Undated work is now counted and disclosed as "dates on N of M
    activities" rather than drawn, so the gap stays visible instead of
    hidden. Applied at the renderer entry points, after _with_workdays,
    which deliberately passes unparseable dates through untouched."""
    kept: list[dict] = []
    real_total = real_dated = 0
    for t in (data.get("tasks") or []):
        real = t.get("type") != "phase"
        try:
            _d(t["start"]), _d(t["end"])
        except (KeyError, ValueError, TypeError):
            if real:
                real_total += 1
            logger.info("gantt_undated_activity name=%r", t.get("name"))
            continue
        if real:
            real_total += 1
            real_dated += 1
        kept.append(t)
    return {**data, "tasks": kept,
            "_date_coverage": (real_dated, real_total)}


def _build_gantt_wb(data: dict, *, today: date | None = None,
                    detail_cols: bool = False):
    """Build the Gantt View workbook (shared by the simple and detailed
    renderers; the detailed one appends sheets to the same workbook).

    detail_cols (detailed style only): adds % Done and Effort columns
    BETWEEN Assigned To and the day grid, and a live completed-portion
    overlay on each bar driven by the % cell (edit the percent in Excel
    and the done-portion redraws, same live-grid rule as everything
    else). Column positions A-I are unchanged so every $D/$E/$F formula
    is shared between styles; only the day-grid origin shifts.

    LIVE GRID (Scott 2026-07-15): the timeline bars are conditional
    formatting formulas over real date cells, not painted fills — edit a
    Start/End date in Excel and the bar redraws; flip the Status dropdown
    and the dot recolors; the today column tracks TODAY(). Row 1 is a
    hidden axis of real dates the formulas compare against (the axis
    itself is fixed at build; bars clip at its edges). Weak viewers that
    skip conditional formatting show a plain grid — real Excel and Google
    Sheets render fully."""
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    def fill(hex6):
        return PatternFill("solid", fgColor="FF" + hex6)

    def dxf_fill(hex6):
        return PatternFill(start_color="FF" + hex6, end_color="FF" + hex6,
                           fill_type="solid")

    tasks = data.get("tasks") or []
    if not tasks:
        raise ValueError("no tasks extracted")
    today = today or date.today()
    start = min(_d(t["start"]) for t in tasks)
    end = max(_d(t["end"]) for t in tasks)
    # +28 days of runway so a user can push dates right and the bars
    # still have axis to land on (was +14; critic pass 2026-07-21 —
    # dependency push makes multi-week slides one edit away)
    days = [(start + timedelta(n))
            for n in range((end - start).days + 29)][:180]

    by_id = {t["id"]: t for t in tasks}
    phases = [t for t in tasks if t.get("type") == "phase"]
    children: dict = {p["id"]: [] for p in phases}
    for t in tasks:
        if t.get("type") != "phase" and t.get("parent_id") in children:
            children[t["parent_id"]].append(t)
    # A task whose parent_id matches no phase used to be dropped from the
    # timeline in silence: the schema asks for phases, and when the model
    # returns a flat list instead (live 2026-07-30, 9 of 9 tasks with
    # parent_id null) the Gantt View rendered the project row and nothing
    # else, while Slip and Receipts listed all nine. A grouping the meeting
    # never made is not a reason to lose the work, so orphans hang off the
    # project row in extraction order.
    _orphans = [t for t in tasks
                if t.get("type") != "phase"
                and t.get("parent_id") not in children]
    if _orphans:
        logger.info("gantt_orphan_tasks count=%d phases=%d",
                    len(_orphans), len(phases))

    def at_risk(t):
        # blocked; never started though its start passed; or OVERDUE —
        # end behind today while anything but complete (folded in from
        # the retired Progress sheet, Scott 2026-07-21)
        return t["status"] == "blocked" or (
            t["status"] == "not_started" and _d(t["start"]) < today) or (
            t["status"] != "complete" and _d(t["end"]) < today)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt View"

    # As-of stamp + one-line status strip (PM audit 2026-07-23: a plan
    # without a timestamp is a rumor, and the 40-second reader needs a
    # sentence, not a chart). Static snapshot text; the as-of date is
    # exactly what makes that honest.
    _real = [t for t in tasks if t.get("type") != "phase"]
    _wsum = sum((_d(t["end"]) - _d(t["start"])).days for t in _real)
    _psum = sum((_d(t["end"]) - _d(t["start"])).days * (
        (t.get("percent_complete") or 0) / 100
        if isinstance(t.get("percent_complete"), int)
        else (1.0 if t.get("status") == "complete" else 0.0))
        for t in _real)
    _overall = round(100 * _psum / _wsum) if _wsum else None
    _risk_n = sum(1 for t in _real if at_risk(t))
    _next_ms = min(
        (t for t in _real if t.get("type") == "milestone"
         and t.get("status") != "complete" and _d(t["start"]) >= today),
        key=lambda t: _d(t["start"]), default=None)
    _bits = [f"As of the {data.get('meeting_date') or today.isoformat()} "
             f"standup", f"generated {today.isoformat()}",
             f"{len(_real)} tasks"]
    # date coverage on the export itself (PM review 2026-08-11): an
    # activity nobody dated is not on the timeline, and the reader must
    # see that from the sheet, never discover it later
    _cov = data.get("_date_coverage")
    if _cov and _cov[0] < _cov[1]:
        _bits.append(f"dates on {_cov[0]} of {_cov[1]} activities")
    if _overall is not None:
        _bits.append(f"{_overall}% complete overall")
    # "At risk" is evaluated LIVE against today, not against the meeting the
    # plan came from, which is right for a plan you keep using and confusing
    # for one you read cold: a workbook built from a Jul 20 standup and
    # opened on Jul 31 shows everything due in between as at risk, and the
    # count grows every day the file sits there. So the strip says which
    # date the count is against rather than leaving the reader to assume.
    _bits.append(f"{_risk_n} at risk as of {today.strftime('%b %d')}"
                 if _risk_n else f"none at risk as of {today.strftime('%b %d')}")
    if _next_ms:
        _bits.append(f"next milestone: {_next_ms['name']} "
                     f"{_d(_next_ms['start']).strftime('%b %d')}")
    strip = ws.cell(2, 2, "  " + "  ·  ".join(_bits))
    strip.font = Font(bold=True, size=9, color="FF3D4653")
    # A:dot B:name C:risk D:status E:start F:end G:predecessors H:chip I:owner
    FIRST_DAY_COL = 13 if detail_cols else 11
    PCT_COL = "J"   # % Done, only written when detail_cols
    FLOAT_COL = "L"  # Total float in working days, detail_cols only
    KEY_TOP = 4          # status key block under the 3 header rows
    # key + overlay/blank/phase legend lines + project row; the detailed
    # style adds one more legend line (what Float means)
    first_bar_row = KEY_TOP + (10 if detail_cols else 8)

    # Pre-pass: worksheet row of every task, so Predecessors can cite
    # rows in either direction (forward deps included).
    row_of: dict = {}
    r = first_bar_row
    for phase in phases:
        r += 1
        row_of[phase["id"]] = r
        for t in children.get(phase["id"], []):
            r += 1
            row_of[t["id"]] = r
    for t in _orphans:
        r += 1
        row_of[t["id"]] = r
    last_row = r
    last_col = FIRST_DAY_COL + len(days) - 1
    grid = lambda row: (f"{get_column_letter(FIRST_DAY_COL)}{row}:"  # noqa: E731
                        f"{get_column_letter(last_col)}{row}")

    # Row 1 (hidden): the real-date axis every bar formula compares
    # against. Rows 2-3: week-of labels + day letters (visual only).
    for i, d in enumerate(days):
        col = FIRST_DAY_COL + i
        ws.column_dimensions[get_column_letter(col)].width = 3
        ax = ws.cell(1, col, d)
        ax.number_format = "yyyy-mm-dd"
        if d.weekday() == 0:
            c = ws.cell(2, col, f"Week of {d.strftime('%b %d')}")
            c.font = Font(size=8, bold=True)
        letter = "MTWTFSS"[d.weekday()]
        c = ws.cell(3, col, letter)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(size=8, bold=True,
                      color="FF" + (_C["risk"] if d.weekday() >= 5 else "3D4653"))
        if d == today:
            c.font = Font(size=8, bold=True, color="FF" + _C["risk"])
    ws.row_dimensions[1].hidden = True
    # beyond-chart indicator column (critic pass 2026-07-21): a bar
    # pushed past the axis edge used to vanish silently; now the row's
    # last cell shows a red arrow, live off the End cell.
    ws.column_dimensions[get_column_letter(last_col + 1)].width = 3
    _heads = [("", 7), ("Task Name", 38), ("At\nRisk", 5), ("Status", 12),
              ("Start\nDate", 11), ("End\nDate", 11), ("Predecessors", 12),
              ("", 4), ("Assigned To", 16)]
    _heads += [("%\nDone", 7)]
    if detail_cols:
        _heads += [("Effort", 10), ("Float\nwork days", 9)]
    for col, (head, width) in enumerate(_heads, start=1):
        c = ws.cell(3, col, head)
        c.font = Font(bold=True, size=9)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.fill = fill("E9E9E9")   # header band (Scott: "lost the cool highlighting")
        ws.column_dimensions[get_column_letter(col)].width = width

    # status key block — amber header band like the reference
    row = KEY_TOP
    kc = ws.cell(row, 2, "  STATUS KEY")
    kc.font = Font(bold=True, size=9)
    kc.fill = fill("FDF3E3")
    ws.cell(row, 1).fill = fill("F5A623")
    for label, key in (("COMPLETED", "complete"), ("IN PROGRESS", "in_progress"),
                       ("ON HOLD", "on_hold"), ("NOT STARTED", "not_started")):
        row += 1
        ws.cell(row, 1, "●").font = Font(color="FF" + _C["status"][key], size=11)
        ws.cell(row, 2, f"      {label}").font = Font(size=8)
    lg = ws.cell(row + 1, 2,
                 "      █ darker bar section = completed share, drives "
                 "from % Done (deep red when the task is at risk)")
    lg.font = Font(size=8, color="FF" + _C["status"]["complete"])
    lg2 = ws.cell(row + 2, 2,
                  "      blank % Done = nobody stated it in a meeting; "
                  "values are never estimated")
    lg2.font = Font(size=8, color="FF9AA4AF")
    lg3 = ws.cell(row + 3, 2,
                  "      phase % = duration-weighted; unstarted tasks "
                  "count as 0, Complete tasks as 100")
    lg3.font = Font(size=8, color="FF9AA4AF")
    if detail_cols:
        lg4 = ws.cell(row + 4, 2,
                      "      Float = working days this task can slip before "
                      "the project end moves; 0 means it is on the critical "
                      "path (live, recalculates when you edit a date)")
        lg4.font = Font(size=8, color="FF" + _C["risk_done"])
        lg5 = ws.cell(row + 5, 2,
                      "      shaded End Date = the meeting stated a condition "
                      "on that date, not a firm commitment; hover the cell "
                      "for the condition")
        lg5.font = Font(size=8, color="FF9A3412")
    row += 6 if detail_cols else 4

    def date_cells(r, s, e, hex_color=None, size=8, formulas=None):
        """Write Start/End as dates, or as live formulas when the row's
        dates are DRIVEN (dependency push, phase/project rollups). The
        format and font are identical either way; a driven cell shows the
        same date until an upstream edit moves it."""
        for col, d in ((5, s), (6, e)):
            f = (formulas or (None, None))[col - 5]
            c = ws.cell(r, col, f if f else d)
            c.number_format = "yyyy-mm-dd"
            c.font = Font(size=size,
                          color="FF" + (hex_color or "3D4653"))

    # Dependency-driven scheduling (Scott 2026-07-21: "if someone were to
    # change something it wouldn't push out any other task"): a task with
    # predecessors gets Start = MAX over predecessors of (driver cell +
    # the LAG the extracted schedule already shows), and End = Start +
    # its extracted duration. SS predecessors drive from the pred's
    # Start cell, everything else from its End cell. Initial values are
    # byte-for-byte the extracted dates; editing a predecessor pushes
    # every dependent (and the bars redraw — they key off these cells).
    # A cyclic dependency graph disables ALL date formulas (static
    # fallback) rather than shipping circular references; users can
    # still edit dates, they just don't cascade.
    _adj = {t["id"]: [d for d in (t.get("depends_on") or []) if d in by_id]
            for t in tasks}
    _indeg = {tid: len(deps) for tid, deps in _adj.items()}
    _queue = [tid for tid, n in _indeg.items() if n == 0]
    _seen = 0
    _dependents = {tid: [] for tid in _adj}
    for tid, deps in _adj.items():
        for d in deps:
            _dependents[d].append(tid)
    while _queue:
        u = _queue.pop()
        _seen += 1
        for v in _dependents[u]:
            _indeg[v] -= 1
            if _indeg[v] == 0:
                _queue.append(v)
    _deps_acyclic = _seen == len(_adj)

    def dep_formulas(t) -> tuple[str, str] | None:
        """(start_formula, end_formula) for a dependent task, or None
        when its dates stay static."""
        deps = _adj.get(t["id"]) or []
        if not deps or not _deps_acyclic:
            return None
        terms = []
        for dep in deps:
            p = by_id[dep]
            prow = row_of.get(dep)
            if prow is None:
                return None
            if _dep_code(p, t) == "SS":
                terms.append(_workday_formula(
                    f"E{prow}", _workday_offset(_d(p["start"]), _d(t["start"]))))
            else:
                terms.append(_workday_formula(
                    f"F{prow}", _workday_offset(_d(p["end"]), _d(t["start"]))))
        ef = "=" + (terms[0] if len(terms) == 1
                    else "MAX(" + ",".join(terms) + ")")
        r = row_of[t["id"]]
        ff = "=" + _workday_formula(
            f"E{r}", _workday_offset(_d(t["start"]), _d(t["end"])))
        return ef, ff

    def bar_rules(r, bar_hex, risk_aware=False, pct_overlay=False):
        """The live bars, drawn twice from the same date cells so every
        viewer shows them (Scott's Numbers finding 2026-07-16: Numbers
        computes formulas but refuses conditional FORMATTING, so
        fill-only bars vanished there):

        1. A full-block character (█) via per-cell formula, font-colored
           to the bar — Numbers renders colored character bars; in Excel
           the same-color character melts invisibly into the fill.
        2. Conditional-formatting fills for Excel/Sheets, with the FONT
           recolored in the same rule so a live status flip (risk red)
           recolors the characters too and nothing clashes.
        """
        E, F = f"$E{r}", f"$F{r}"
        ax = f"{get_column_letter(FIRST_DAY_COL)}$1"
        in_range = f"AND({ax}>={E},{ax}<={F})"
        for i in range(len(days)):
            col = FIRST_DAY_COL + i
            L = get_column_letter(col)
            c = ws.cell(r, col, f'=IF(AND({L}$1>={E},{L}$1<={F}),"█","")')
            c.font = Font(color="FF" + bar_hex, size=9)
            c.alignment = Alignment(horizontal="center")
        P = f"${PCT_COL}{r}"
        _in_done = f"{ax}>={E},{ax}<={E}+({F}-{E})*{P}"
        _riskcond = (f"OR($D{r}=\"Blocked\","
                     f"AND($D{r}=\"Not Started\",{E}<TODAY()),"
                     f"AND($D{r}<>\"Complete\",$D{r}<>\"\",{F}<TODAY()))")
        if pct_overlay and risk_aware:
            # At-risk rows mirror the healthy scheme in the red family
            # (Scott 2026-07-23): completed share deep red, remaining
            # share light red, so a late task still shows what's banked.
            ws.conditional_formatting.add(grid(r), FormulaRule(
                formula=[f"AND({P}<>\"\",{_in_done},{_riskcond})"],
                fill=dxf_fill(_C["risk_done"]),
                font=Font(color="FF" + _C["risk_done"]), stopIfTrue=True))
        if risk_aware:
            risky = f"AND({ax}>={E},{ax}<={F},{_riskcond})"
            ws.conditional_formatting.add(grid(r), FormulaRule(
                formula=[risky], fill=dxf_fill(_C["risk_rest"]),
                font=Font(color="FF" + _C["risk_rest"]), stopIfTrue=True))
        if pct_overlay:
            # Healthy rows: completed share in status-complete blue over
            # the gray base bar. Blank % means no overlay.
            ws.conditional_formatting.add(grid(r), FormulaRule(
                formula=[f"AND({P}<>\"\",{_in_done})"],
                fill=dxf_fill(_C["status"]["complete"]),
                font=Font(color="FF" + _C["status"]["complete"]),
                stopIfTrue=True))
        ws.conditional_formatting.add(grid(r), FormulaRule(
            formula=[in_range], fill=dxf_fill(bar_hex),
            font=Font(color="FF" + bar_hex), stopIfTrue=True))
        ind = ws.cell(r, last_col + 1,
                      f'=IF($F{r}>{get_column_letter(last_col)}$1,"→","")')
        ind.font = Font(bold=True, color="FF" + _C["risk"])
        ind.alignment = Alignment(horizontal="center")

    # project row — dates roll up live from the phase rows
    ws.cell(row, 2, f"  {data.get('project') or 'Project'}").font = \
        Font(bold=True, color="FFFFFFFF")
    for c in range(1, FIRST_DAY_COL):
        ws.cell(row, c).fill = fill(_C["project"])
    _proj_formulas = None
    if phases and _deps_acyclic:
        _prows = [row_of[p["id"]] for p in phases]
        _proj_formulas = (
            "=MIN(" + ",".join(f"E{r}" for r in _prows) + ")",
            "=MAX(" + ",".join(f"F{r}" for r in _prows) + ")")
    date_cells(row, start, end, hex_color="FFFFFF", size=9,
               formulas=_proj_formulas)
    ws.row_dimensions[row].height = 12
    bar_rules(row, _C["project"])
    row += 1

    dv = DataValidation(
        type="list",
        formula1='"Complete,In Progress,On Hold,Not Started,Blocked"',
        allow_blank=True)
    ws.add_data_validation(dv)

    chip_cache: dict = {}

    def render_task(t):
        """One task or milestone row: a phase's child, or an orphan
        hanging off the project row when the extraction returned no
        usable grouping."""
        nonlocal row
        assert row == row_of[t["id"]]
        risky = at_risk(t)
        text_hex = _C["risk"] if risky else "3D4653"
        ws.cell(row, 1, "●" if t["type"] != "milestone" else "").font = \
            Font(color="FF" + _C["status"].get(t["status"], _C["bar"]))
        name = ("          🏁 " if t["type"] == "milestone" else "          ") + t["name"]
        ws.cell(row, 2, name).font = Font(size=9, color="FF" + text_hex)
        fl = ws.cell(row, 3, "⚑" if risky else "⚐")
        fl.font = Font(color="FF" + (_C["risk"] if risky else "9AA4AF"))
        fl.alignment = Alignment(horizontal="center")
        st = ws.cell(row, 4, _STATUS_LABELS.get(t.get("status", ""), ""))
        st.font = Font(size=8, color="FF" + text_hex)
        dv.add(f"D{row}")
        date_cells(row, _d(t["start"]), _d(t["end"]), hex_color=text_hex,
                   formulas=dep_formulas(t))
        # A conditional date is the one place the sheet would otherwise
        # assert more than the meeting did: "release August 3 if beta's
        # quiet" renders identically to a firm commitment. Mark the END
        # cell rather than the name, which slip matches on.
        cond = str(t.get("conditional") or "").strip()
        if cond:
            ec = ws.cell(row, 6)
            ec.fill = fill(_C["conditional"])
            ec.comment = Comment(f"Conditional: {cond}", "Shoulder Surf")
            ec.comment.width, ec.comment.height = 240, 60
        # Predecessors: Smartsheet nomenclature, dates-derived (FS
        # default; SS/FF only when the extracted dates say so)
        codes = ", ".join(
            f"{row_of[dep]}{_dep_code(by_id[dep], t)}"
            for dep in (t.get("depends_on") or []) if dep in by_id and dep in row_of)
        if codes:
            pc = ws.cell(row, 7, codes)
            pc.font = Font(size=8, color="FF" + text_hex)
            pc.alignment = Alignment(horizontal="center")
        owner = (t.get("owner") or "").strip()
        if owner:
            # chip = colored initials; FULL NAME beside it (Scott's
            # review: "I don't get the full name, just the letter")
            initials = "".join(w[0] for w in owner.split()[:2]).upper()
            hex_c = chip_cache.setdefault(
                owner, _C["chips"][int(hashlib.sha256(owner.encode()).hexdigest(), 16) % len(_C["chips"])])
            chip = ws.cell(row, 8, initials)
            chip.fill = fill(hex_c)
            chip.font = Font(bold=True, size=8, color="FFFFFFFF")
            chip.alignment = Alignment(horizontal="center")
            nm = ws.cell(row, 9, owner)
            nm.font = Font(size=8, color="FF" + text_hex)
        pct = t.get("percent_complete")
        if isinstance(pct, int) and 0 <= pct <= 100:
            pcell = ws.cell(row, 10, pct / 100)
            pcell.number_format = "0%"
            pcell.font = Font(size=8, color="FF" + text_hex)
            pcell.alignment = Alignment(horizontal="center")
        if detail_cols:
            eff = t.get("effort")
            if isinstance(eff, str) and eff.strip():
                ecell = ws.cell(row, 11, eff.strip())
                ecell.font = Font(size=8, color="FF" + text_hex)
                ecell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].outline_level = 2
        if t["type"] == "milestone":
            # The ◆ marker is a formula so it moves with the date —
            # SAME formula shape as the bar cells (Scott 2026-07-16:
            # Excel's inconsistent-formula check stamped green
            # triangles along every milestone row because its formula
            # differed from its neighbors'). A milestone's start
            # equals its end, so the range test marks exactly one day.
            for i in range(len(days)):
                col = FIRST_DAY_COL + i
                L = get_column_letter(col)
                m = ws.cell(row, col,
                            f'=IF(AND({L}$1>=$E{row},{L}$1<=$F{row}),'
                            f'"◆","")')
                m.font = Font(color="FF" + _C["risk"], bold=True)
                m.alignment = Alignment(horizontal="center")
            mi = ws.cell(row, last_col + 1,
                         f'=IF($F{row}>{get_column_letter(last_col)}$1,'
                         f'"→","")')
            mi.font = Font(bold=True, color="FF" + _C["risk"])
            mi.alignment = Alignment(horizontal="center")
        else:
            bar_rules(row, _C["bar"], risk_aware=True, pct_overlay=True)
        row += 1

    for phase in phases:
        assert row == row_of[phase["id"]]
        ws.cell(row, 1, "●").font = Font(
            color="FF" + _C["status"].get(phase.get("status", "in_progress"), _C["bar"]))
        ws.cell(row, 2, f"  −  {phase['name']}").font = Font(bold=True, size=9)
        ws.cell(row, 4, _STATUS_LABELS.get(phase.get("status", ""), "")).font = Font(size=8)
        _kids = children.get(phase["id"], [])
        _phase_formulas = None
        if _kids and _deps_acyclic:
            # children occupy the contiguous rows right under the phase
            _phase_formulas = (f"=MIN(E{row + 1}:E{row + len(_kids)})",
                               f"=MAX(F{row + 1}:F{row + len(_kids)})")
        date_cells(row, _d(phase["start"]), _d(phase["end"]),
                   formulas=_phase_formulas)
        _weighted = any((_d(k["end"]) - _d(k["start"])).days > 0
                        for k in _kids)
        if _kids and _weighted:
            # phase % Done rolls up live, duration-weighted over ALL
            # non-milestone children (Scott 2026-07-23): stated percent
            # when spoken, 100 when status is Complete, 0 otherwise —
            # unstarted work counts against the phase instead of being
            # ignored. Milestones drop out via zero duration weight.
            # Plain range arithmetic only (the N() Excel trap).
            c1, cn = row + 1, row + len(_kids)
            pf = ('=IF(SUMPRODUCT(($F{c1}:$F{cn}-$E{c1}:$E{cn}))=0,"",'
                  'SUMPRODUCT(($F{c1}:$F{cn}-$E{c1}:$E{cn})*'
                  '(($J{c1}:$J{cn})+($J{c1}:$J{cn}="")*'
                  '($D{c1}:$D{cn}="Complete")))/'
                  'SUMPRODUCT(($F{c1}:$F{cn}-$E{c1}:$E{cn})))'
                  ).replace("{c1}", str(c1)).replace("{cn}", str(cn))
            ppc = ws.cell(row, 10, pf)
            ppc.number_format = "0%"
            ppc.font = Font(bold=True, size=8)
            ppc.alignment = Alignment(horizontal="center")
        ws.row_dimensions[row].outline_level = 1
        ws.row_dimensions[row].height = 12
        bar_rules(row, _C["summary"], pct_overlay=True)
        dv.add(f"D{row}")
        row += 1
        for t in children.get(phase["id"], []):
            render_task(t)

    for t in _orphans:
        render_task(t)

    # ---- Total float and the critical path (detailed style) -------------
    # CPM backward pass, emitted as LIVE formulas rather than numbers baked
    # in at generation. Static float next to live dates would repeat the
    # mistake working-day scheduling just fixed: the sheet contradicting
    # itself the moment someone edits. Two hidden helper columns carry the
    # late-start chain; the visible Float column is the difference.
    #
    #   project finish PF = MAX of every real task's End
    #   LS(t), no successors      = PF back by t's duration
    #   LS(t), with successors    = MIN over s of the latest start that
    #                               still lets s start on time, following
    #                               the same SS-drives-from-Start /
    #                               otherwise-from-End rule as the forward
    #                               pass, with the same extracted lags
    #   Float(t) = working days between t's Start and its late start
    #
    # Float 0 means the task is on the critical path: any slip moves the
    # project end. A cyclic graph skips the whole pass (the forward pass
    # already falls back to static dates there, and circular references
    # would be worse than a blank column).
    if detail_cols:
        LS_COL, PF_COL = last_col + 2, last_col + 3
        _cpm = [t for t in tasks
                if t.get("type") != "phase" and t["id"] in row_of]
        if _cpm and _deps_acyclic:
            _pf = f"${get_column_letter(PF_COL)}$1"
            ws.cell(1, PF_COL, "=MAX(" + ",".join(
                f"F{row_of[t['id']]}" for t in _cpm) + ")")
            _ls_of = lambda tid: f"${get_column_letter(LS_COL)}${row_of[tid]}"  # noqa: E731
            for t in _cpm:
                r = row_of[t["id"]]
                # Duration reads LIVE off the row's own dates rather than
                # baking in the extracted length. It matters in exactly the
                # case people actually hit: overwrite a task's End by hand
                # and a constant duration would leave its float unchanged
                # while its slack has visibly been spent.
                dur = f"(NETWORKDAYS(E{r},F{r})-1)"
                back = lambda expr: f"WORKDAY({expr},-{dur})"  # noqa: E731
                succs = [by_id[sid] for sid in _dependents.get(t["id"], [])
                         if by_id.get(sid) is not None
                         and by_id[sid].get("type") != "phase"
                         and sid in row_of]
                if not succs:
                    expr = back(_pf)
                else:
                    bounds = []
                    for sc in succs:
                        ls_s = _ls_of(sc["id"])
                        if _dep_code(t, sc) == "SS":
                            # a start-to-start successor constrains this
                            # task's start directly, so no duration term
                            lag = _workday_offset(_d(t["start"]), _d(sc["start"]))
                            bounds.append(_workday_formula(ls_s, -lag))
                        else:
                            lag = _workday_offset(_d(t["end"]), _d(sc["start"]))
                            bounds.append(back(_workday_formula(ls_s, -lag)))
                    expr = (bounds[0] if len(bounds) == 1
                            else "MIN(" + ",".join(bounds) + ")")
                ws.cell(r, LS_COL, "=" + expr)
                fc = ws.cell(r, 12,
                             f"=NETWORKDAYS(E{r},{_ls_of(t['id'])})-1")
                fc.number_format = "0"
                fc.font = Font(size=8, color="FF3D4653")
                fc.alignment = Alignment(horizontal="center")
            # zero float reads as the critical path, live
            ws.conditional_formatting.add(
                f"{FLOAT_COL}{first_bar_row}:{FLOAT_COL}{last_row}",
                CellIsRule(operator="lessThanOrEqual", formula=["0"],
                           font=Font(bold=True, size=8,
                                     color="FF" + _C["risk_done"]),
                           fill=dxf_fill(_C["risk_rest"])))
        for _hc in (LS_COL, PF_COL):
            ws.column_dimensions[get_column_letter(_hc)].hidden = True

    # grid-wide dynamics AFTER the bar rules so bars win: the today
    # column tracks TODAY(); weekends shade by formula
    ax0 = f"{get_column_letter(FIRST_DAY_COL)}$1"
    full_grid = (f"{get_column_letter(FIRST_DAY_COL)}{first_bar_row}:"
                 f"{get_column_letter(last_col)}{last_row}")
    ws.conditional_formatting.add(full_grid, FormulaRule(
        formula=[f"{ax0}=TODAY()"], fill=dxf_fill(_C["today"]), stopIfTrue=True))
    ws.conditional_formatting.add(full_grid, FormulaRule(
        formula=[f"WEEKDAY({ax0},2)>5"], fill=dxf_fill(_C["weekend"])))
    # live status dots: flipping the Status dropdown recolors column A.
    # Guarded: a plan whose tasks all miss their phase (model returns a flat
    # list, or a parent_id that matches nothing) writes the project row and
    # stops, and the range below inverts. openpyxl rejects it with a bare
    # TypeError, which reaches the user as a failed build rather than a
    # thin one.
    if last_row > first_bar_row:
        dot_range = f"A{first_bar_row + 1}:A{last_row}"
        for key, label in _STATUS_LABELS.items():
            ws.conditional_formatting.add(dot_range, FormulaRule(
                formula=[f'$D{first_bar_row + 1}="{label}"'],
                font=Font(color="FF" + _C["status"].get(key, _C["bar"]))))

    ws.freeze_panes = f"{get_column_letter(FIRST_DAY_COL)}4"
    ws.sheet_properties.outlinePr.summaryBelow = False
    # Print setup (PM audit 2026-07-23: print-to-PDF for a steering
    # meeting must not sprawl across six portrait pages). Landscape,
    # squeeze to one page wide, header rows repeat on every page.
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.print_area = (f"A2:{get_column_letter(last_col + 1)}"
                     f"{last_row + 2}")
    ws.print_title_rows = "2:3"
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    # Milestone rows compute "◆" between bar rows computing "█", and the
    # Start/End columns now mix static dates with dependency formulas:
    # both trip Excel's inconsistent-formula checker (Scott 2026-07-21,
    # row-12 screenshot). Declare the day grid AND the date columns
    # ignored for that check, the same mark Excel writes when a user
    # clicks "Ignore Error". openpyxl 3.1 has no serializer for it, so
    # _normalize_zip injects the element from this stash.
    _ind_L = get_column_letter(last_col + 1)
    wb._gp_ignored_errors = {
        ws.title: (f"{full_grid} E{first_bar_row}:F{last_row}"
                   f" {_ind_L}{first_bar_row}:{_ind_L}{last_row}"
                   f" J{first_bar_row}:J{last_row}")}
    return wb, {"row_of": row_of, "first_bar_row": first_bar_row,
                "last_row": last_row, "sheet": ws.title}


def _serialize_wb(wb) -> bytes:
    # Determinism is a CLAIMED property (same-plan-same-bytes, asserted by
    # the acceptance test and relied on for artifact byte-stability), but
    # openpyxl stamps wall-clock time in two places: docProps/core.xml
    # created/modified, and every zip member's DOS mtime (2s resolution).
    # Renders straddling a second boundary produced different bytes —
    # a latent CI flake that struck twice on 2026-07-19. Freeze both.
    from datetime import datetime as _dt
    wb.properties.created = _dt(2026, 1, 1)
    wb.properties.modified = _dt(2026, 1, 1)
    # sheet title -> zip member name, for the ignored-errors injection
    ignored = {
        f"xl/worksheets/sheet{i + 1}.xml": sqref
        for i, title in enumerate(wb.sheetnames)
        for t, sqref in (getattr(wb, "_gp_ignored_errors", {}) or {}).items()
        if t == title
    }
    buf = BytesIO()
    wb.save(buf)
    return _normalize_zip(buf.getvalue(), ignored_errors=ignored)


def _slip_key(name: str) -> str:
    """Task identity across extraction runs: normalized name. Extraction
    won't spell a task identically every time ("Payments integration" vs
    "Payments Integration"), so match case/space/punctuation-blind and
    let anything that still misses fall out as "first tracked" rather
    than fabricate a lineage."""
    return re.sub(r"[^a-z0-9]+", " ", (name or "").lower()).strip()


def _solid(argb: str):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=argb)


def _norm_quote(s: str) -> str:
    """Quote identity for matching a receipt against a commitment."""
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


_FIELD_LABELS = {
    "end": "Due date", "end_date": "Due date", "due": "Due date",
    "start": "Start date", "start_date": "Start date",
    "owner": "Owner", "status": "Status", "effort": "Effort estimate",
    "percent_complete": "Progress", "percent": "Progress",
    "depends_on": "Dependency", "name": "Task", "conditional": "Condition",
}

_DATE_FIELDS = {"end", "end_date", "due", "start", "start_date"}


def _field_label(field: str) -> str:
    """Receipts is a customer-facing sheet: our schema keys are not.
    Anything unmapped falls back to the raw key made readable rather than
    dropped, so a new schema field surfaces as odd wording, never as a
    blank cell."""
    key = str(field or "").strip()
    return _FIELD_LABELS.get(key.lower(), key.replace("_", " ").capitalize())


def _commitments(task: dict) -> list[dict]:
    """The due dates somebody actually SPOKE for this task, oldest first.

    The plan-version trail can only see what changed between generations;
    this sees what changed between meetings, which is the thing a project
    manager is actually asking about. It also survives what the version
    trail cannot: a rename by the extraction model, a project regenerated
    out of order, and a FIRST generation, where there are no prior
    versions at all and slip used to be structurally empty.

    Entries the model could not date sort after the ones it could, in the
    order given, since the schema asks for them oldest first."""
    out = []
    for c in (task.get("commitments") or []):
        if not isinstance(c, dict):
            continue
        try:
            when = _d(c["date"])
        except (KeyError, ValueError, TypeError):
            continue
        as_of = None
        if c.get("as_of"):
            try:
                as_of = _d(c["as_of"])
            except (ValueError, TypeError):
                as_of = None
        out.append({"date": when, "as_of": as_of,
                    "quote": str(c.get("quote") or "").strip(),
                    "speaker": str(c.get("speaker") or "").strip(),
                    "reason": str(c.get("reason") or "").strip()})
    dated = [c for c in out if c["as_of"]]
    undated = [c for c in out if not c["as_of"]]
    dated.sort(key=lambda c: c["as_of"])
    return dated + undated


def _wd_inclusive(a: date, b: date) -> int:
    """Working days from a to b counting BOTH ends, matching Excel's
    NETWORKDAYS. The S-curve mixes Python-computed history with live
    NETWORKDAYS formulas in the same chart, so the two have to count the
    same way or baseline and planned would be weighted differently."""
    if b < a:
        return 0
    return _workday_offset(a, b) + (1 if a.weekday() < 5 else 0)


def _earned_days(tasks: list[dict]) -> float:
    """Working days of work banked, in ABSOLUTE days, not a share.

    The S-curve divides this by ONE denominator (the current plan's total),
    which is what keeps Reported monotonic. Dividing each version by its own
    scope is what made progress appear to go backwards: the 2026-07-13
    standup stretched payments and added a task with no stated percent, so
    the same banked work over a bigger plan read as 18.5% falling to 16.5%.
    A cumulative progress line that drops when scope GROWS is the classic
    way to make a real plan look like a failing one."""
    num = 0.0
    for t in tasks:
        if t.get("type") in ("phase", "milestone"):
            continue
        try:
            dur = _wd_inclusive(_d(t["start"]), _d(t["end"]))
        except (KeyError, ValueError, TypeError):
            continue
        if dur <= 0:
            continue
        pct = t.get("percent_complete")
        if not isinstance(pct, int):
            pct = 100 if t.get("status") == "complete" else 0
        num += dur * min(max(pct, 0), 100) / 100.0
    return num


def _total_days(tasks: list[dict]) -> float:
    """Scheduled working days across real tasks: the S-curve's denominator."""
    tot = 0.0
    for t in tasks:
        if t.get("type") in ("phase", "milestone"):
            continue
        try:
            dur = _wd_inclusive(_d(t["start"]), _d(t["end"]))
        except (KeyError, ValueError, TypeError):
            continue
        if dur > 0:
            tot += dur
    return tot


def _weighted_progress(tasks: list[dict]) -> float | None:
    """Duration-weighted percent complete over real tasks, using the SAME
    rule the phase rollups use on the view: the stated percent when
    somebody said one, 100 when the status is Complete, 0 otherwise, so
    unstarted work counts against the total instead of being ignored.
    Milestones drop out by carrying zero duration weight."""
    num = den = 0.0
    for t in tasks:
        # Milestones are moments, not work; phases are rollups of the rows
        # below them. Counting either would double-weight the curve.
        if t.get("type") in ("phase", "milestone"):
            continue
        try:
            dur = _wd_inclusive(_d(t["start"]), _d(t["end"]))
        except (KeyError, ValueError, TypeError):
            continue
        if dur <= 0:
            continue
        pct = t.get("percent_complete")
        if not isinstance(pct, int):
            pct = 100 if t.get("status") == "complete" else 0
        num += dur * min(max(pct, 0), 100) / 100.0
        den += dur
    return (num / den) if den else None


def _planned_days(tasks: list[dict], upto: date) -> float:
    """Scheduled working days falling on or before `upto`, in ABSOLUTE
    days, not a share.

    Absolute for the same reason `_earned_days` is, and it returns days
    rather than a share so that no caller can accidentally normalise a
    plan version against its own scope. Every series on this chart
    divides by ONE denominator, the plan as it stands today.

    This function used to return `done / total` over the tasks it was
    handed, which meant the baseline was a share of the FIRST plan's
    scope while Current plan and Reported were shares of today's. Two
    percentages of different wholes plotted on one axis render scope
    GROWTH as if it were slippage: with a first plan of 13 working days
    against today's 100, both plans have banked the identical 3 days by a
    given week and read as 23% against 3%. The baseline then sits under
    the current plan for the whole chart and looks like a project running
    late (Scott, 2026-08-13). Same failure `_earned_days` documents for
    Reported, which was fixed there and missed here."""
    done = 0.0
    for t in tasks:
        if t.get("type") in ("phase", "milestone"):
            continue
        try:
            s0, e0 = _d(t["start"]), _d(t["end"])
        except (KeyError, ValueError, TypeError):
            continue
        dur = _wd_inclusive(s0, e0)
        if dur <= 0:
            continue
        if upto >= s0:
            done += min(_wd_inclusive(s0, min(e0, upto)), dur)
    return done


def _scurve_weeks(tasks: list[dict], history: list[dict] | None) -> list[date]:
    """Week-ending Fridays spanning every plan version we know about."""
    dates = []
    for src in [{"tasks": tasks}] + list(history or []):
        for t in (src.get("tasks") or []):
            for k in ("start", "end"):
                try:
                    dates.append(_d(t[k]))
                except (KeyError, ValueError, TypeError):
                    pass
    if not dates:
        return []
    lo, hi = min(dates), max(dates)
    cur = lo + timedelta(days=(4 - lo.weekday()) % 7)   # first Friday on/after
    out = []
    while cur <= hi + timedelta(days=7) and len(out) < 80:
        out.append(cur)
        cur += timedelta(days=7)
    return out


def _compute_slip(tasks: list[dict], history: list[dict],
                  *, as_of: str | None = None) -> list[dict]:
    """Per-task due-date movement, from what was SAID and what was PLANNED.

    Two sources, deliberately merged into one trail:

      spoken   the commitments the extraction pulled out of the meetings
               ("by July 10", "pushed to the 17th", "realistically the
               24th"). Available on the very first generation, and immune
               to the extraction renaming a task between runs.
      planned  the same task's end date in each earlier plan VERSION,
               matched by normalized name. Catches movement nobody said
               out loud, and covers projects whose history predates the
               commitments field.

    Entries are keyed by the meeting they belong to, so a date that was
    both spoken and captured in that meeting's version counts ONCE. Where
    the two disagree for the same meeting the spoken one wins: somebody
    said it, we inferred the other. Consecutive equal dates collapse, so
    restating a date unchanged is not a move.

    baseline = the earliest date we can attribute to this task; moves =
    number of changes after it; the caller computes variance as current
    minus baseline. A task with exactly one entry is "first tracked",
    honest, not padded.

    `as_of` is THIS plan's meeting date, and versions at or after it are
    dropped: they are earlier BUILDS of the same meeting, not earlier
    meetings, and this build supersedes them. Without that, regenerating
    a plan reads its own previous attempt as history and every wobble in
    extraction becomes a move the meeting never made. Sibling of the
    one-version-per-as-of rule in plan_snapshots.history, which dedupes
    rebuilds inside the history but could not see the current one."""
    cur_as_of = None
    if as_of:
        try:
            cur_as_of = _d(as_of)
        except (ValueError, TypeError):
            cur_as_of = None
    versions = []
    for ver in history or []:
        try:
            v_as_of = _d(ver["as_of"])
        except (KeyError, ValueError, TypeError):
            continue
        if cur_as_of and v_as_of >= cur_as_of:
            continue
        versions.append((v_as_of, ver))
    out = []
    for t in tasks:
        if t.get("type") == "phase":
            continue
        key = _slip_key(t.get("name"))
        # keyed by as-of date so the two sources dedupe against each other;
        # undated spoken entries keep their given order at the front
        seq: list[dict] = []
        by_as_of: dict[date, dict] = {}
        for c in _commitments(t):
            entry = {"as_of": c["as_of"], "date": c["date"], "spoken": True,
                     "quote": c["quote"], "reason": c["reason"]}
            if c["as_of"] is None:
                seq.append(entry)
            else:
                by_as_of[c["as_of"]] = entry
        for v_as_of, ver in versions:
            for ht in ver.get("tasks") or []:
                if ht.get("type") != "phase" and _slip_key(ht.get("name")) == key:
                    try:
                        # Snapshots predate working-day snapping, so put the
                        # historical end through the same rule; otherwise the
                        # first regeneration after that change reads a
                        # weekend-to-Friday snap as real slip.
                        when = _snap_span(_d(ht.get("start") or ht["end"]),
                                          _d(ht["end"]))[1]
                    except (KeyError, ValueError, TypeError):
                        break
                    as_of = v_as_of
                    # spoken wins for the same meeting; see docstring
                    by_as_of.setdefault(as_of, {
                        "as_of": as_of, "date": when, "spoken": False,
                        "quote": "", "reason": ""})
                    break
        seq += [by_as_of[k] for k in sorted(by_as_of)]
        cur_end = _d(t["end"])
        if not seq or seq[-1]["date"] != cur_end:
            seq.append({"as_of": None, "date": cur_end, "spoken": False,
                        "quote": "", "reason": "", "current": True})
        changes = [seq[0]]
        for item in seq[1:]:
            if item["date"] != changes[-1]["date"]:
                changes.append(item)
        out.append({
            "task": t, "baseline": changes[0]["date"],
            "baseline_as_of": changes[0]["as_of"],
            "baseline_spoken": changes[0].get("spoken", False),
            "current": cur_end, "moves": len(changes) - 1,
            "trail": changes, "first_tracked": len(seq) == 1,
            "spoken_any": any(c.get("spoken") for c in changes),
        })
    return out


def _build_scurve_sheet(wb, data: dict, history: list[dict] | None,
                        layout: dict, today: date) -> None:
    """Progress sheet: reported progress from meetings against the plan.

    Retitled and reframed after a professional PM reviewed a live export
    (2026-08-11, validated against earned-value convention): this is a
    meeting-derived progress curve, NOT a time-phased planned-value
    S-curve, and the labeling must not overpromise. The series:

      First plan (baseline)  where the FIRST plan version said the work
                would be by each week. Static: it is what was promised,
                and editing today's dates must not rewrite it. Absent
                until a project has a second meeting.
      Current plan  where the CURRENT plan puts it. Live formulas over
                the Gantt View's own date cells, so it re-draws when you
                push a task, exactly like the bars do. Solid up to the
                data date, dashed past it, because right of the data date
                nothing has reported and the line is plan only.
      Reported  duration-weighted percent complete as of each meeting,
                using the same rule as the phase rollups: the stated
                percent when somebody said one, 100 when Complete, 0
                otherwise. Held flat between meetings because that is
                genuinely all we know, and STOPPED at the data date (the
                last meeting that reported progress): carrying it further
                would render actuals for weeks that have not reported.

    A vertical marker stands on the data date, and the sheet carries a
    date-coverage line ("dates on N of M activities") so undated work is
    visible instead of silently missing. Everything is weighted by
    working days, counted inclusively so the Python-computed history and
    the live NETWORKDAYS formulas agree.
    """
    from openpyxl.chart import LineChart, Reference
    import copy

    from openpyxl.chart.error_bar import ErrorBars
    from openpyxl.chart.legend import LegendEntry
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.text import RichText
    from openpyxl.drawing.line import LineProperties
    from openpyxl.drawing.text import (
        CharacterProperties,
        Paragraph,
        ParagraphProperties,
    )
    from openpyxl.styles import Alignment, Font

    tasks = [t for t in (data.get("tasks") or [])
             if t.get("type") not in ("phase", "milestone")]
    weeks = _scurve_weeks(data.get("tasks") or [], history)
    if not tasks or not weeks:
        return

    row_of = layout["row_of"]
    gsheet = layout["sheet"]
    rows = [row_of[t["id"]] for t in tasks if t.get("id") in row_of]
    if not rows:
        return

    # "Progress", not "S-Curve": the tab name is a claim too, and this
    # sheet is not an EVM S-curve (PM review 2026-08-11)
    sc = wb.create_sheet("Progress")
    navy = "1F3A5F"
    sc["A1"] = "Reported progress (from meetings)"
    sc["A1"].font = Font(bold=True, size=13, color="FF" + navy)
    # What the weighting is actually based on, said out loud. Every series
    # here weighs a task by its scheduled working days, which is a proxy for
    # size, not a measure of it: ten days waiting on a vendor outweighs three
    # days of hard work. Stating the count of tasks that gave us a real
    # effort figure lets the reader judge how rough the proxy is.
    _real = [t for t in tasks if t.get("type") not in ("phase", "milestone")]
    _stated = sum(1 for t in _real
                  if isinstance(t.get("effort"), str) and t["effort"].strip())
    # Rows 2-4 are merged wrapped notes: give each row enough height for
    # its whole text and anchor the text to the TOP. A merged cell never
    # auto-grows its row, and the default bottom anchor made an undersized
    # row show only its LAST lines, so the printed sheet opened the note
    # mid sentence (PM review round two, 2026-08-12).
    sc["A2"] = ("Baseline is the first plan version and never moves. "
                "Every line here is a share of the plan as it stands "
                "today, the baseline included, so a first plan that "
                "covered less work than today's tops out below 100% "
                "instead of finishing at it. "
                "Current plan follows the dates on the Gantt View and "
                "redraws when you edit them. Reported is % complete as of "
                "each meeting, held flat between meetings because that is "
                "all the meetings said, and it stops at the data date. "
                "The vertical line marks the last meeting that reported "
                "progress. Right of that line nothing has reported yet, so "
                "the chart shows plan only, drawn dashed. The dots are the "
                "meetings themselves. This is a meeting-derived progress "
                "curve measured against the plan as it stands today, not "
                "an earned-value S-curve, so adding scope raises what is "
                "left rather than pushing the line back down.")
    sc["A2"].font = Font(size=9, color="FF666666", italic=True)
    sc["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    sc.merge_cells("A2:G2")
    sc.row_dimensions[2].height = 80
    sc["A3"] = (
        f"Weighting: every task counts for its scheduled working days, "
        f"which stands in for size. {_stated} of {len(_real)} tasks had an "
        f"effort figure stated in a meeting; none of the numbers here are "
        f"weighted by it.")
    sc["A3"].font = Font(size=8, color="FF9AA4AF", italic=True)
    sc["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    sc.merge_cells("A3:G3")
    sc.row_dimensions[3].height = 26
    # date coverage, computed from the source activities and rendered on
    # the export itself so an undated-activity gap is visible, never
    # hidden (PM review 2026-08-11)
    _dated_n, _total_n = (data.get("_date_coverage")
                          or (len(_real), len(_real)))
    _cov_line = f"Date coverage: dates on {_dated_n} of {_total_n} activities."
    if _dated_n < _total_n:
        _cov_line += (" Activities the meetings never dated are not on the "
                      "timeline and not in these curves.")
    sc["A4"] = _cov_line
    sc["A4"].font = Font(size=8, italic=True, bold=_dated_n < _total_n,
                         color="FF9A3412" if _dated_n < _total_n
                         else "FF9AA4AF")
    sc["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    sc.merge_cells("A4:G4")
    sc.row_dimensions[4].height = 24

    # total scheduled working days, live, so Current plan is a real share
    denom = "+".join(f"NETWORKDAYS('{gsheet}'!$E{r},'{gsheet}'!$F{r})"
                     for r in rows)
    sc["J1"] = "=" + denom
    sc.column_dimensions["J"].hidden = True

    hdr = ["Week ending", "First plan (baseline)", "Current plan",
           "Plan after data date (planned only)", "Reported (from meetings)",
           "At a meeting", "Data date"]
    for c, h in enumerate(hdr, start=1):
        cell = sc.cell(5, c, h)
        cell.font = Font(bold=True, size=9, color="FFFFFFFF")
        cell.fill = _solid("FF" + navy)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    sc.row_dimensions[5].height = 34
    sc.column_dimensions["A"].width = 14
    for col in "BCDEFG":
        sc.column_dimensions[col].width = 12

    # Reported: one observation per plan version, then today's plan. Held
    # flat between meetings, blank before the first meeting we have, and
    # stopped at the data date.
    # ONE denominator for every point: the plan as it stands today. Earned
    # work is absolute, so adding scope raises what is left to do without
    # ever un-banking what was already done.
    _denom = _total_days(data.get("tasks") or [])
    stamps: list[tuple[date, float]] = []
    for ver in (history or []):
        try:
            stamps.append((_d(ver["as_of"]),
                           _earned_days(ver.get("tasks") or []) / _denom))
        except (KeyError, ValueError, TypeError, ZeroDivisionError):
            continue
    if _denom:
        stamps.append((_d(data.get("meeting_date")) if data.get("meeting_date")
                       else today,
                       _earned_days(data.get("tasks") or []) / _denom))
    stamps = sorted((d, v) for d, v in stamps if v is not None)

    # The data date: the last meeting that reported progress. Reported and
    # its flat carry stop here, because a value in a later week would be
    # an actual for a week no meeting has reported (PM review 2026-08-11,
    # earned-value convention). Right of it the chart is plan only.
    data_date = stamps[-1][0] if stamps else None
    dw_idx = None
    if data_date is not None:
        dw_idx = next((i for i, wk in enumerate(weeks) if wk >= data_date),
                      len(weeks) - 1)

    base_tasks = (history or [{}])[0].get("tasks") if history else None

    for i, wk in enumerate(weeks):
        r = 6 + i
        # Column A is the TEXT label the chart plots. A real date here makes
        # Excel build a date axis and label every day between the first and
        # last point: 44 ticks for 7 values, which crushed the labels and
        # pushed the axis title on top of them (2026-07-31). The live date
        # the Planned formula needs moves to the hidden helper column, which
        # formulas read happily. It cannot go the other way round: a chart
        # will not plot categories out of a hidden column.
        dc = sc.cell(r, 1, wk.strftime("%b %-d"))
        dc.font = Font(size=8)
        dc.alignment = Alignment(horizontal="center")
        hd = sc.cell(r, 10, wk)
        hd.number_format = "yyyy-mm-dd"
        if base_tasks and _denom:
            # Over TODAY's denominator, never the first plan's own. See
            # _planned_days: dividing each version by its own scope drew
            # scope growth as slippage.
            sc.cell(r, 2, _planned_days(base_tasks, wk) / _denom
                    ).number_format = "0%"
        terms = "+".join(
            f"MAX(0,NETWORKDAYS('{gsheet}'!$E{rr},"
            f"MIN('{gsheet}'!$F{rr},$J{r})))" for rr in rows)
        pf = f"=IFERROR(({terms})/$J$1,\"\")"
        # The current plan is ONE line with two rendering identities: solid
        # up to the data date, dashed past it, split across two columns so
        # the chart can style them apart. Both stay live formulas, and they
        # share the data date week so the line stays connected.
        if dw_idx is None or i <= dw_idx:
            sc.cell(r, 3, pf).number_format = "0%"
        if dw_idx is not None and dw_idx <= i and dw_idx < len(weeks) - 1:
            sc.cell(r, 4, pf).number_format = "0%"
        # Reported stops at the data date. The flat carry BETWEEN meetings
        # before it stays: that is honest, it is all the meetings said.
        reported = [v for d, v in stamps if d <= wk]
        if reported and dw_idx is not None and i <= dw_idx:
            sc.cell(r, 5, reported[-1]).number_format = "0%"
        # The same value again, but ONLY in a week that actually contains a
        # meeting. Plotted as markers with no line, it separates the
        # things we were told from the flat carry between them, which the
        # line alone cannot say.
        observed = [v for d, v in stamps if wk - timedelta(days=6) <= d <= wk]
        if observed:
            sc.cell(r, 6, observed[-1]).number_format = "0%"
        # anchor for the vertical time-now marker: the series' y error bar
        # drops from 100% to the axis at exactly this category
        if dw_idx is not None and i == dw_idx:
            sc.cell(r, 7, 1.0).number_format = "0%"

    last = 5 + len(weeks)
    chart = LineChart()
    # No chart title and no axis titles: A1 already names the sheet, the
    # header row already says "Week ending", and a percent axis needs no
    # label. Every one of them was spending space the plot wanted, and the
    # x-axis title was landing on top of the tick labels.
    chart.style = 12
    chart.height, chart.width = 9, 20
    chart.y_axis.numFmt = "0%"
    # Progress cannot exceed 100%, and Excel's autoscale was running the
    # axis to 120% and spending a fifth of the plot height on impossible
    # values (Scott's screenshot 2026-07-31).
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 1
    # Without this the category axis renders the date serials (46199,
    # 46206) instead of dates: openpyxl writes the categories as a numeric
    # reference and Excel has no format to apply unless the axis carries
    # one. Caught in the shipped 2026-07-30 workbook.
    chart.x_axis.numFmt = "mmm d"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.add_data(Reference(sc, min_col=2, max_col=7, min_row=5, max_row=last),
                   titles_from_data=True)
    chart.set_categories(Reference(sc, min_col=1, min_row=6, max_row=last))
    # Legend placement, chosen from the DATA rather than fixed (Scott
    # 2026-08-01, after the lines ran through it). An S-curve climbs left to
    # right, so the free corner depends on how the project is doing:
    #   bottom-right  the usual case, curves finish high and leave it empty
    #   top-left      a struggling project keeps the curves low on the right,
    #                 which fills the bottom-right and empties the top-left
    #   right rail    neither corner is clear, so do not sit on the plot
    # Baseline and Planned are computed here the same way the sheet computes
    # them, so this reads the real shape rather than assuming one.
    _left_third = weeks[:max(1, len(weeks) // 3)]
    _right = weeks[-1]
    _data_wk = weeks[dw_idx] if dw_idx is not None else None
    _series_at = lambda wk: [v for v in (
        (_planned_days(base_tasks, wk) / _denom)
        if (base_tasks and _denom) else None,
        (_planned_days(data.get("tasks") or [], wk) / _denom)
        if _denom else None,
        # Reported no longer extends past the data date, so it is not ink
        # to dodge out there
        (next((v for d, v in reversed(stamps) if d <= wk), None)
         if _data_wk is not None and wk <= _data_wk else None),
    ) if v is not None]
    _right_min = min(_series_at(_right), default=0.0)
    _left_max = max((max(_series_at(wk), default=0.0) for wk in _left_third),
                    default=1.0)
    # The vertical data date marker is ink too: it runs the full plot
    # height, so any corner it passes through is not free, whatever the
    # curves are doing.
    _dd_pos = ((dw_idx + 0.5) / len(weeks)) if dw_idx is not None else None
    _clear = lambda x0, x1: (_dd_pos is None  # noqa: E731
                             or not (x0 - 0.02 <= _dd_pos <= x1 + 0.02))
    if _right_min >= 0.45 and _clear(0.70, 0.98):
        _corner = (0.70, 0.52, 0.28, 0.30)      # inside, bottom right
    elif _left_max <= 0.55 and _clear(0.14, 0.42):
        _corner = (0.14, 0.08, 0.28, 0.30)      # inside, top left
    else:
        _corner = None                          # nowhere safe: outside right
    chart.legend.position = "r"
    if _corner:
        from openpyxl.chart.layout import Layout, ManualLayout
        _x, _y, _w, _h = _corner
        chart.legend.overlay = True
        chart.legend.layout = Layout(manualLayout=ManualLayout(
            xMode="edge", yMode="edge", x=_x, y=_y, w=_w, h=_h))
    # Both inside placements sit the legend ON the plot, and a legend with
    # no fill is transparent, so the curves ran straight through the text
    # (Scott, 2026-08-13: "I had to move it so I could actually see the
    # lines"). Corner dodging picks the emptiest corner, it cannot promise
    # an empty one. A white panel with a light border makes the legend
    # opaque furniture instead of one more layer of ink.
    _legend_fill = GraphicalProperties(solidFill="FFFFFF")
    _legend_fill.ln = LineProperties(solidFill="D9D9D9")
    chart.legend.spPr = _legend_fill
    # Six series are plotted, four of them real curves. "At a meeting" and
    # "Data date" are furniture: the meeting dots and the vertical marker.
    # Listing them spent legend rows that the box did not have, and the
    # entries that fell off the end were the real ones, which is how
    # Reported (the green line) ended up plotted with nothing naming it.
    # Drop the two helpers so every line a reader can see is named.
    _helpers = [i for i, _h in enumerate(hdr[1:])
                if _h in ("At a meeting", "Data date")]
    if _helpers:
        chart.legend.legendEntry = [LegendEntry(idx=i, delete=True)
                                    for i in _helpers]
    # Chart text does not inherit the cell font, and the 10pt default was
    # small against a 20-wide plot.
    _axis_text = RichText(p=[Paragraph(pPr=ParagraphProperties(
        defRPr=CharacterProperties(sz=1000)), endParaRPr=CharacterProperties(sz=1000))])
    chart.x_axis.txPr = _axis_text
    chart.y_axis.txPr = copy.deepcopy(_axis_text)
    # Three identities, three hues, assigned in fixed order and validated for
    # colourblind separation rather than eyeballed. Baseline is the promise
    # and is dashed because it never moves; Current plan follows the sheet,
    # in one hue across both of its columns; Reported is what actually
    # happened, and it ends at the data date.
    for ser, hexc in zip(chart.series,
                         ("B5651D", "1F4E9C", "1F4E9C", "2E9E4F")):
        ser.smooth = False
        ser.graphicalProperties.line.solidFill = hexc
        ser.graphicalProperties.line.width = 22000     # ~1.7pt
    chart.series[0].graphicalProperties.line.dashStyle = "dash"
    # Right of the data date the current plan is planned work only, so it
    # renders dashed in the same hue: same identity, no claim of actuals.
    chart.series[2].graphicalProperties.line.dashStyle = "dash"
    # The meeting observations are the SAME entity as Reported, so they keep
    # its colour and carry identity by shape instead: dots, no line.
    obs = chart.series[4]
    obs.marker = Marker(symbol="circle", size=7)
    obs.graphicalProperties.line.noFill = True
    obs.marker.graphicalProperties.solidFill = "2E9E4F"
    obs.marker.graphicalProperties.line.solidFill = "2E9E4F"
    # The time-now marker: a one point series at 100% over the data date
    # week whose y error bar drops the full axis height, which draws the
    # vertical line a category axis line chart cannot otherwise draw.
    dd_ser = chart.series[5]
    dd_ser.smooth = False
    dd_ser.graphicalProperties.line.noFill = True
    dd_ser.marker = Marker(symbol="dash", size=7)
    dd_ser.marker.graphicalProperties.solidFill = "3D4653"
    dd_ser.marker.graphicalProperties.line.solidFill = "3D4653"
    dd_ser.errBars = ErrorBars(
        errDir="y", errBarType="minus", errValType="fixedVal", val=1,
        noEndCap=True,
        spPr=GraphicalProperties(ln=LineProperties(solidFill="3D4653",
                                                   w=12700)))
    sc.add_chart(chart, "I5")
    if not base_tasks:
        sc.cell(last + 2, 1,
                "Baseline appears once this project has a second plan "
                "version to compare against.").font = Font(
                    size=8, color="FF9AA4AF", italic=True)
    # Print setup, same rationale as the Gantt View's: a print-to-PDF of
    # this sheet must come out as ONE page carrying the title, the notes,
    # the coverage line, the table and the WHOLE chart. Without it the
    # default pagination sliced the chart across page boundaries, which is
    # how users saw the export (PM review round two, 2026-08-12). V/row 28
    # bound the chart's floating extent right of column I.
    from openpyxl.worksheet.properties import PageSetupProperties
    sc.print_area = f"A1:V{max(last + 2, 28)}"
    sc.page_setup.orientation = "landscape"
    sc.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sc.page_setup.fitToWidth = 1
    sc.page_setup.fitToHeight = 1


def render_gantt_detailed(data: dict, *, today: date | None = None,
                          history: list[dict] | None = None) -> bytes:
    """Detailed variant, LEAN (Scott 2026-07-21): the Gantt View plus
    exactly two sheets, Slip and Receipts. Everything task-level lives ON
    the timeline view: live dependency-driven dates, % Done and Effort
    (STRICTLY blank when nobody stated them; the extraction schema
    forbids estimating), the completed-portion bar overlay, overdue
    folded into the live risk rules, beyond-chart arrows. Phase % rolls
    up only when stated percents cover at least half the phase duration.
    Slip compares this plan against the project's prior snapshot
    versions (as-of ordered by meeting date). Receipts quotes the
    meeting line behind every extracted value. The retired Progress and
    Workload sheets are deliberate cuts, not omissions: Progress became
    redundant with the on-view columns, and active-per-week counted
    calendar spans dressed up as workload (the only-what-the-meeting-
    knows rule)."""
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.hyperlink import Hyperlink

    data = _split_undated(_with_workdays(_points_are_milestones(data)))
    wb, _layout = _build_gantt_wb(data, today=today, detail_cols=True)
    tasks = data.get("tasks") or []
    rows = [t for t in tasks if t.get("type") != "phase"]

    navy, amber_lt, red_lt, gray_lt = "1F3A5F", "FBF0D5", "F6D3CE", "F2F3F5"
    hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="FF" + navy)
    sub_font = Font(size=9, color="FF666666", italic=True)
    thin = Side(style="thin", color="FFD9D9D9")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Receipts numbering: plan order, commitments then evidence ----
    # The two extraction fields carry DIFFERENT quotes, which is not what I
    # assumed when Standing first shipped: the model files date lines under
    # commitments and files percent, status and effort lines under evidence,
    # so a date's provenance lives only in commitments. Receipts is the union
    # or it silently stops showing the line behind every due date.
    #
    # Standing comes from the DATE, not from matching quote strings: a
    # commitment still standing is one whose date is the task's current end.
    # Quote matching was the first attempt and it read empty on live data,
    # because the same utterance is rarely spelled identically twice.
    receipts: list[tuple[str, dict, dict]] = []   # (ref, task, receipt item)
    for t in rows:
        seen: set[str] = set()
        try:
            cur_end = _d(t["end"])
        except (KeyError, ValueError, TypeError):
            cur_end = None
        for c in _commitments(t):
            if not c["quote"]:
                continue
            seen.add(_norm_quote(c["quote"]))
            receipts.append((f"R{len(receipts) + 1}", t, {
                "field": "end",
                "quote": c["quote"],
                "speaker": c["speaker"],
                "meeting_date": c["as_of"].isoformat() if c["as_of"] else "",
                "standing": ("current" if cur_end and c["date"] == cur_end
                             else "superseded"),
            }))
        for ev in (t.get("evidence") or []):
            if not isinstance(ev, dict) or not str(ev.get("quote") or "").strip():
                continue
            if _norm_quote(ev.get("quote")) in seen:
                continue
            receipts.append((f"R{len(receipts) + 1}", t, ev))

    # ---- Slip sheet (v2: snapshot history) ----
    sl = wb.create_sheet("Slip")
    sl["A1"] = "Slip, how due dates moved across plan versions"
    sl["A1"].font = Font(bold=True, size=13, color="FF" + navy)
    sl["A2"] = ("First committed is the earliest due date anyone stated for "
                "the task, taken from the meetings themselves; where nobody "
                "stated one, it falls back to the earliest generated plan "
                "that tracked the task. Variance is current minus first "
                "committed: positive is late, negative is early. The trail "
                "shows each move, the meeting it came from, and the reason "
                "when somebody gave one.")
    sl["A2"].font = sub_font
    sl["A2"].alignment = Alignment(wrap_text=True)
    sl.merge_cells("A2:H2")
    sl.row_dimensions[2].height = 26
    slip_heads = ["Task", "Owner", "First committed", "As of", "Current due",
                  "Times moved", "Variance (days)", "Trail"]
    for ci, h in enumerate(slip_heads, 1):
        c = sl.cell(4, ci, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    slip_rows = _compute_slip(tasks, history,
                              as_of=data.get("meeting_date"))
    for ri, srow in enumerate(slip_rows, 5):
        t = srow["task"]
        sl.cell(ri, 1, t["name"]).font = Font(bold=True, size=9)
        sl.cell(ri, 2, (t.get("owner") or "").strip())
        bc = sl.cell(ri, 3, srow["baseline"])
        bc.number_format = "yyyy-mm-dd"
        _as_of = srow["baseline_as_of"]
        sl.cell(ri, 4, "first tracked now" if srow["first_tracked"]
                else (_as_of.strftime("%Y-%m-%d") if _as_of else "stated")
                ).font = Font(size=8, color="FF666666")
        cc = sl.cell(ri, 5, srow["current"])
        cc.number_format = "yyyy-mm-dd"
        mv = sl.cell(ri, 6, srow["moves"])
        mv.alignment = Alignment(horizontal="center")
        sc = sl.cell(ri, 7, f"=E{ri}-C{ri}")
        sc.number_format = "0"
        sc.alignment = Alignment(horizontal="center")
        parts = []
        for item in srow["trail"]:
            seg = item["date"].strftime("%b %d")
            if item.get("current"):
                seg += " (current)"
            elif item["as_of"]:
                seg += f" (said {item['as_of'].strftime('%b %d')})"
            if item.get("reason"):
                seg += f": {item['reason']}"
            parts.append(seg)
        trail = " → ".join(parts)
        sl.cell(ri, 8, trail if srow["moves"] else "").font = Font(size=8)
        if srow["moves"] >= 2:
            for ci in range(1, 9):
                sl.cell(ri, ci).fill = PatternFill("solid",
                                                   fgColor="FF" + amber_lt)
            mv.font = Font(bold=True, color="FF9A3412")
        for ci in range(1, 9):
            sl.cell(ri, ci).border = box
    if slip_rows:
        # positive variance reads red, live (critic pass 2026-07-21);
        # negative is a task that came IN and should not read like a
        # defect, so it gets its own green rather than sharing the red
        sl.conditional_formatting.add(
            f"G5:G{4 + len(slip_rows)}",
            CellIsRule(operator="greaterThan", formula=["0"],
                       font=Font(bold=True, color="FF9A1B12")))
        sl.conditional_formatting.add(
            f"G5:G{4 + len(slip_rows)}",
            CellIsRule(operator="lessThan", formula=["0"],
                       font=Font(bold=True, color="FF1E7A3C")))
    note_r = 5 + len(slip_rows) + 1
    if not any(s["spoken_any"] for s in slip_rows) and not (history or []):
        sl.cell(note_r, 1,
                "Nobody stated a due date more than once in these meetings, "
                "so there is nothing to compare yet: every future gantt for "
                "this project adds a dated version to compare "
                "against.").font = sub_font
    elif not (history or []):
        sl.cell(note_r, 1,
                "This is the first generated plan for this project, so the "
                "trail above is what people said in the meetings. Future "
                "gantts add plan versions to it.").font = sub_font
    for col, w in {"A": 30, "B": 14, "C": 12, "D": 15, "E": 12, "F": 11,
                   "G": 10, "H": 52}.items():
        sl.column_dimensions[col].width = w
    sl.freeze_panes = "A5"

    # ---- Receipts sheet ----
    rc = wb.create_sheet("Receipts")
    rc["A1"] = "Receipts, the meeting line behind every value"
    rc["A1"].font = Font(bold=True, size=13, color="FF" + navy)
    rc["A2"] = ("The plan is generated from the meetings, so provenance is "
                "automatic. Quotes are verbatim from the source content.")
    rc["A2"].font = sub_font
    rc["A3"] = ("Forwarding note: this sheet quotes people by name. "
                "Consider removing it before sending the workbook outside "
                "the team.")
    rc["A3"].font = Font(size=9, bold=True, color="FF9A3412")
    for ci, h in enumerate(["Ref", "Task", "Supports", "Meeting", "Standing",
                            "Speaker", "Verbatim line"], 1):
        c = rc.cell(4, ci, h)
        c.font = hdr_font
        c.fill = hdr_fill
    for ri, (ref, t, ev) in enumerate(receipts, 5):
        rc.cell(ri, 1, ref).font = Font(bold=True, size=9)
        _nc = rc.cell(ri, 2, t["name"])
        _nc.font = Font(size=9)
        # Click a receipt, land on the row it justifies. Cheap, and it turns
        # Receipts from an audit table into something you actually navigate.
        _r = _layout["row_of"].get(t.get("id"))
        if _r:
            # LOCATION, not target. Assigning a string to cell.hyperlink makes
            # openpyxl write an EXTERNAL relationship (TargetMode="External")
            # pointing at "#'Sheet'!B16", which Excel rejects outright: it
            # opened the 2026-07-31 workbook with "we found a problem with
            # some content". LibreOffice accepted it, which is exactly the
            # trap that made the N() bug survive too.
            _nc.hyperlink = Hyperlink(
                ref=_nc.coordinate, location=f"'{_layout['sheet']}'!B{_r}",
                display=str(t["name"]))
            _nc.font = Font(size=9, color="FF1F4E9C", underline="single")
        rc.cell(ri, 3, _field_label(ev.get("field"))).font = Font(size=9)
        md = str(ev.get("meeting_date") or "").strip()
        rc.cell(ri, 4, md).font = Font(size=8, color="FF666666")
        standing = str(ev.get("standing") or "")
        sc_cell = rc.cell(ri, 5, standing)
        sc_cell.font = Font(size=8, italic=True,
                            color="FF9A3412" if standing == "superseded"
                            else "FF666666")
        rc.cell(ri, 6, str(ev.get("speaker") or "")).font = Font(size=9)
        q = rc.cell(ri, 7, f'"{str(ev.get("quote")).strip()}"')
        q.font = Font(size=9)
        q.alignment = Alignment(wrap_text=True)
        for ci in range(1, 8):
            rc.cell(ri, ci).border = box
            if ri % 2 == 0:
                rc.cell(ri, ci).fill = PatternFill("solid",
                                                   fgColor="FF" + gray_lt)
    for col, w in {"A": 6, "B": 30, "C": 16, "D": 12, "E": 12, "F": 12,
                   "G": 64}.items():
        rc.column_dimensions[col].width = w
    rc.freeze_panes = "A5"

    _build_scurve_sheet(wb, data, history, _layout, today or date.today())

    return _serialize_wb(wb)


# The user's style word maps to a concrete registry entry at arm time
# (offers store the FAMILY template; see the template lane in chat.py).
STYLE_TO_TEMPLATE = {"simple": "gantt_smartsheet", "detailed": "gantt_detailed"}


def _normalize_zip(blob: bytes,
                   ignored_errors: dict[str, str] | None = None) -> bytes:
    """Re-pack the xlsx with fixed member timestamps (content, order, and
    compression preserved) so identical content is identical bytes.

    ignored_errors maps zip member name -> A1 range: injects the
    <ignoredErrors> element (the mark Excel writes on "Ignore Error") so
    the inconsistent-formula checker stays quiet over the day grid.

    POSITION MATTERS. CT_Worksheet is a sequence, and ignoredErrors comes
    BEFORE drawing / legacyDrawing / tableParts, not at the end. Appending
    it before </worksheet> was fine for as long as those elements never
    appeared. The conditional-date cell comment shipped 2026-07-31 put a
    <legacyDrawing> on the Gantt View, the injected element landed after it,
    and Excel refused the whole sheet: "we found a problem with some
    content", recovery log naming sheet1.xml. LibreOffice opened it without
    complaint, which is why recalc-green said nothing. So insert before the
    first element that must follow it, and fall back to </worksheet>."""
    import zipfile
    src = zipfile.ZipFile(BytesIO(blob))
    out = BytesIO()
    import re as _re
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as z:
        for name in src.namelist():
            data = src.read(name)
            sqref = (ignored_errors or {}).get(name)
            if sqref:
                blk = (b'<ignoredErrors><ignoredError sqref="'
                       + sqref.encode() + b'" formula="1"/></ignoredErrors>')
                # everything that must come AFTER ignoredErrors, in order
                anchor = min(
                    (i for i in (data.find(t) for t in
                                 (b"<drawing", b"<legacyDrawing", b"<oleObjects",
                                  b"<controls", b"<tableParts", b"<extLst"))
                     if i != -1),
                    default=data.find(b"</worksheet>"))
                data = data[:anchor] + blk + data[anchor:]
            if name == "docProps/core.xml":
                # openpyxl overwrites dcterms:modified with wall-clock at
                # save time (setting wb.properties beforehand is futile) —
                # pin both stamps here instead.
                for tag in (b"created", b"modified"):
                    # keep the element's own attributes (openpyxl declares
                    # xmlns:xsi element-scoped) — pin only the text value
                    data = _re.sub(
                        b"(<dcterms:" + tag + b"[^>]*>)[^<]*(</dcterms:"
                        + tag + b">)",
                        lambda m: m.group(1) + b"2026-01-01T00:00:00Z"
                        + m.group(2),
                        data)
            zi = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            zi.compress_type = zipfile.ZIP_DEFLATED
            z.writestr(zi, data)
    return out.getvalue()


TEMPLATES = {
    "gantt_smartsheet": {
        "hints": ("gantt", "project timeline", "project plan chart",
                  "timeline chart", "diagrama de gantt", "ガントチャート"),
        "extraction_prompt": _GANTT_SCHEMA_PROMPT,
        "renderer": render_gantt,
        "format": "xlsx",
        "media_type": XLSX_MIME,
        "filename": "Gantt.xlsx",
        "expected_seconds": 25,  # measured 2026-07-30, thinking off (was 45)
        "thinking": "disabled",  # see the detailed entry
        "offer_noun": "my polished Gantt chart (collapsible phases, status "
                      "colors, stated progress on the bars, critical "
                      "dates, a native Excel file)",
    },
    # Detailed variant (v1, 2026-07-21). hints EMPTY on purpose:
    # match_template never picks it directly — the family match is always
    # gantt_smartsheet, and the user's style choice (reply word or saved
    # per-project preference) swaps to this entry at arm time in chat.py.
    "gantt_detailed": {
        "hints": (),
        "extraction_prompt": _GANTT_DETAILED_SCHEMA_PROMPT,
        "renderer": render_gantt_detailed,
        "format": "xlsx",
        "media_type": XLSX_MIME,
        "filename": "Gantt_Detailed.xlsx",
        "expected_seconds": 30,  # measured 2026-07-30, thinking off (was 60)
        "max_tokens": 12000,     # evidence quotes fatten the JSON
        # Extraction is transcription with a schema, not deliberation, and
        # on models that think by default the reasoning shares max_tokens
        # with the answer. Measured on the demo standups, Sonnet 5, 2 runs
        # each: thinking on 87s and 108s, $0.11 to $0.12, and one of the two
        # burned all 12000 tokens thinking and returned NOTHING parseable;
        # thinking off 20s and 23s, $0.036 to $0.040, with byte-identical
        # extraction quality (same 9 tasks, same stated percents, same
        # commitment trail). Speed and cost are the small win here; the real
        # one is that a truncated turn is a failed generation for the user.
        "thinking": "disabled",
        "offer_noun": "my detailed Gantt workbook (the live timeline plus a "
                      "slip history of how due dates moved, stated effort, "
                      "and a receipts sheet quoting the meeting line "
                      "behind every value)",
    },
}


def artifact_filename(template: dict, plan: dict) -> str:
    """Distinctive artifact name: <Project>_<Base>_<MMDDYY>.<ext>.
    Five identical Project_Gantt.xlsx rows in the client's References
    made artifacts indistinguishable (Scott 2026-07-14) — the project
    slug comes from the extracted plan, and the stamp is the MEETING
    date the extraction read from the content (Scott's call: the
    artifact describes the meeting, not the build; also keeps the name
    deterministic per plan). Falls back to the UTC build date when the
    plan carries no parsable date. No project -> <Base>_<MMDDYY>.<ext>."""
    base, ext = template["filename"].rsplit(".", 1)
    slug = re.sub(r"[^A-Za-z0-9]+", "_",
                  str(plan.get("project") or "")).strip("_")[:40]
    try:
        stamp = datetime.strptime(
            str(plan.get("meeting_date")), "%Y-%m-%d").strftime("%m%d%y")
    except ValueError:
        stamp = datetime.now(timezone.utc).strftime("%m%d%y")
    return "_".join(p for p in (slug, base, stamp) if p) + "." + ext


def match_template(text: str, format: str | None = None) -> str | None:
    """Scan the WHOLE given text (capped), not a tail window: the first
    live miss was a 400-word library prompt that said "Gantt" once, in its
    opening sentence — a tail slice cut the keyword out of its own prompt.
    Callers pass the full assembled content DELIBERATELY, unlike the
    intent checks (#420): anaphoric asks ("make IT into an excel doc")
    carry the template keyword only in history, and that case is live-
    proven wanted (2026-07-13 16:52 offer).

    `format` is the classifier's read of the DESIRED output ("docx"...):
    a template that builds a different format is vetoed. This is what
    contains the history scan's false-positive class — live 2026-07-14
    21:58Z, a Word roles-doc ask drew the xlsx Gantt offer off 'gantt'
    in carried history; the veto blocks it while anaphora keeps working."""
    hay = (text or "")[-50000:].lower()
    for tid, t in TEMPLATES.items():
        if format and t["format"] != format:
            continue
        if any(h in hay and not _negated(hay, h) for h in t["hints"]):
            return tid
    return None


# Vocabulary that reads like the Gantt registry's territory without ever
# naming a match_template hint. Field case (Scott, 2026-08-11): "a detailed
# project plan that has our view and also the progress reported progress
# curve" said "project plan" and "progress curve", missed every hint, and
# silently rode the freeform sandbox lane past a template built for exactly
# that ask. Ruling: both outputs are legitimate; an ambiguous ask gets ONE
# short question describing the two versions in user terms, and an
# unambiguous ask routes with no question. en/es/ja, the same trilingual
# pattern as the intent prefilter.
AMBIGUOUS_PLAN_HINTS = (
    "project plan", "project schedule", "progress curve", "s-curve",
    "s curve", "slip history", "progress s curve",
    "plan de proyecto", "plan del proyecto", "curva de progreso",
    "cronograma",
    "プロジェクト計画",
    "進捗曲線", "工程表",
)


# A hint that the user is REFUSING, not requesting. Substring matching
# cannot tell "make me a project plan" from "I don't want a project
# plan", and the second one is what a person types precisely because we
# just got it wrong: live 2026-08-16, Scott answered the version
# question with "I don't want a project plan. I'm looking for a Test
# plan document." and we served the Gantt question straight back. A
# correction that re-triggers the thing being corrected is the worst
# moment to be literal, so a negation within a short reach of the hint
# vetoes it.
_NEGATORS = (
    "don't want", "dont want", "do not want", "not a", "not the",
    "instead of", "rather than", "no ", "not looking for",
    "isn't", "isnt", "is not",
    "no quiero", "en lugar de",
    "ne veux pas", "au lieu de",
    "ではなく", "いらない",
)
_NEGATION_REACH = 40


def _negated(hay: str, hint: str) -> bool:
    """Is this hint occurrence inside a refusal?

    Only the text BEFORE the hint counts, within a short reach: "not a
    project plan" negates, while "a project plan, not a test plan" does
    not negate the project plan. Checked per occurrence, so one negated
    mention does not excuse a later genuine one.
    """
    start = 0
    while True:
        i = hay.find(hint, start)
        if i == -1:
            return True  # every occurrence was negated
        window = hay[max(0, i - _NEGATION_REACH):i]
        if not any(n in window for n in _NEGATORS):
            return False  # this one is a real ask
        start = i + len(hint)


def ambiguous_plan_ask(text: str, format: str | None = None) -> bool:
    """True when the ask reads like plan/progress territory without
    matching a template hint, so the caller must ask which version the
    user wants BEFORE generating (Scott's ruling, 2026-08-11). Callers
    pass the QUESTION PORTION, not assembled history: a plan word in
    carried history must not re-question every later file ask (the #420
    lesson applied to disambiguation). The format veto mirrors
    match_template: a non-xlsx wish is unambiguously custom because the
    registry only builds xlsx here."""
    if format not in (None, "xlsx"):
        return False
    hay = (text or "")[-4000:].lower()
    if match_template(hay, format=format):
        return False
    return any(h in hay and not _negated(hay, h)
               for h in AMBIGUOUS_PLAN_HINTS)


def parse_extraction(text: str) -> dict:
    """JSON recovery from the extraction turn. Models sometimes narrate
    around the object or append a rendering despite the output-only-JSON
    instruction (live 2026-07-13 19:16Z: prose + valid plan JSON + a full
    HTML page — the old first-{-to-last-} slice ended inside the HTML's
    CSS braces and failed on a turn that carried a perfectly good plan).
    Decode balanced objects wherever they start and prefer the one that
    looks like a plan; fall back to the largest object found."""
    t = text or ""
    dec = json.JSONDecoder()
    candidates: list[dict] = []
    i = t.find("{")
    while i != -1:
        try:
            obj, end = dec.raw_decode(t, i)
            if isinstance(obj, dict):
                if "tasks" in obj:
                    return obj
                candidates.append(obj)
                i = t.find("{", end)
                continue
        except json.JSONDecodeError:
            pass
        i = t.find("{", i + 1)
    if candidates:
        return max(candidates, key=lambda o: len(json.dumps(o)))
    raise ValueError("no JSON object in extraction text")
