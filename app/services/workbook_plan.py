"""Render a model-authored sheet plan into a workbook.

The generation lane hands the whole job to the provider: the model writes
Python, runs it in a container, and streams back a finished file. That
works, but it spends most of its tokens orchestrating rather than
authoring. Measured 2026-08-14 on the same transcript and ask, Sonnet 5
read 212,886 cached input tokens building the file itself versus 4,433 in
a single authoring pass, a 3.6x cost difference, and produced FEWER rows
because the budget went to tool calls.

So we split the job. The model returns a PLAN (what sheets exist, what
columns they have, what the rows are) and this module renders it. The
model keeps the part it is good at, deciding what belongs in the
document. We keep the part code is good at, producing the same file the
same way every time.

WHAT THIS IS FOR: authored, tabular content, where the model composes
rows out of a transcript, memory, or reasoning. It is NOT for computing
over a supplied file (a pivot of an attached CSV) or for generating rows
at volume (a 300-row combinatorial sweep). Those still belong on the
code-execution lane, which can read data and loop. See
`needs_code_execution()` for the routing signal.

House style is enforced here rather than requested in the prompt, because
asking is unreliable: both baseline runs were told the rules and both
still shipped spaced hyphens into tab names and title rows.
"""

from __future__ import annotations

import io
import re
from typing import Any

from app.services.text_hygiene import normalize_dashes

# Reuse the Gantt lane's palette so a scenario matrix and a schedule look
# like they came from the same product.
_HEADER_BG = "3D4653"
_TITLE_FG = "3D4653"
_BAND_BG = "F7F8FA"
_GRID_FG = "E9E9E9"

_MAX_TAB = 31
_ILLEGAL_TAB = re.compile(r"[\[\]:*?/\\]")
# " - " reads as an aside, which is the one thing the house rule bans.
# Opus wrote "Dependent - Caregiver" as a tab name while its own README
# wrote "Dependent / Caregiver" correctly, so fix it rather than ask.
_SPACED_HYPHEN = re.compile(r"\s+-\s+")

_NUMBER_FORMATS = {
    "text": None,
    "number": "#,##0.##",
    "currency": "$#,##0.00",
    "percent": "0.0%",
    "date": "yyyy-mm-dd",
}

# What the model fills in. Passed as a forced tool call so the JSON is
# schema-valid on arrival and we never parse prose.
WORKBOOK_PLAN_SCHEMA: dict[str, Any] = {
    "type": "object",
    "required": ["filename", "sheets"],
    "properties": {
        "filename": {
            "type": "string",
            "description": "Lowercase, underscores, ending in .xlsx.",
        },
        "title": {
            "type": "string",
            "description": "Human title for the workbook as a whole.",
        },
        "needs_computation": {
            "type": "boolean",
            "description": (
                "True only if producing this workbook requires computing "
                "over a supplied data file, or generating rows "
                "programmatically at a volume you cannot write out by "
                "hand. If true, stop and say so instead of guessing "
                "values."
            ),
        },
        "readme": {
            "type": "object",
            "description": "Cover sheet. Omit only for trivial workbooks.",
            "properties": {
                "purpose": {"type": "string"},
                "scope": {
                    "type": "string",
                    "description": (
                        "What is covered and, importantly, what was "
                        "deliberately left out and why."
                    ),
                },
                "how_to_read": {"type": "string"},
                "reviewers": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Real names from the source material.",
                },
                "handoff": {"type": "string"},
            },
        },
        "sheets": {
            "type": "array",
            "minItems": 1,
            "items": {
                "type": "object",
                "required": ["name", "columns", "rows"],
                "properties": {
                    "name": {
                        "type": "string",
                        "description": "Tab name, 31 chars or fewer.",
                    },
                    "subtitle": {
                        "type": "string",
                        "description": "One line under the tab title.",
                    },
                    "orientation": {
                        "type": "string",
                        "enum": ["records", "matrix"],
                        "description": (
                            "'records' when every row is one item of the "
                            "same kind. 'matrix' when the first column "
                            "labels the row and the other columns are the "
                            "things being compared; that freezes the label "
                            "column so wide grids stay readable."
                        ),
                    },
                    "columns": {
                        "type": "array",
                        "minItems": 1,
                        "items": {
                            "type": "object",
                            "required": ["key", "label"],
                            "properties": {
                                "key": {
                                    "type": "string",
                                    "description": "Key used in rows[].",
                                },
                                "label": {"type": "string"},
                                "width": {
                                    "type": "integer",
                                    "description": "Chars, 8 to 80.",
                                },
                                "wrap": {"type": "boolean"},
                                "format": {
                                    "type": "string",
                                    "enum": list(_NUMBER_FORMATS),
                                },
                                "computed": {
                                    "type": "object",
                                    "description": (
                                        "Declare the intent and we emit the "
                                        "Excel formula. Do NOT write "
                                        "formulas yourself: a wrong cell "
                                        "reference is a silent wrong "
                                        "answer. Leave this column out of "
                                        "rows[]."
                                    ),
                                    "required": ["op", "of"],
                                    "properties": {
                                        "op": {
                                            "type": "string",
                                            "enum": [
                                                "sum", "difference",
                                                "product", "ratio",
                                            ],
                                        },
                                        "of": {
                                            "type": "array",
                                            "items": {"type": "string"},
                                            "description": "Column keys.",
                                        },
                                    },
                                },
                            },
                        },
                    },
                    "rows": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "additionalProperties": {
                                "type": ["string", "number", "boolean", "null"]
                            },
                        },
                    },
                },
            },
        },
    },
}


def needs_code_execution(plan: dict) -> bool:
    """Whether this plan should fall back to the provider's file lane.

    The model tells us directly. A plan that needs to compute over an
    attachment, or to generate rows it cannot write out, is one we should
    not render from authored data, because the values would be guesses.
    """
    return bool(plan.get("needs_computation"))


def _clean(value: Any) -> Any:
    """House style, applied at the cell rather than asked for in a prompt."""
    if not isinstance(value, str):
        return value
    return _SPACED_HYPHEN.sub(", ", normalize_dashes(value)).strip()


def _tab_name(raw: str, taken: set[str]) -> str:
    """Excel-legal, house-legal, and unique.

    A spaced hyphen becomes a comma rather than a slash: Excel forbids
    "/" in a sheet name, so substituting one only to strip it in the next
    pass leaves a gap where the punctuation was.
    """
    name = _SPACED_HYPHEN.sub(", ", normalize_dashes(str(raw or "Sheet")))
    name = _ILLEGAL_TAB.sub(" ", name)
    name = re.sub(r"\s+", " ", name).strip() or "Sheet"
    name = name[:_MAX_TAB].strip()
    base, n = name, 2
    while name.lower() in taken:
        suffix = f" {n}"
        name = base[: _MAX_TAB - len(suffix)] + suffix
        n += 1
    taken.add(name.lower())
    return name


def _formula(op: str, letters: list[str], row: int) -> str | None:
    """Build the formula from declared intent, wrapped so errors stay blank."""
    refs = [f"{c}{row}" for c in letters]
    if not refs:
        return None
    if op == "sum":
        body = f"SUM({','.join(refs)})"
    elif op == "difference":
        body = refs[0] if len(refs) == 1 else refs[0] + "-" + "-".join(refs[1:])
    elif op == "product":
        body = "*".join(refs)
    elif op == "ratio" and len(refs) >= 2:
        body = f"{refs[0]}/{refs[1]}"
    else:
        return None
    return f'=IFERROR({body},"")'


def render_workbook(plan: dict) -> bytes:
    """Render a sheet plan to xlsx bytes."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    sheets = plan.get("sheets") or []
    if not sheets:
        raise ValueError("plan has no sheets")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    taken: set[str] = set()

    thin = Side(style="thin", color="FF" + _GRID_FG)
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="FF" + _HEADER_BG)
    band_fill = PatternFill("solid", fgColor="FF" + _BAND_BG)

    if plan.get("readme"):
        _render_readme(wb, plan, taken, Font, Alignment)

    for spec in sheets:
        cols = spec.get("columns") or []
        rows = spec.get("rows") or []
        if not cols:
            continue
        ws = wb.create_sheet(_tab_name(spec.get("name"), taken))
        letters = {c["key"]: get_column_letter(i)
                   for i, c in enumerate(cols, start=1)}

        title = _clean(spec.get("subtitle") or spec.get("name") or "")
        tc = ws.cell(1, 1, title)
        tc.font = Font(bold=True, size=12, color="FF" + _TITLE_FG)
        tc.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 22

        for i, col in enumerate(cols, start=1):
            hc = ws.cell(2, i, _clean(col.get("label") or col["key"]))
            hc.font = Font(bold=True, size=10, color="FFFFFFFF")
            hc.fill = head_fill
            hc.alignment = Alignment(wrap_text=True, vertical="center")
            hc.border = grid
            width = col.get("width")
            if not isinstance(width, int):
                # Size to the content rather than guessing, so a
                # Description column is not the width of an ID column.
                longest = max(
                    [len(str(col.get("label") or ""))]
                    + [len(str(r.get(col["key"], ""))) for r in rows[:200]]
                )
                width = longest + 2
            ws.column_dimensions[get_column_letter(i)].width = (
                max(8, min(int(width), 80)))
        ws.row_dimensions[2].height = 30

        for r, row in enumerate(rows, start=3):
            for i, col in enumerate(cols, start=1):
                computed = col.get("computed")
                if computed:
                    value = _formula(
                        computed.get("op", ""),
                        [letters[k] for k in computed.get("of", [])
                         if k in letters],
                        r)
                else:
                    value = _clean(row.get(col["key"]))
                cell = ws.cell(r, i, value)
                cell.font = Font(size=10)
                cell.border = grid
                cell.alignment = Alignment(
                    wrap_text=bool(col.get("wrap", True)),
                    vertical="top")
                fmt = _NUMBER_FORMATS.get(col.get("format") or "text")
                if fmt:
                    cell.number_format = fmt
                if r % 2 == 0:
                    cell.fill = band_fill

        # A totals row, when the plan asks for one. A budget without
        # column totals is a list of numbers, not a budget, and we build
        # it as a live SUM so editing a line updates it.
        totals = [k for k in (spec.get("totals") or []) if k in letters]
        if totals and rows:
            trow = len(rows) + 3
            label = ws.cell(trow, 1, "Total")
            label.font = Font(bold=True, size=10)
            label.border = grid
            for i, col in enumerate(cols, start=1):
                cell = ws.cell(trow, i)
                cell.border = grid
                if col["key"] in totals:
                    letter = letters[col["key"]]
                    cell.value = (f'=IFERROR(SUM({letter}3:'
                                  f'{letter}{len(rows) + 2}),"")')
                    cell.font = Font(bold=True, size=10)
                    fmt = _NUMBER_FORMATS.get(col.get("format") or "text")
                    if fmt:
                        cell.number_format = fmt

        last = get_column_letter(len(cols))
        if rows:
            ws.auto_filter.ref = f"A2:{last}{len(rows) + 2}"
        # A matrix's first column is the row label, so it has to stay on
        # screen the way the header row does.
        ws.freeze_panes = ("B3" if spec.get("orientation") == "matrix"
                           else "A3")
        ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_readme(wb, plan: dict, taken: set[str], Font, Alignment) -> None:
    """Cover sheet, always built the same way."""
    rm = plan["readme"]
    ws = wb.create_sheet(_tab_name("README", taken))
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 96
    ws.sheet_view.showGridLines = False

    tc = ws.cell(1, 1, _clean(plan.get("title") or "Workbook"))
    tc.font = Font(bold=True, size=14, color="FF" + _TITLE_FG)
    ws.row_dimensions[1].height = 24

    fields = [("Purpose", "purpose"), ("Scope", "scope"),
              ("How to read", "how_to_read"), ("Reviewers", "reviewers"),
              ("Handoff", "handoff")]
    row = 3
    for label, key in fields:
        value = rm.get(key)
        if not value:
            continue
        if isinstance(value, list):
            value = ", ".join(str(v) for v in value)
        lc = ws.cell(row, 1, label)
        lc.font = Font(bold=True, size=10, color="FF" + _TITLE_FG)
        lc.alignment = Alignment(vertical="top")
        vc = ws.cell(row, 2, _clean(value))
        vc.font = Font(size=10)
        vc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = max(
            18, 14 * (len(str(value)) // 90 + 1))
        row += 2
