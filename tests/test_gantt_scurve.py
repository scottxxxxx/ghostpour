"""Progress sheet: reported progress from meetings against the plan.

Series that are deliberately not the same kind of thing:

  First plan (baseline)  what the FIRST plan version said, static,
            because editing today's dates must not rewrite what was
            promised.
  Current plan  what the CURRENT plan says, live formulas over the Gantt
            View's own date cells so it redraws like the bars do. Solid
            up to the data date, dashed past it.
  Reported  duration-weighted percent complete as of each meeting, held
            flat in between because that is genuinely all we know, and
            stopped at the data date (the last meeting that reported):
            actuals never render for weeks that have not reported.

All weighted by working days counted inclusively, matching Excel's
NETWORKDAYS, so the Python-computed history and the live formulas weight
the same work the same way. Reframed per the 2026-08-11 PM review: this
is a meeting-derived progress curve, not an EVM S-curve, and the sheet
says so.
"""

from __future__ import annotations

import datetime
import io
import shutil
import subprocess

import openpyxl
import pytest

from app.services.doc_templates import (
    _planned_share,
    _wd_inclusive,
    _weighted_progress,
    render_gantt,
    render_gantt_detailed,
)

D = datetime.date.fromisoformat

_PLAN = {
    "project": "Field Kit",
    "meeting_date": "2026-07-27",
    "tasks": [
        {"id": 1, "name": "Phase", "type": "phase", "parent_id": None,
         "owner": None, "status": "in_progress", "start": "2026-07-06",
         "end": "2026-07-31", "depends_on": []},
        {"id": 2, "name": "A", "type": "task", "parent_id": 1, "owner": None,
         "status": "complete", "start": "2026-07-06", "end": "2026-07-10",
         "depends_on": [], "percent_complete": 100},
        {"id": 3, "name": "B", "type": "task", "parent_id": 1, "owner": None,
         "status": "in_progress", "start": "2026-07-13", "end": "2026-07-24",
         "depends_on": [2], "percent_complete": 50},
        {"id": 4, "name": "C", "type": "task", "parent_id": 1, "owner": None,
         "status": "not_started", "start": "2026-07-27", "end": "2026-07-31",
         "depends_on": [3]},
    ],
}

_HISTORY = [{"as_of": "2026-07-13", "tasks": [
    {"id": 2, "name": "A", "type": "task", "start": "2026-07-06",
     "end": "2026-07-10", "percent_complete": 60},
    {"id": 3, "name": "B", "type": "task", "start": "2026-07-13",
     "end": "2026-07-22", "percent_complete": 0},
    {"id": 4, "name": "C", "type": "task", "start": "2026-07-23",
     "end": "2026-07-29"},
]}]


def _book(history=_HISTORY):
    blob = render_gantt_detailed(_PLAN, today=D("2026-07-27"), history=history)
    return openpyxl.load_workbook(io.BytesIO(blob))


# --- counting convention --------------------------------------------------

def test_working_days_are_counted_inclusively_like_networkdays():
    assert _wd_inclusive(D("2026-07-06"), D("2026-07-10")) == 5   # Mon..Fri
    assert _wd_inclusive(D("2026-07-06"), D("2026-07-06")) == 1
    assert _wd_inclusive(D("2026-07-06"), D("2026-07-13")) == 6   # skips a weekend
    assert _wd_inclusive(D("2026-07-13"), D("2026-07-06")) == 0   # inverted


def test_progress_uses_the_same_rule_as_the_phase_rollups():
    tasks = [
        {"type": "task", "start": "2026-07-06", "end": "2026-07-10",
         "percent_complete": 50},                       # 5 wd at 50%
        {"type": "task", "start": "2026-07-13", "end": "2026-07-17",
         "status": "complete"},                         # 5 wd, Complete -> 100
        {"type": "task", "start": "2026-07-20", "end": "2026-07-24"},
    ]                                                   # 5 wd, unstated -> 0
    assert _weighted_progress(tasks) == pytest.approx((2.5 + 5) / 15)


def test_milestones_and_phases_carry_no_weight():
    tasks = [
        {"type": "task", "start": "2026-07-06", "end": "2026-07-10",
         "percent_complete": 0},
        {"type": "milestone", "start": "2026-07-13", "end": "2026-07-13",
         "status": "complete"},
        {"type": "phase", "start": "2026-07-06", "end": "2026-07-13",
         "percent_complete": 100},
    ]
    # a complete milestone must not drag the curve up on its own
    assert _weighted_progress(tasks) == 0


def test_planned_share_walks_from_zero_to_one():
    tasks = _PLAN["tasks"]
    assert _planned_share(tasks, D("2026-07-01")) == 0
    assert _planned_share(tasks, D("2026-07-10")) == pytest.approx(0.25)
    assert _planned_share(tasks, D("2026-08-31")) == 1


# --- the sheet ------------------------------------------------------------

def test_sheet_is_added_to_the_detailed_style_only():
    # tab says "Progress", not "S-Curve": the tab name is a claim too
    assert _book().sheetnames == ["Gantt View", "Slip", "Receipts", "Progress"]
    simple = openpyxl.load_workbook(io.BytesIO(
        render_gantt(_PLAN, today=D("2026-07-27"))))
    assert simple.sheetnames == ["Gantt View"]


def test_title_and_labels_do_not_claim_an_evm_s_curve():
    sc = _book()["Progress"]
    assert sc["A1"].value == "Reported progress (from meetings)"
    assert sc.cell(5, 2).value == "First plan (baseline)"
    assert "planned only" in str(sc.cell(5, 4).value)
    assert "not an earned-value S-curve" in str(sc["A2"].value)


def test_planned_is_live_off_the_gantt_view_dates():
    sc = _book()["Progress"]
    f = sc.cell(6, 3).value
    assert isinstance(f, str) and f.startswith("=IFERROR(")
    assert "'Gantt View'!$E" in f and "NETWORKDAYS" in f, f
    assert "$J$1" in f, "shares one live denominator"


def test_baseline_is_static_history_not_a_formula():
    sc = _book()["Progress"]
    v = sc.cell(6, 2).value
    assert isinstance(v, float), "editing today's plan must not rewrite it"
    assert 0 <= v <= 1


def test_reported_holds_flat_between_meetings_and_is_blank_before_the_first():
    sc = _book()["Progress"]
    col = [sc.cell(r, 5).value for r in range(6, 6 + 5)]
    assert col[0] is None, "nothing reported before the first meeting we have"
    # Every point divides by the CURRENT plan's 20 scheduled working days,
    # never by the version's own scope. At the 2026-07-13 meeting 3 days
    # were banked (A at 60% of 5 days), so 3/20; today 10 (A done, B half
    # of 10 days), so 10/20.
    assert col[1] == pytest.approx(0.15)
    assert col[2] == pytest.approx(0.15), "held flat, never interpolated"
    assert col[3] == pytest.approx(0.5)


def test_reported_stops_at_the_data_date():
    """PM review 2026-08-11: the line used to carry the last value flat
    into future weeks, rendering actuals for weeks that never reported.
    The data date here is 2026-07-27, whose week ends Jul 31; the Aug 7
    week is in the future and must be blank."""
    sc = _book()["Progress"]
    assert sc.cell(9, 5).value == pytest.approx(0.5), "last reported week"
    assert sc.cell(10, 5).value is None, "no actuals past the data date"


def test_the_plan_past_the_data_date_is_its_own_dashed_series():
    sc = _book()["Progress"]
    # solid current plan runs up to the data date week and stops
    assert isinstance(sc.cell(9, 3).value, str)
    assert sc.cell(10, 3).value is None
    # the dashed continuation shares the data date week (connected line)
    # and carries the future weeks, still as live formulas
    assert isinstance(sc.cell(9, 4).value, str)
    assert isinstance(sc.cell(10, 4).value, str)
    assert sc.cell(6, 4).value is None, "nothing left of the data date"


def test_the_data_date_marker_stands_on_the_data_date_week():
    sc = _book()["Progress"]
    col = [sc.cell(r, 7).value for r in range(6, 6 + 5)]
    assert col == [None, None, None, 1.0, None], col
    ch = sc._charts[0]
    dd = ch.series[5]
    assert dd.errBars is not None, "the vertical line is a y error bar"
    assert dd.errBars.errBarType == "minus"
    assert dd.errBars.val == 1.0, "drops the full axis height"
    assert dd.graphicalProperties.line.noFill is True


def test_the_coverage_line_counts_dates_on_the_source_activities():
    undated = {**_PLAN, "tasks": _PLAN["tasks"] + [
        {"id": 8, "name": "Vendor signature", "type": "task", "parent_id": 1,
         "owner": None, "status": "not_started", "start": None, "end": None,
         "depends_on": []}]}
    sc = openpyxl.load_workbook(io.BytesIO(render_gantt_detailed(
        undated, today=D("2026-07-27"), history=_HISTORY)))["Progress"]
    assert "dates on 3 of 4 activities" in str(sc["A4"].value)
    fully = _book()["Progress"]
    assert "dates on 3 of 3 activities" in str(fully["A4"].value)


def test_growing_scope_never_pushes_reported_backwards():
    """The 2026-07-30 finding: Reported fell 18.5% to 16.5% because a
    meeting stretched one task and added another with no stated percent, so
    the same banked work divided by a bigger plan. A cumulative line that
    drops when scope GROWS makes a real plan look like a failing one."""
    import io
    import openpyxl
    from app.services.doc_templates import render_gantt_detailed
    grown = {**_PLAN, "tasks": _PLAN["tasks"] + [
        {"id": 9, "name": "D, added later", "type": "task", "parent_id": 1,
         "owner": None, "status": "not_started", "start": "2026-07-27",
         "end": "2026-08-07", "depends_on": []}]}
    sc = openpyxl.load_workbook(io.BytesIO(render_gantt_detailed(
        grown, today=D("2026-07-27"), history=_HISTORY)))["Progress"]
    seen = [sc.cell(r, 5).value for r in range(6, 21)]
    seen = [v for v in seen if v is not None]
    assert seen == sorted(seen), f"reported went backwards: {seen}"


def test_a_first_ever_plan_says_why_the_baseline_is_missing():
    sc = _book(history=None)["Progress"]
    assert sc.cell(6, 2).value is None
    text = " ".join(str(sc.cell(r, 1).value or "") for r in range(1, sc.max_row + 1))
    assert "second plan version" in text


def test_chart_exists_with_all_series():
    sc = _book()["Progress"]
    assert len(sc._charts) == 1
    # baseline, current plan, its dashed continuation past the data date,
    # reported, the meeting dots, and the data date marker
    assert len(sc._charts[0].series) == 6
    assert sc._charts[0].series[4].marker.symbol == "circle"


def test_header_explains_which_lines_move_and_which_do_not():
    sc = _book()["Progress"]
    note = str(sc["A2"].value)
    assert "never moves" in note
    assert "redraws" in note
    assert "held flat" in note
    assert "stops at the data date" in note
    assert "plan only" in note


# --- behavior, via a real formula engine ----------------------------------

_SOFFICE = shutil.which("soffice")


@pytest.mark.skipif(_SOFFICE is None, reason="needs LibreOffice to recalculate")
def test_recalculated_planned_curve_matches_hand_computed_shares(tmp_path):
    """A is 5 working days, B is 10, C is 5, so 20 total. The planned curve
    should reach a quarter, a half, three quarters and then all of it."""
    src = tmp_path / "curve.xlsx"
    src.write_bytes(render_gantt_detailed(
        _PLAN, today=D("2026-07-27"), history=_HISTORY))
    out = tmp_path / "out"
    out.mkdir()
    subprocess.run(
        [_SOFFICE, "--headless", "--convert-to", "xlsx",
         str(src), "--outdir", str(out)],
        check=True, capture_output=True, timeout=240,
    )
    sc = openpyxl.load_workbook(out / "curve.xlsx", data_only=True)["Progress"]
    assert sc["J1"].value == 20, "total scheduled working days"
    planned = [sc.cell(r, 3).value for r in range(6, 10)]
    assert planned == pytest.approx([0.25, 0.5, 0.75, 1.0])
    # the dashed continuation past the data date computes the same shares
    assert sc.cell(9, 4).value == pytest.approx(1.0)
    assert sc.cell(10, 4).value == pytest.approx(1.0)
    # the baseline plan was a different shape, which is the whole point
    assert sc.cell(6, 2).value == pytest.approx(5 / 18)


def test_the_chart_is_readable_not_just_correct():
    """Fixes from Scott's 2026-07-31 screenshot: the axis had exploded into
    one tick per DAY (44 labels for 7 points) because the categories were
    real dates, which also shoved the axis title on top of the labels."""
    sc = _book()["Progress"]
    # categories are text labels, in a VISIBLE column: a chart cannot plot
    # categories out of a hidden column, which is why the live dates moved
    # to the hidden helper instead of the other way round
    assert sc.cell(6, 1).value == "Jul 10"   # first week in this fixture
    assert sc.column_dimensions["A"].hidden is False
    assert sc.column_dimensions["J"].hidden is True
    ch = sc._charts[0]
    assert ch.title is None, "the sheet header already names it"
    assert ch.x_axis.title is None and ch.y_axis.title is None
    assert ch.y_axis.scaling.max == 1, "progress cannot exceed 100%"
    # right rail, not bottom: bottom interleaved the legend entries
    # with the date labels in Excel (2026-07-31)
    assert ch.legend.position == "r"
    assert ch.x_axis.txPr is not None, "axis text needs its own size"
    # three identities, three validated hues; the plan keeps one hue
    # across its solid and dashed halves
    hexes = [s.graphicalProperties.line.solidFill.srgbClr
             for s in ch.series[:4]]
    assert hexes == ["B5651D", "1F4E9C", "1F4E9C", "2E9E4F"]
    assert ch.series[0].graphicalProperties.line.dashStyle == "dash"
    assert ch.series[2].graphicalProperties.line.dashStyle == "dash", \
        "right of the data date is plan only and renders dashed"
    assert ch.series[1].graphicalProperties.line.dashStyle in (None, "solid")
    assert ch.series[3].graphicalProperties.line.dashStyle in (None, "solid")
    assert all(s.smooth is False for s in ch.series[:4]), "no invented curve"


def _legend_of(plan, history=None):
    import io
    import openpyxl
    from app.services.doc_templates import render_gantt_detailed
    wb = openpyxl.load_workbook(io.BytesIO(render_gantt_detailed(
        plan, today=D("2026-07-27"), history=history)))
    return wb["Progress"]._charts[0].legend


def test_legend_takes_the_bottom_right_when_the_project_finishes_high():
    """Scott 2026-08-01: the lines ran through the legend. The curves climb
    left to right, so the empty corner depends on the shape. The plan here
    runs well past the data date so the vertical marker sits mid plot and
    the bottom right corner is genuinely free."""
    longer = {**_PLAN, "tasks": _PLAN["tasks"] + [
        {"id": 5, "name": "E", "type": "task", "parent_id": 1, "owner": None,
         "status": "not_started", "start": "2026-08-10", "end": "2026-08-21",
         "depends_on": []}]}
    lg = _legend_of(longer, _HISTORY)
    assert lg.overlay is True
    assert lg.layout.manualLayout.x == 0.70
    assert lg.layout.manualLayout.y == 0.52


def test_the_data_date_marker_evicts_the_legend_from_its_corner():
    """The vertical marker runs the full plot height, so a corner it
    passes through is occupied ink even when every curve avoids it. In
    this fixture the data date week sits at the right of the plot, where
    the bottom right corner used to win."""
    lg = _legend_of(_PLAN, _HISTORY)
    assert lg.overlay is True
    assert lg.layout.manualLayout.x == 0.14, "top left instead"
    assert lg.layout.manualLayout.y == 0.08


def test_a_struggling_project_moves_the_legend_to_the_top_left():
    """Curves that stay low on the right fill the bottom-right corner and
    leave the top-left empty. Same reason, opposite answer."""
    stalled = {**_PLAN, "tasks": [
        {**t, "percent_complete": 0,
         "status": "not_started" if t.get("type") != "phase" else t.get("status")}
        for t in _PLAN["tasks"]]}
    lg = _legend_of(stalled)
    assert lg.overlay is True
    assert lg.layout.manualLayout.x == 0.14, "top left"
    assert lg.layout.manualLayout.y == 0.08


def test_the_legend_stays_off_the_plot_when_no_corner_is_clear():
    """Neither corner free is a real case, and sitting on the plot anyway
    would be worse than spending the width."""
    # front-loaded schedule with no progress: Planned is already 71% in the
    # first third (fills the top left) while Reported flatlines at 0 (fills
    # the bottom right). Rare, and exactly the case worth not guessing at.
    awkward = {"project": "Stuck", "meeting_date": "2026-07-14", "tasks": [
        {"id": 1, "name": "P", "type": "phase", "parent_id": None,
         "status": "in_progress", "start": "2026-07-06", "end": "2026-07-14",
         "depends_on": []},
        {"id": 2, "name": "Bulk", "type": "task", "parent_id": 1,
         "status": "not_started", "start": "2026-07-06", "end": "2026-07-10",
         "depends_on": [], "percent_complete": 0},
        {"id": 3, "name": "Tail", "type": "task", "parent_id": 1,
         "status": "not_started", "start": "2026-07-13", "end": "2026-07-14",
         "depends_on": [2], "percent_complete": 0}]}
    lg = _legend_of(awkward)
    assert lg.position == "r"
    assert lg.layout is None, "outside the plot, not overlaid on it"
