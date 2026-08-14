"""Meetings are not tasks (Scott, 2026-08-13).

A live export came back with 29 rows of which 22 were meeting titles:
"Weekly Status Sync on CKS and ABM Items", "Project Status Review: SDK
Fixes, Deployment Readiness, and Scope Concerns". Every bar was a single
day, because a meeting happens on one date, and they were chained one per
working day into a critical path made of somebody's calendar. The progress
curve then measured the share of MEETINGS whose subject sounded finished
and fell from 50% to 38% when later meetings reopened topics, which is not
a project going backwards.

Two halves, because a prompt constrains and does not prove:

  instructed  the extraction prompt now says what earns a row, and names
              the meeting headings in the memory block as containers
              rather than work.
  structural  a task whose start equals its end is retyped as a
              milestone, so points can never be weighed as spans.

The structural half is deliberately not a name heuristic. Whether a string
looks like a meeting title is a guess; whether a row has a duration is a
fact.
"""

from __future__ import annotations

import datetime
import io

import openpyxl
import pytest

from app.services.doc_templates import (
    _GANTT_DETAILED_SCHEMA_PROMPT,
    _GANTT_SCHEMA_PROMPT,
    _points_are_milestones,
    _weighted_progress,
    render_gantt_detailed,
)

D = datetime.date.fromisoformat


def _task(i, name, start, end, **kw):
    return {"id": i, "name": name, "type": "task", "parent_id": 1,
            "owner": None, "status": "complete", "start": start,
            "end": end, "depends_on": [], **kw}


# The shape the ABM export actually came back in: a phase, then a run of
# one-day rows named after the meetings they came from.
_MEETING_SHAPED = {
    "project": "Helpdesk AI",
    "meeting_date": "2026-08-11",
    "tasks": [
        {"id": 1, "name": "Rollout", "type": "phase", "parent_id": None,
         "owner": None, "status": "in_progress", "start": "2026-07-13",
         "end": "2026-07-17", "depends_on": []},
        _task(2, "KB Retrieval QA and Production Rollout Planning",
              "2026-07-13", "2026-07-13"),
        _task(3, "Weekly Status Sync on CKS and ABM Items",
              "2026-07-14", "2026-07-14"),
        _task(4, "Project Status Review: SDK Fixes and Scope Concerns",
              "2026-07-15", "2026-07-15", status="blocked"),
    ],
}


def test_a_row_with_no_duration_is_a_point_not_work():
    out = _points_are_milestones(_MEETING_SHAPED)
    types = [t["type"] for t in out["tasks"]]
    assert types == ["phase", "milestone", "milestone", "milestone"]
    # everything a reader can see survives the retype
    kept = out["tasks"][3]
    assert kept["name"] == "Project Status Review: SDK Fixes and Scope Concerns"
    assert kept["status"] == "blocked"
    assert kept["start"] == "2026-07-15"


def test_real_spans_are_left_alone():
    """The guard fires on degenerate data only. A plan with genuine
    durations must come through untouched."""
    plan = {"tasks": [
        {"id": 1, "name": "Phase", "type": "phase", "parent_id": None,
         "start": "2026-07-06", "end": "2026-07-31"},
        _task(2, "Structured cabling", "2026-06-24", "2026-07-21"),
        _task(3, "Switch stack build", "2026-06-29", "2026-08-05"),
    ]}
    assert _points_are_milestones(plan) is plan, "no copy when nothing changed"


def test_the_input_is_not_mutated():
    """Callers hand us the extracted plan and may keep using it."""
    before = [t["type"] for t in _MEETING_SHAPED["tasks"]]
    _points_are_milestones(_MEETING_SHAPED)
    assert [t["type"] for t in _MEETING_SHAPED["tasks"]] == before


def test_points_cannot_manufacture_a_progress_percentage():
    """The actual damage. Weighed as one-day tasks these rows produce a
    percent complete out of nothing; as milestones they carry no weight,
    which is why the retype is the fix rather than cosmetic."""
    as_tasks = _MEETING_SHAPED["tasks"][1:]
    assert _weighted_progress(as_tasks) == pytest.approx(2 / 3)
    retyped = _points_are_milestones(_MEETING_SHAPED)["tasks"][1:]
    # None, not 0: there is nothing here to weigh, which is a refusal to
    # state a percentage rather than a claim that no progress was made.
    # "67% complete" and "0% complete" are both fabrications from this data.
    assert _weighted_progress(retyped) is None


def test_a_plan_made_only_of_points_charts_no_progress_curve():
    """The honest outcome: nothing in this plan states how long any work
    takes, so there is no curve to draw. Previously it drew one from the
    meetings, which is the bug."""
    wb = openpyxl.load_workbook(io.BytesIO(render_gantt_detailed(
        _MEETING_SHAPED, today=D("2026-08-11"))))
    assert "Progress" not in wb.sheetnames
    # the rows themselves are still on the timeline, as points
    names = [wb["Gantt View"].cell(r, 2).value or ""
             for r in range(1, wb["Gantt View"].max_row + 1)]
    assert any("Weekly Status Sync" in str(n) for n in names)


def test_both_prompts_say_a_meeting_is_not_a_row():
    """The instructed half. Cheap guard so nobody edits this back out
    without meeting the reason it exists."""
    for prompt in (_GANTT_SCHEMA_PROMPT, _GANTT_DETAILED_SCHEMA_PROMPT):
        assert "A meeting is not work." in prompt
        assert "never name a row after one" in prompt
        assert "grouped under meeting headings" in prompt
        # and the old blanket instruction is gone
        assert "Extract every task and milestone discussed. Output" not in prompt


def test_the_prompts_still_forbid_dashes():
    """The no-dash rule sits next to the new text; a careless edit to one
    is exactly how the other goes missing."""
    for prompt in (_GANTT_SCHEMA_PROMPT, _GANTT_DETAILED_SCHEMA_PROMPT):
        assert "without em " in prompt or "without em or en dashes" in prompt
        assert "—" not in prompt and "–" not in prompt
