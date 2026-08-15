"""The model plans the workbook, we build it.

Measured 2026-08-14 on the real Cigna ask, with the exact system blocks
and user message pulled off the wire and only the mechanism swapped:
handing the whole job to the provider's code-execution lane cost $0.3930
for 47 scenarios, while asking the same model for a plan and rendering it
here cost $0.2824 for 46. Most of the premium is the server-side loop
re-reading its own conversation, 212,886 cached input tokens on the
Sonnet run against 4,433 for a single authoring pass.

These tests cover the half that is ours. The model's judgment about what
belongs in the document is not testable here; the guarantees are.

The house-style tests earn their place. Both file-lane runs were TOLD the
no-dash rule in the system prompt and both shipped violations anyway:
Opus wrote the tab name "Dependent - Caregiver" while its own README
wrote it correctly, and Sonnet put a spaced hyphen in the title row of
all four tabs. Asking is not a control. Sanitizing is.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.services.workbook_plan import (
    WORKBOOK_PLAN_SCHEMA,
    needs_code_execution,
    render_workbook,
)


def _load(plan: dict):
    return openpyxl.load_workbook(io.BytesIO(render_workbook(plan)))


def _plan(**over) -> dict:
    base = {
        "filename": "x.xlsx",
        "sheets": [{
            "name": "Refill",
            "columns": [
                {"key": "id", "label": "ID"},
                {"key": "desc", "label": "Description"},
            ],
            "rows": [
                {"id": "RF-01", "desc": "Standard refill available"},
                {"id": "RF-02", "desc": "Too soon to refill"},
            ],
        }],
    }
    base.update(over)
    return base


def test_header_and_data_land_where_the_reader_expects() -> None:
    ws = _load(_plan())["Refill"]
    assert ws["A1"].value == "Refill"          # title
    assert [ws["A2"].value, ws["B2"].value] == ["ID", "Description"]
    assert ws["A3"].value == "RF-01"
    assert ws["A4"].value == "RF-02"
    # Header stays on screen and the block is filterable. The provider
    # lane set autofilter on 0 of 13 tabs across both baseline runs.
    assert ws.freeze_panes == "A3"
    assert ws.auto_filter.ref == "A2:B4"


def test_house_style_is_enforced_rather_than_requested() -> None:
    """Dashes the model wrote anyway do not reach the file."""
    plan = _plan(
        title="Scenarios — Cigna",
        readme={"purpose": "Cover P1 – P2 intents"},
        sheets=[{
            "name": "Dependent - Caregiver",
            "subtitle": "Refill Intent - Test Scenarios",
            "columns": [{"key": "a", "label": "Note — detail"}],
            "rows": [{"a": "Member calls — caregiver on file"}],
        }],
    )
    wb = _load(plan)
    blob = "\n".join(
        str(c.value) for ws in wb.worksheets
        for row in ws.iter_rows() for c in row if c.value is not None)
    blob += "\n" + "\n".join(wb.sheetnames)

    assert "—" not in blob and "–" not in blob
    # " - " reads as an aside, which is the case the rule is actually about
    assert " - " not in blob


def test_tab_names_are_excel_legal_unique_and_bounded() -> None:
    plan = _plan(sheets=[
        {"name": "Q3 P/L [draft]", "columns": [{"key": "a", "label": "A"}],
         "rows": [{"a": 1}]},
        {"name": "R" * 60, "columns": [{"key": "a", "label": "A"}],
         "rows": [{"a": 1}]},
        {"name": "R" * 60, "columns": [{"key": "a", "label": "A"}],
         "rows": [{"a": 1}]},
    ])
    names = _load(plan).sheetnames
    assert not any(ch in n for n in names for ch in r"[]:*?/\ ".strip())
    assert all(len(n) <= 31 for n in names)
    assert len(set(names)) == len(names), "duplicate tabs would raise"


def test_computed_columns_emit_a_formula_from_declared_intent() -> None:
    """The model says what it wants; we write the reference.

    A model writing its own "=SUMIF(...)" is a silent wrong answer that
    looks authoritative, and neither lane can verify it.
    """
    plan = _plan(sheets=[{
        "name": "Budget",
        "columns": [
            {"key": "lo", "label": "Low", "format": "currency"},
            {"key": "hi", "label": "High", "format": "currency"},
            {"key": "tot", "label": "Total",
             "computed": {"op": "sum", "of": ["lo", "hi"]}},
        ],
        "rows": [{"lo": 10, "hi": 25}, {"lo": 5, "hi": 5}],
    }])
    ws = _load(plan)["Budget"]
    assert ws["C3"].value == '=IFERROR(SUM(A3,B3),"")'
    assert ws["C4"].value == '=IFERROR(SUM(A4,B4),"")'
    assert ws["A3"].number_format == "$#,##0.00"


@pytest.mark.parametrize("op,expected", [
    ("difference", '=IFERROR(A3-B3,"")'),
    ("product", '=IFERROR(A3*B3,"")'),
    ("ratio", '=IFERROR(A3/B3,"")'),
])
def test_each_declared_op_has_a_formula(op: str, expected: str) -> None:
    plan = _plan(sheets=[{
        "name": "S",
        "columns": [
            {"key": "a", "label": "A"}, {"key": "b", "label": "B"},
            {"key": "c", "label": "C", "computed": {"op": op,
                                                    "of": ["a", "b"]}},
        ],
        "rows": [{"a": 1, "b": 2}],
    }])
    assert _load(plan)["S"]["C3"].value == expected


def test_matrix_orientation_freezes_the_label_column() -> None:
    """A comparison grid's first column labels the row, so it has to
    stay on screen the way the header row does."""
    plan = _plan(sheets=[{
        "name": "Vendors", "orientation": "matrix",
        "columns": [{"key": "feature", "label": "Feature"},
                    {"key": "v1", "label": "Vendor A"}],
        "rows": [{"feature": "SSO", "v1": "yes"}],
    }])
    assert _load(plan)["Vendors"].freeze_panes == "B3"


def test_a_row_missing_a_declared_key_renders_blank_not_broken() -> None:
    """Key drift is the predicted failure mode of this design.

    It did not occur in any measured run, but a sparse optional column is
    normal and must not shift the row or raise.
    """
    plan = _plan(sheets=[{
        "name": "S",
        "columns": [{"key": "id", "label": "ID"},
                    {"key": "notes", "label": "Notes"}],
        "rows": [{"id": "A1", "notes": "seen"}, {"id": "A2"}],
    }])
    ws = _load(plan)["S"]
    assert ws["A4"].value == "A2"
    assert ws["B4"].value in (None, "")


def test_needs_computation_is_the_signal_to_use_the_provider_lane() -> None:
    """Authoring renders here; computing over supplied data does not.

    The model is asked to declare it, so a pivot of an attached CSV or a
    generated combinatorial sweep falls back instead of rendering values
    the model guessed.
    """
    assert needs_code_execution({"needs_computation": True}) is True
    assert needs_code_execution({"needs_computation": False}) is False
    assert needs_code_execution({}) is False


def test_readme_is_rendered_when_present_and_skipped_when_not() -> None:
    with_readme = _load(_plan(title="T", readme={"purpose": "why",
                                                 "reviewers": ["Mike"]}))
    assert "README" in with_readme.sheetnames
    body = "\n".join(str(c.value) for row in with_readme["README"].iter_rows()
                     for c in row if c.value is not None)
    assert "why" in body and "Mike" in body
    assert "README" not in _load(_plan()).sheetnames


def test_schema_is_a_usable_tool_input_schema() -> None:
    """It is passed to the provider as a tool input_schema, so a
    malformed shape fails the request rather than degrading."""
    s = WORKBOOK_PLAN_SCHEMA
    assert s["type"] == "object"
    assert "sheets" in s["required"]
    sheet = s["properties"]["sheets"]["items"]
    assert set(sheet["required"]) == {"name", "columns", "rows"}
    col = sheet["properties"]["columns"]["items"]
    assert set(col["required"]) == {"key", "label"}
    # The model must never hand us a formula string.
    assert "computed" in col["properties"]
    assert col["properties"]["computed"]["required"] == ["op", "of"]


def test_empty_plan_is_rejected_rather_than_written_as_an_empty_file() -> None:
    with pytest.raises(ValueError):
        render_workbook({"filename": "x.xlsx", "sheets": []})
