"""Slip from what people SAID, plus the review pass that came with it.

The live 2026-07-30 run exposed the weakness: slip matched tasks by
normalized name across plan versions, the extraction model renamed
"Payments Integration" to "Payments provider integration", and a task
that had visibly moved twice in the meetings reported zero moves. The
quotes behind those moves were sitting in Receipts the whole time.

So the commitments a person actually spoke are now a first-class source,
merged with the plan-version trail rather than replacing it. Everything
here is about the pair: what the meeting said, and what the plan did.

Also covered, from the same pass: the S-curve category axis (which was
rendering date serials), user-facing field labels on Receipts, superseded
quotes, and conditional dates.
"""

from __future__ import annotations

import datetime
import io
import re
import zipfile

import openpyxl

# Payments moves Jul 10 -> Jul 17 -> Jul 24 entirely through what people
# said, and this plan has NO snapshot history at all: the case that used
# to render an empty Slip sheet.
_PLAN = {
    "project": "Field Kit",
    "meeting_date": "2026-07-20",
    "tasks": [
        {"id": 1, "name": "Release 1.2", "type": "phase", "parent_id": None,
         "owner": None, "status": "in_progress", "start": "2026-06-22",
         "end": "2026-08-03", "depends_on": []},
        {"id": 2, "name": "Payments integration", "type": "task",
         "parent_id": 1, "owner": "Maya Chen", "status": "in_progress",
         "start": "2026-06-22", "end": "2026-07-24", "depends_on": [],
         "percent_complete": 70, "effort": None, "conditional": None,
         "commitments": [
             {"date": "2026-07-10", "as_of": "2026-06-22",
              "quote": "payments integration should land by July 10",
              "speaker": "Maya", "reason": None},
             {"date": "2026-07-17", "as_of": "2026-07-06",
              "quote": "payments is pushed to the 17th though",
              "speaker": "Maya",
              "reason": "sandbox creds took a week to arrive"},
             {"date": "2026-07-24", "as_of": "2026-07-13",
              "quote": "realistically the 24th now", "speaker": "Maya",
              "reason": "webhook signature verification failing on their side"},
         ],
         "evidence": [
             {"field": "end", "quote": "payments integration should land by July 10",
              "speaker": "Maya", "meeting_date": "2026-06-22"},
             {"field": "end", "quote": "realistically the 24th now",
              "speaker": "Maya", "meeting_date": "2026-07-13"},
             {"field": "percent_complete", "quote": "payments is 70 percent",
              "speaker": "Maya", "meeting_date": "2026-07-20"},
         ]},
        {"id": 3, "name": "Offline sync", "type": "task", "parent_id": 1,
         "owner": "Jordan Lee", "status": "in_progress", "start": "2026-07-06",
         "end": "2026-07-22", "depends_on": [], "percent_complete": None,
         "effort": "3 days", "conditional": None,
         "commitments": [
             {"date": "2026-07-22", "as_of": "2026-06-29",
              "quote": "I'll have it by the 22nd", "speaker": "Jordan",
              "reason": None},
         ],
         "evidence": [{"field": "effort", "quote": "three days of actual work",
                       "speaker": "Jordan", "meeting_date": "2026-06-29"}]},
        {"id": 4, "name": "General release", "type": "milestone",
         "parent_id": 1, "owner": "Alex", "status": "not_started",
         "start": "2026-08-03", "end": "2026-08-03", "depends_on": [],
         "conditional": "if beta is quiet",
         "commitments": [
             {"date": "2026-08-03", "as_of": "2026-06-22",
              "quote": "release August 3 if beta's quiet", "speaker": "Alex",
              "reason": None},
         ],
         "evidence": [{"field": "end",
                       "quote": "release August 3 if beta's quiet",
                       "speaker": "Alex", "meeting_date": "2026-06-22"}]},
    ],
}

TODAY = datetime.date(2026, 7, 21)


def _wb(plan=None, history=None):
    from app.services.doc_templates import render_gantt_detailed
    return openpyxl.load_workbook(io.BytesIO(render_gantt_detailed(
        plan or _PLAN, today=TODAY, history=history)))


def _slip_rows(ws):
    return {str(ws.cell(r, 1).value): r for r in range(5, 40)
            if ws.cell(r, 1).value}


# --- slip from spoken commitments ---------------------------------------

def test_spoken_moves_are_tracked_with_no_plan_history_at_all():
    from app.services.doc_templates import _compute_slip
    rows = {r["task"]["name"]: r for r in _compute_slip(_PLAN["tasks"], [])}
    p = rows["Payments integration"]
    assert str(p["baseline"]) == "2026-07-10"       # what was FIRST said
    assert str(p["current"]) == "2026-07-24"
    assert p["moves"] == 2
    assert p["first_tracked"] is False              # history is not the source
    assert p["spoken_any"] is True


def test_a_date_stated_once_is_not_a_move():
    from app.services.doc_templates import _compute_slip
    rows = {r["task"]["name"]: r for r in _compute_slip(_PLAN["tasks"], [])}
    s = rows["Offline sync"]
    assert s["moves"] == 0
    assert s["baseline"] == s["current"]


def test_spoken_and_planned_merge_without_double_counting():
    """A version whose end date matches what was said in that same meeting
    is the same event, not a second move."""
    from app.services.doc_templates import _compute_slip
    history = [
        {"as_of": "2026-06-22", "tasks": [
            {"id": 9, "name": "payments integration", "type": "task",
             "start": "2026-06-22", "end": "2026-07-10"}]},
        {"as_of": "2026-07-06", "tasks": [
            {"id": 9, "name": "Payments Integration", "type": "task",
             "start": "2026-06-22", "end": "2026-07-17"}]},
    ]
    rows = {r["task"]["name"]: r
            for r in _compute_slip(_PLAN["tasks"], history)}
    p = rows["Payments integration"]
    assert p["moves"] == 2                      # not 4
    assert str(p["baseline"]) == "2026-07-10"


def test_plan_history_still_counts_when_nobody_spoke_a_date():
    from app.services.doc_templates import _compute_slip
    silent = {**_PLAN, "tasks": [
        {**t, "commitments": []} for t in _PLAN["tasks"]]}
    history = [{"as_of": "2026-06-22", "tasks": [
        {"id": 9, "name": "Payments integration", "type": "task",
         "start": "2026-06-22", "end": "2026-07-10"}]}]
    rows = {r["task"]["name"]: r for r in _compute_slip(silent["tasks"], history)}
    p = rows["Payments integration"]
    assert p["moves"] == 1 and str(p["baseline"]) == "2026-07-10"
    assert p["spoken_any"] is False


def test_slip_sheet_reads_as_variance_and_carries_the_reason():
    sl = _wb()["Slip"]
    assert sl.cell(4, 3).value == "First committed"
    assert sl.cell(4, 7).value == "Variance (days)"
    pr = _slip_rows(sl)["Payments integration"]
    assert sl.cell(pr, 3).value.date() == datetime.date(2026, 7, 10)
    assert sl.cell(pr, 6).value == 2
    trail = str(sl.cell(pr, 8).value)
    assert "said Jun 22" in trail
    assert "sandbox creds took a week" in trail   # the WHY, when stated


def test_early_variance_is_not_styled_as_a_defect():
    """A task that came in reads green, not the red that means late."""
    sl = _wb()["Slip"]
    fonts = [rule.dxf.font.color.rgb
             for rng in sl.conditional_formatting for rule in rng.rules
             if rule.dxf is not None and rule.dxf.font is not None]
    assert "FF9A1B12" in fonts        # late
    assert "FF1E7A3C" in fonts        # early


# --- receipts -----------------------------------------------------------

def test_receipts_never_shows_a_schema_key():
    rc = _wb()["Receipts"]
    labels = {str(rc.cell(r, 3).value) for r in range(5, 20)
              if rc.cell(r, 3).value}
    assert labels <= {"Due date", "Progress", "Effort estimate"}
    assert not any("_" in x for x in labels)


def test_receipts_marks_a_superseded_quote():
    rc = _wb()["Receipts"]
    by_quote = {str(rc.cell(r, 7).value): r for r in range(5, 20)
                if rc.cell(r, 7).value}
    old = next(r for q, r in by_quote.items() if "by July 10" in q)
    new = next(r for q, r in by_quote.items() if "the 24th now" in q)
    assert rc.cell(old, 5).value == "superseded"
    assert rc.cell(new, 5).value == "current"


def test_receipts_carries_the_meeting_date():
    rc = _wb()["Receipts"]
    dates = {str(rc.cell(r, 4).value) for r in range(5, 20)
             if rc.cell(r, 4).value}
    assert "2026-06-22" in dates and "2026-07-13" in dates


# --- conditional dates --------------------------------------------------

def test_conditional_date_is_marked_where_the_date_lives():
    gv = _wb()["Gantt View"]
    # task rows only: the row 2 summary strip names the next milestone too
    rows = {str(gv.cell(r, 2).value).replace("🏁", "").strip(): r
            for r in range(4, gv.max_row + 1) if gv.cell(r, 2).value}
    rel = next(r for name, r in rows.items() if "General release" in name)
    end = gv.cell(rel, 6)
    assert end.comment is not None
    assert "if beta is quiet" in end.comment.text
    assert end.fill.start_color.rgb == "FFFBF0D5"
    # the task NAME stays clean: slip matches on it across runs
    assert "conditional" not in str(gv.cell(rel, 2).value).lower()


def test_unconditional_dates_carry_no_marker():
    gv = _wb()["Gantt View"]
    # task rows only: the row 2 summary strip names the next milestone too
    rows = {str(gv.cell(r, 2).value).replace("🏁", "").strip(): r
            for r in range(4, gv.max_row + 1) if gv.cell(r, 2).value}
    sync = next(r for name, r in rows.items() if "Offline sync" in name)
    assert gv.cell(sync, 6).comment is None


# --- s-curve ------------------------------------------------------------

def test_scurve_axis_formats_dates_instead_of_serials():
    """The defect in the shipped 2026-07-30 workbook: the category axis
    had no format, so Excel drew 46199 where the table said Jun 26."""
    from app.services.doc_templates import render_gantt_detailed
    blob = render_gantt_detailed(_PLAN, today=TODAY, history=[
        {"as_of": "2026-06-22", "tasks": _PLAN["tasks"]}])
    with zipfile.ZipFile(io.BytesIO(blob)) as z:
        chart = next(n for n in z.namelist() if "charts/chart" in n)
        xml = z.read(chart).decode()
    cat = re.search(r"<catAx>.*?</catAx>", xml, re.S)
    assert cat, "no category axis in the chart"
    assert 'formatCode="mmm d"' in cat.group(0)


def test_scurve_separates_observations_from_the_carry():
    wb = _wb(history=[{"as_of": "2026-06-22", "tasks": _PLAN["tasks"]}])
    sc = wb["S-Curve"]
    assert sc.cell(4, 5).value == "At a meeting"
    reported = [sc.cell(r, 4).value for r in range(5, 5 + 12)]
    observed = [sc.cell(r, 5).value for r in range(5, 5 + 12)]
    # the line is carried forward on every week, the dots are not
    assert len([v for v in reported if v is not None]) > \
        len([v for v in observed if v is not None])
    assert any(v is not None for v in observed)


def test_scurve_discloses_what_the_weighting_is_based_on():
    sc = _wb()["S-Curve"]
    note = str(sc["A3"].value)
    assert "scheduled working days" in note
    assert "1 of 2 tasks" in note      # only Offline sync stated an effort


def test_determinism_holds_with_commitments_in_play():
    from app.services.doc_templates import render_gantt_detailed
    a = render_gantt_detailed(_PLAN, today=TODAY)
    b = render_gantt_detailed(_PLAN, today=TODAY)
    assert a == b


# --- rebuilds of the same meeting ---------------------------------------

def test_a_rebuild_of_the_same_meeting_is_not_a_move():
    """Regenerating a plan must not read its own previous attempt as
    history. Live 2026-07-30: three tasks showed moves that were nothing
    but extraction wobble between two builds of the July 20 standup."""
    from app.services.doc_templates import _compute_slip
    prior_build = [{"as_of": "2026-07-20", "tasks": [
        {"id": 9, "name": "Offline sync", "type": "task",
         "start": "2026-07-06", "end": "2026-07-24"}]}]
    rows = {r["task"]["name"]: r for r in _compute_slip(
        _PLAN["tasks"], prior_build, as_of=_PLAN["meeting_date"])}
    s = rows["Offline sync"]
    assert s["moves"] == 0
    assert str(s["current"]) == "2026-07-22"


def test_a_genuinely_earlier_meeting_still_counts():
    from app.services.doc_templates import _compute_slip
    # Jul 06, a meeting of its own: not the Jun 29 one Jordan spoke at
    # (spoken wins there) and not a rebuild of Jul 20
    earlier = [{"as_of": "2026-07-06", "tasks": [
        {"id": 9, "name": "Offline sync", "type": "task",
         "start": "2026-07-06", "end": "2026-07-29"}]}]
    rows = {r["task"]["name"]: r for r in _compute_slip(
        _PLAN["tasks"], earlier, as_of=_PLAN["meeting_date"])}
    s = rows["Offline sync"]
    # spoken Jul 22, then that version's Jul 29, then back to Jul 22: two
    # real changes, and the version is what makes the middle one visible
    assert s["moves"] == 2
    assert [str(i["date"]) for i in s["trail"]] == [
        "2026-07-22", "2026-07-29", "2026-07-22"]


def test_the_rendered_sheet_passes_its_own_meeting_date_through():
    wb = _wb(history=[{"as_of": "2026-07-20", "tasks": [
        {"id": 9, "name": "Offline sync", "type": "task",
         "start": "2026-07-06", "end": "2026-07-24"}]}])
    sl = wb["Slip"]
    r = _slip_rows(sl)["Offline sync"]
    assert sl.cell(r, 6).value == 0


# --- receipts as the union of both quote sources ------------------------

_LIVE_SHAPE = {
    "project": "Field Kit", "meeting_date": "2026-07-20",
    "tasks": [
        {"id": 7, "name": "Release 1.2", "type": "phase", "parent_id": None,
         "owner": None, "status": "in_progress", "start": "2026-06-22",
         "end": "2026-07-24", "depends_on": []},
        {"id": 1, "name": "Payments", "type": "task", "parent_id": 7,
         "owner": "Maya", "status": "in_progress", "start": "2026-06-22",
         "end": "2026-07-24", "depends_on": [], "percent_complete": 70,
         # what prod actually returns: the DATE lines live only in
         # commitments, evidence carries the other fields, and the two
         # never quote the same words
         "commitments": [
             {"date": "2026-07-10", "as_of": "2026-06-22",
              "quote": "payments integration should land by July 10.",
              "speaker": "Maya", "reason": None},
             {"date": "2026-07-24", "as_of": "2026-07-13",
              "quote": "realistically the 24th now.", "speaker": "Maya",
              "reason": None},
         ],
         "evidence": [
             {"field": "percent_complete", "speaker": "Maya",
              "quote": "I'd say payments is 70 percent", "meeting_date": "2026-07-20"},
         ]},
    ],
}


def test_receipts_keep_the_line_behind_a_due_date():
    """The regression this fixture exists for: date provenance lives only
    in commitments, so an evidence-only Receipts sheet loses it."""
    rc = _wb(plan=_LIVE_SHAPE)["Receipts"]
    quotes = [str(rc.cell(r, 7).value) for r in range(5, 20)
              if rc.cell(r, 7).value]
    assert any("by July 10" in q for q in quotes)
    assert any("the 24th now" in q for q in quotes)
    assert any("70 percent" in q for q in quotes)


def test_standing_comes_from_the_date_not_the_wording():
    rc = _wb(plan=_LIVE_SHAPE)["Receipts"]
    by_quote = {str(rc.cell(r, 7).value): r for r in range(5, 20)
                if rc.cell(r, 7).value}
    old = next(r for q, r in by_quote.items() if "by July 10" in q)
    new = next(r for q, r in by_quote.items() if "the 24th now" in q)
    assert rc.cell(old, 5).value == "superseded"   # not the current end
    assert rc.cell(new, 5).value == "current"      # equals the current end
    assert rc.cell(old, 3).value == "Due date"
    assert rc.cell(old, 4).value == "2026-06-22"   # the meeting it was said in


def test_a_quote_in_both_fields_is_not_listed_twice():
    phase, payments = _LIVE_SHAPE["tasks"]
    plan = {**_LIVE_SHAPE, "tasks": [phase, {
        **payments,
        # the same line filed under BOTH fields
        "evidence": [{"field": "end", "speaker": "Maya",
                      "quote": "realistically the 24th now.",
                      "meeting_date": "2026-07-13"}],
    }]}
    rc = _wb(plan=plan)["Receipts"]
    quotes = [str(rc.cell(r, 7).value) for r in range(5, 20)
              if rc.cell(r, 7).value]
    assert sum(1 for q in quotes if "the 24th now" in q) == 1


def test_a_plan_whose_tasks_miss_their_phase_still_renders():
    """openpyxl rejects the inverted status-dot range with a bare
    TypeError, which reaches the user as a failed build."""
    from app.services.doc_templates import render_gantt_detailed
    orphaned = {"project": "Field Kit", "meeting_date": "2026-07-20",
                "tasks": [{"id": 1, "name": "Payments", "type": "task",
                           "parent_id": 99, "owner": "Maya",
                           "status": "in_progress", "start": "2026-06-22",
                           "end": "2026-07-24", "depends_on": []}]}
    assert render_gantt_detailed(orphaned, today=TODAY)
