"""Gantt scheduling runs on working days, not calendar days.

The grid always greyed weekends out (a `WEEKDAY()>5` rule) while the date
arithmetic counted calendar days. The sheet therefore contradicted itself the
moment anyone edited it: pushing a predecessor could land a dependent's Start
on a Saturday, and the bar would then be drawn straight across the weekend the
same screen had just shaded as non-working.

Weekends only. Holidays are locale-specific and a standup never states them,
so inventing a calendar would break the only-what-the-meeting-knows rule.
"""

from __future__ import annotations

import datetime
import io

import openpyxl

from app.services.doc_templates import (
    _snap_span,
    _with_workdays,
    _workday_formula,
    _workday_offset,
    render_gantt,
)

D = datetime.date.fromisoformat


def _rows(gv):
    out = {}
    for r in range(1, gv.max_row + 1):
        v = gv.cell(r, 2).value
        if isinstance(v, str) and v.strip():
            out[v.strip()] = r
    return out


# --- helpers --------------------------------------------------------------

def test_snap_pulls_a_span_inward_onto_working_days():
    # Sat start -> Mon; Sun end -> Fri
    assert _snap_span(D("2026-07-11"), D("2026-07-17")) == (D("2026-07-13"), D("2026-07-17"))
    assert _snap_span(D("2026-07-06"), D("2026-07-12")) == (D("2026-07-06"), D("2026-07-10"))
    # weekday spans are untouched, which keeps the extracted dates exact
    assert _snap_span(D("2026-07-06"), D("2026-07-10")) == (D("2026-07-06"), D("2026-07-10"))


def test_a_weekend_only_span_collapses_instead_of_inverting():
    # Sat..Sun would snap to Mon..Fri-before, i.e. end < start. One day wins.
    s, e = _snap_span(D("2026-07-11"), D("2026-07-12"))
    assert s == e == D("2026-07-13")


def test_workday_offset_is_the_k_that_excel_would_need():
    # Tue 21st -> Mon 27th is 6 calendar days but 4 working days
    assert _workday_offset(D("2026-07-21"), D("2026-07-27")) == 4
    assert _workday_offset(D("2026-07-10"), D("2026-07-15")) == 3
    assert _workday_offset(D("2026-07-15"), D("2026-07-15")) == 0
    assert _workday_offset(D("2026-07-15"), D("2026-07-10")) == -3


def test_zero_offset_stays_a_bare_reference():
    # The initial value has to equal the extracted date byte for byte, and
    # WORKDAY(x, 0) is not a guaranteed identity across viewers.
    assert _workday_formula("E9", 0) == "E9"
    assert _workday_formula("E9", 4) == "WORKDAY(E9,4)"
    assert _workday_formula("F9", -2) == "WORKDAY(F9,-2)"


def test_with_workdays_does_not_mutate_the_caller():
    data = {"tasks": [{"id": 1, "name": "T", "start": "2026-07-11",
                       "end": "2026-07-12"}]}
    out = _with_workdays(data)
    assert data["tasks"][0]["start"] == "2026-07-11", "input must be untouched"
    assert out["tasks"][0]["start"] == "2026-07-13"


def test_with_workdays_passes_through_unparseable_dates():
    # A malformed span must not take the whole render down; the existing
    # date handling downstream already raises where it matters.
    data = {"tasks": [{"id": 1, "name": "T", "start": "not-a-date",
                       "end": "2026-07-12"}]}
    assert _with_workdays(data)["tasks"][0]["start"] == "not-a-date"


# --- rendered output ------------------------------------------------------

_PLAN = {
    "project": "Weekend Test",
    "tasks": [
        {"id": 1, "name": "Phase", "type": "phase", "parent_id": None,
         "owner": None, "status": "in_progress", "start": "2026-07-06",
         "end": "2026-07-24", "depends_on": []},
        # ends on a Friday
        {"id": 2, "name": "Design", "type": "task", "parent_id": 1,
         "owner": None, "status": "complete", "start": "2026-07-06",
         "end": "2026-07-10", "depends_on": []},
        # extracted start is the following Monday: 1 working day after Friday
        {"id": 3, "name": "Build", "type": "task", "parent_id": 1,
         "owner": None, "status": "in_progress", "start": "2026-07-13",
         "end": "2026-07-17", "depends_on": [2]},
        # a span the model put on a Saturday
        {"id": 4, "name": "Weekend task", "type": "task", "parent_id": 1,
         "owner": None, "status": "not_started", "start": "2026-07-18",
         "end": "2026-07-24", "depends_on": []},
    ],
}


def _gv(plan):
    blob = render_gantt(plan, today=D("2026-07-13"))
    return openpyxl.load_workbook(io.BytesIO(blob))["Gantt View"]


def test_dependent_start_moves_in_working_days():
    gv = _gv(_PLAN)
    rows = _rows(gv)
    # Fri 10th -> Mon 13th is 3 calendar days and 1 working day
    assert gv.cell(rows["Build"], 5).value == f"=WORKDAY(F{rows['Design']},1)"


def test_duration_is_counted_in_working_days():
    gv = _gv(_PLAN)
    r = _rows(gv)["Build"]
    # Mon 13th -> Fri 17th: 4 calendar days, 4 working days (same here), and
    # the formula must be WORKDAY so a push keeps the length in work terms
    assert gv.cell(r, 6).value == f"=WORKDAY(E{r},4)"


def test_a_weekend_start_is_rendered_on_a_weekday():
    gv = _gv(_PLAN)
    r = _rows(gv)["Weekend task"]
    start = gv.cell(r, 5).value
    assert isinstance(start, datetime.datetime)
    assert start.date() == D("2026-07-20"), "Sat 18th should render as Mon 20th"
    assert start.weekday() < 5


def test_initial_values_still_match_the_extracted_dates_for_weekday_spans():
    """The live-grid promise: nothing looks different until someone edits."""
    gv = _gv(_PLAN)
    r = _rows(gv)["Design"]
    assert gv.cell(r, 5).value.date() == D("2026-07-06")
    assert gv.cell(r, 6).value.date() == D("2026-07-10")


def test_every_rendered_date_lands_on_a_working_day():
    gv = _gv(_PLAN)
    for r in range(1, gv.max_row + 1):
        for c in (5, 6):
            v = gv.cell(r, c).value
            if isinstance(v, datetime.datetime):
                assert v.weekday() < 5, f"row {r} col {c} is a weekend: {v}"


def test_slip_ignores_a_weekend_snap_in_the_snapshot_history():
    """Snapshots predate this change, so a historical Saturday end must be
    put through the same rule. Otherwise the first regeneration after the
    change reports a weekend-to-Friday snap as real slip."""
    from app.services.doc_templates import _compute_slip

    tasks = [{"id": 1, "name": "Build", "type": "task",
              "start": "2026-07-13", "end": "2026-07-17"}]
    history = [{"as_of": "2026-07-06", "tasks": [
        {"id": 1, "name": "Build", "type": "task",
         "start": "2026-07-13", "end": "2026-07-18"},  # a Saturday
    ]}]
    row = _compute_slip(tasks, history)[0]
    assert row["baseline"] == D("2026-07-17"), "Sat 18th baseline snaps to Fri 17th"
    assert row["moves"] == 0
    assert row["current"] == row["baseline"]
