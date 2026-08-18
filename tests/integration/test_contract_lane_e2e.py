"""The contract lane, driven end to end through the real router.

WHY THIS FILE EXISTS. The contract lane shipped with a suite of tests
that asserted source TEXT ("assert '...' in CHAT_SRC"). Those prove a
line was written. They cannot prove the path runs, and five separate
defects reached a live device in a row, each surfacing only once the
previous fix let a run travel far enough to hit it:

  1. the explicit-command fast path minted no offer, so the lane was
     unreachable from the phrasings most likely to want it
  2. the routing value kept its "anthropic/" prefix, so the provider
     answered 400 "The requested model is not available"
  3. a string `readme` raised inside the renderer
  4. `should_stream` did not exclude the lane, so a build took the chat
     stream: a 180s cap against a served 170s promise, and no
     generation events, so the client showed no build at all
  5. `raw_response_json` is a JSON string and the lane read it as a
     dict, so a finished 40-scenario workbook was thrown away

Every one of those is caught by driving a confirmed turn through the
real app with only the provider stubbed. That is what this file does.
The rule it encodes: the contract lane is exercised, not read.
"""

from __future__ import annotations

import asyncio
import io
import json
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest

from app.models.chat import ChatResponse

# The wire shape the Anthropic adapter actually hands back: `content` is
# a list of blocks, and the whole thing arrives as a STRING on
# ChatResponse.raw_response_json. Copied from a real 2026-08-16 response
# (sonnet-4-6, 183s, $0.1722, stop_reason tool_use) and trimmed.
def _emit_artifact_response(sheets: list[dict] | None = None,
                            readme: object = None) -> ChatResponse:
    payload = {
        "id": "msg_01",
        "stop_reason": "tool_use",
        "content": [{
            "type": "tool_use",
            "id": "toolu_01",
            "name": "emit_artifact",
            "input": {
                "filename": "cigna_demo_test_scenarios.xlsx",
                "title": "Cigna Demo Test Scenarios",
                "needs_computation": False,
                "readme": readme if readme is not None else {
                    "purpose": "Scenarios for the demo build.",
                    "scope": "P1 intents only.",
                    "reviewers": ["Mike"],
                },
                "sheets": sheets if sheets is not None else [
                    {"name": "Eligibility", "rows": [
                        {"id": "ELG-01", "scenario": "Active member",
                         "type": "Happy path", "preconditions": "None",
                         "test_data": "CIG100001",
                         "expected": "Eligible - coverage returned",
                         "notes": "n/a"},
                        {"id": "ELG-02", "scenario": "Termed member",
                         "type": "Failure mode", "preconditions": "None",
                         "test_data": "CIG100002",
                         "expected": "Not eligible", "notes": ""},
                    ]},
                    {"name": "Refill - Caregiver", "rows": [
                        {"id": "RFL-01", "scenario": "Caregiver refill",
                         "type": "Edge case", "preconditions": "Linked",
                         "test_data": "CIG100003",
                         "expected": "Refill accepted", "notes": ""},
                    ]},
                ],
            },
        }],
        "usage": {"input_tokens": 4000, "output_tokens": 10255},
    }
    return ChatResponse(
        text="",
        input_tokens=4000,
        output_tokens=10255,
        model="claude-sonnet-4-6",
        provider="anthropic",
        usage={"input_tokens": 4000, "output_tokens": 10255},
        raw_request_json="{}",
        # THE POINT: a string, not a dict. Defect 5 lived here.
        raw_response_json=json.dumps(payload),
    )


@pytest.fixture
def contract_lane(client, pro_user):
    """Turn the canary on for this user, and make the classifier name
    the artifact without a network call."""
    from app.main import app

    cfg = app.state.remote_configs
    docs = cfg.setdefault("client-config", {}).setdefault("documents", {})
    gen = docs.setdefault("generation", {})
    before = gen.get("contract_lane_users")
    gen["contract_lane_users"] = [pro_user["user_id"]]

    intent = {"file_request": True, "format": "xlsx", "gist": "test scenarios",
              "artifact": "test_plan", "artifact_confidence": "high"}
    with patch("app.services.document_generation.classify_generation_intent",
               new_callable=AsyncMock, return_value=intent):
        yield client
    if before is None:
        gen.pop("contract_lane_users", None)
    else:
        gen["contract_lane_users"] = before


def _send(client, pro_user, *, stream: bool = False, text: str | None = None,
          **meta):
    """One turn, as iOS sends it.

    Note what is NOT set: `generation_confirmed`. An explicit ask
    ("create a test plan spreadsheet") arms the build on this very turn
    through the fast path, which is the road defect 1 made unreachable
    and the one most real asks take. The confirm round trip is covered
    separately below.
    """
    return client.post(
        "/v1/chat",
        json={
            "provider": "auto",
            "model": "auto",
            "user_content": text or "create a test plan spreadsheet from the meeting",
            # The client assembles the prompt; PostMeetingChat sends one.
            "system_prompt": "You are helping with a meeting.",
            "stream": stream,
            "metadata": {
                "prompt_mode": "PostMeetingChat",
                "call_type": "meeting_chat",
                **meta,
            },
        },
        headers=pro_user["headers"],
    )


def _events(resp) -> list[tuple[str, dict]]:
    """A confirmed build is answered as SSE on EVERY surface, streaming
    or not (handoff Part 2): started -> progress -> result. Parsing it
    is part of what the client has to do, so the test does it too."""
    out = []
    for chunk in resp.text.split("\n\n"):
        if not chunk.strip():
            continue
        name = data = None
        for line in chunk.splitlines():
            if line.startswith("event: "):
                name = line[7:]
            elif line.startswith("data: "):
                data = line[6:]
        if name:
            out.append((name, json.loads(data) if data else {}))
    return out


def _result_of(resp) -> dict:
    for name, data in _events(resp):
        if name == "generation_result":
            return data
    raise AssertionError(
        f"no generation_result in stream: {[n for n, _ in _events(resp)]}")


def _files_from(resp) -> list[dict]:
    return _result_of(resp).get("generated_files") or []


class TestContractLaneBuildsAFile:
    def test_a_confirmed_ask_returns_a_real_workbook(self, contract_lane, pro_user):
        """The whole point, in one assertion chain: ask, get xlsx bytes
        that openpyxl can read, with the rows the model emitted."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            resp = _send(contract_lane, pro_user, stream=False)

        assert resp.status_code == 200, resp.text
        files = _files_from(resp)
        assert files, f"no file was produced: {resp.text[:600]}"

        # Files are STAGED, not inlined: the envelope carries a
        # file_id the client fetches. Pull it the way the app does, so
        # the download path is covered too.
        row = files[0]
        assert row["name"].endswith(".xlsx"), row
        got = contract_lane.get(f"/v1/generated-files/{row['file_id']}",
                                headers=pro_user["headers"])
        assert got.status_code == 200, got.text
        wb = openpyxl.load_workbook(io.BytesIO(got.content))
        assert "README" in wb.sheetnames
        rows = sum(ws.max_row - 2 for ws in wb.worksheets if ws.title != "README")
        assert rows == 3

    def test_the_model_string_reaches_the_provider_without_its_prefix(
            self, contract_lane, pro_user):
        """Defect 2: the routing config's value is "<provider>/<model>"
        and the model field goes on the wire verbatim, so an unsplit
        value asked Anthropic for "anthropic/claude-sonnet-4-6" and got
        400 back."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(contract_lane, pro_user, stream=False)

        sent = route.await_args.args[1]
        assert "/" not in sent.model, sent.model
        assert sent.provider == "anthropic"

    def test_the_turn_is_billed_as_artifact_generation(
            self, contract_lane, pro_user, tmp_db_path):
        """The dedicated call_type is what makes per-artifact cost and
        demand answerable at all. It is also the dial's only hook."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            _send(contract_lane, pro_user, stream=False)

        import sqlite3
        conn = sqlite3.connect(tmp_db_path)
        rows = conn.execute(
            "SELECT call_type, metadata FROM usage_log ORDER BY rowid DESC"
        ).fetchall()
        conn.close()
        kinds = [r[0] for r in rows]
        assert "artifact_generation" in kinds, kinds
        metas = [json.loads(r[1] or "{}") for r in rows
                 if r[0] == "artifact_generation"]
        assert any(m.get("artifact_type") == "test_plan" for m in metas), metas


class TestContractLaneTransport:
    def test_a_streamed_build_announces_itself(self, contract_lane, pro_user):
        """Defect 4. A build must not ride the interactive chat stream:
        that path caps at 180s (below the 170s we promise for this very
        contract) and emits no generation events, so the client has
        nothing to render a build from and shows a plain thinking
        indicator for three minutes."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            resp = _send(contract_lane, pro_user, stream=True)

        assert resp.status_code == 200
        body = resp.text
        assert "event: generation_started" in body, body[:400]
        started = json.loads(
            body.split("event: generation_started\ndata: ")[1].split("\n")[0])
        # The served expectation is the contract's own, not the flat 150.
        assert started["expected_seconds"] == 170, started


class TestContractLaneDegradesHonestly:
    @pytest.mark.parametrize("readme", ["a plain paragraph", None, ["a", "b"]])
    def test_an_odd_readme_shape_still_yields_a_file(
            self, contract_lane, pro_user, readme):
        """Defect 3. The contract declares readme an object, but a tool
        schema is advisory unless the tool is strict, and losing the
        whole workbook over the cover sheet is the worst trade here."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response(readme=readme)):
            resp = _send(contract_lane, pro_user, stream=False)
        assert resp.status_code == 200
        assert _files_from(resp), resp.text[:400]

    def test_no_tool_call_falls_back_to_text_rather_than_500(
            self, contract_lane, pro_user):
        """The model answering in prose is a miss, not an outage."""
        plain = ChatResponse(
            text="Here is a plan in chat instead.",
            input_tokens=10, output_tokens=10, model="claude-sonnet-4-6",
            provider="anthropic", usage={},
            raw_response_json=json.dumps(
                {"content": [{"type": "text", "text": "..."}]}),
        )
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock, return_value=plain):
            resp = _send(contract_lane, pro_user, stream=False)
        assert resp.status_code == 200
        assert not _files_from(resp)
        assert _result_of(resp).get("text")

    def test_the_lane_stays_dark_for_everyone_else(self, client, pro_user):
        """No canary entry, no contract lane — the whole reason the
        rollout is safe."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(client, pro_user, stream=False)
        if route.await_args:
            sent = route.await_args.args[1]
            assert not getattr(sent, "tools", None), (
                "contract tools armed for an unlisted user")


class TestTheOfferRoundTrip:
    """The road a phrasing without a format noun takes, and the one
    Scott used on device: ask -> "want the file?" -> "Yes" -> build.

    The artifact is chosen on the FIRST turn and has to survive onto the
    second, because the confirm send carries chat history rather than
    the original ask. That storage is the only reason the second turn
    knows which contract to use.
    """

    def test_an_unprefixed_ask_offers_first_then_builds(
            self, contract_lane, pro_user):
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            offer = _send(contract_lane, pro_user,
                          text="create a test plan from what we discussed")

        assert offer.status_code == 200
        body = offer.json()
        cta = ((body.get("feature_state") or {}).get("cta") or {})
        assert cta.get("kind") == "generation_offer", body
        offer_id = (cta.get("details") or {}).get("offer_id")
        assert offer_id, f"no offer minted: {body}"
        # Nothing was built yet, and the model was never asked to.
        assert not route.await_count or not getattr(
            route.await_args.args[1], "tools", None)

        # The reply interpreter is its own small model call; stub the
        # judgement, not the lane.
        yes = {"confirm": True, "format": "xlsx", "style": None,
               "version": None}
        with patch("app.services.document_generation.interpret_offer_reply",
                   new_callable=AsyncMock, return_value=yes), \
             patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            confirmed = _send(contract_lane, pro_user, text="Yes",
                              offer_id=offer_id, reply_text="Yes")

        assert confirmed.status_code == 200, confirmed.text
        assert _files_from(confirmed), confirmed.text[:600]


class TestAPausedBuildResumes:
    """The last of the family. Every defect on this lane was behaviour
    keyed on `request.generation`, which the contract lane deliberately
    does not set because it ships a tool schema instead of using the
    sandbox. The pause_turn continuation loop was the final one.

    It matters most for the case Scott actually wanted: research the
    topic, then build the artifact. Mixing web_search into a build turn
    means the server-side tool loop can hit its iteration limit, and a
    paused contract turn ends with no emit_artifact block at all — so
    the lane finds no tool call, falls back to text, and the user waits
    out a three minute build for nothing.
    """

    def test_a_paused_turn_is_resumed_rather_than_dropped(self):
        import asyncio
        from unittest.mock import MagicMock

        from app.models.chat import ChatRequest
        from app.services.providers.anthropic import AnthropicAdapter

        paused = {"stop_reason": "pause_turn", "content": [
            {"type": "server_tool_use", "name": "web_search", "id": "s1"}],
            "container": {"id": "cont_1"}, "usage": {}}
        finished = json.loads(_emit_artifact_response().raw_response_json)

        adapter = AnthropicAdapter.__new__(AnthropicAdapter)
        adapter.base_url = "https://example/v1/messages"
        adapter.api_key = "k"
        adapter._build_headers = lambda: {}

        posts = []

        async def fake_post(url, body, headers, timeout=None):
            posts.append(body)
            data = paused if len(posts) == 1 else finished
            return 200, data, "{}", json.dumps(data)

        adapter._post = fake_post
        req = ChatRequest(provider="anthropic", model="claude-sonnet-4-6",
                          user_content="x", system_prompt="s",
                          tools=[{"name": "emit_artifact",
                                  "input_schema": {"type": "object"}}])
        resp = asyncio.run(adapter.send_request(req))

        assert len(posts) == 2, "a paused build was not resumed"
        # The continuation reuses the container and appends the partial.
        assert posts[1].get("container") == "cont_1"
        assert posts[1]["messages"][-1]["role"] == "assistant"
        # And the finished artifact survives to the caller.
        blocks = json.loads(resp.raw_response_json)["content"]
        assert any(b.get("name") == "emit_artifact" for b in blocks)

    def test_a_plain_chat_turn_is_still_never_resumed(self):
        """The loop is for builds. An ordinary turn that somehow pauses
        must not be silently re-billed."""
        from app.models.chat import ChatRequest
        from app.services.providers.anthropic import AnthropicAdapter as A

        assert A._is_build(ChatRequest(
            provider="anthropic", model="m", user_content="x")) is False


def test_all_three_build_lanes_are_recognised_as_builds():
    """The adapter's build test decides two things: the 400s timeout and
    whether a paused turn resumes. Each lane signals differently, and a
    lane the adapter cannot recognise silently gets a chat turn's
    treatment — which is how the contract lane ended up on a 180s cap
    against its own 170s promise.

      sandbox   `generation=True`
      contract  a tool schema
      template  neither, so the router marks it
    """
    from app.models.chat import ChatRequest
    from app.services.providers.anthropic import AnthropicAdapter as A

    chat = ChatRequest(provider="anthropic", model="m", user_content="x")
    sandbox = ChatRequest(provider="anthropic", model="m", user_content="x",
                          generation=True)
    contract = ChatRequest(provider="anthropic", model="m", user_content="x",
                           tools=[{"name": "emit_artifact"}])
    template = ChatRequest(provider="anthropic", model="m", user_content="x",
                           metadata={"build_lane": "template"})

    assert A._is_build(chat) is False
    for req in (sandbox, contract, template):
        assert A._is_build(req) is True


def test_the_template_lane_marks_itself():
    """The mark has to be set where the lane arms, or the flag above is
    decoration."""
    import pathlib
    src = (pathlib.Path(__file__).resolve().parents[2]
           / "app/routers/chat.py").read_text()
    arm = src.index("elif _gen_armed and _template_id:")
    assert '"build_lane": "template"' in src[arm:arm + 900]


class TestResearchIntentSurvivesTheOffer:
    """A researched file ask goes: turn 1 mints "want the file?", turn 2
    builds. The search flag rides turn 1. If it does not reach turn 2,
    the user gets a file with no research in it and nothing says so —
    failure shaped exactly like success.

    The offer already remembers ask_content and images for this reason.
    Research intent belongs in the same bag.
    """

    def test_the_offer_remembers_and_the_confirm_inherits(self):
        from app.services import generation_offers

        oid = generation_offers.create(
            "u-search", "xlsx", "g", artifact_id="test_plan",
            search_enabled=True)
        assert generation_offers.peek("u-search", oid)["search_enabled"] is True
        # peek must not consume: the reply still has an offer to claim.
        assert generation_offers.take("u-search", oid)["search_enabled"] is True
        assert generation_offers.peek("u-search", oid) is None

    def test_an_ask_without_research_stays_without_it(self):
        from app.services import generation_offers
        oid = generation_offers.create("u-plain", "xlsx", "g")
        assert generation_offers.take("u-plain", oid)["search_enabled"] is False

    def test_the_flag_is_inherited_before_the_cap_gate_not_after(self):
        """The load-bearing detail. Inheriting after the gate would
        attach the search tool with no tier or monthly cap ruling on it
        and no counter movement — a cap bypass wearing a bugfix."""
        import pathlib
        src = (pathlib.Path(__file__).resolve().parents[2]
               / "app/routers/chat.py").read_text()
        inherit = src.index("search_intent_inherited_from_offer")
        # Anchor on the gate itself, not on an import of its module —
        # search_caps is also imported by an unrelated helper earlier in
        # the file, which is not the thing that must come second.
        gate = src.index("if body.get_meta(\"search_enabled\"):", inherit - 4000)
        assert inherit < gate, "inheritance must precede the search gate"
        # And it must sit inside the same request flow, not in a helper.
        assert 0 < gate - inherit < 2000

    def test_the_lane_question_carries_it_across_a_remint(self):
        """An ambiguous plan ask re-mints the offer to ask which version.
        The intent has to survive that hop too, or research is lost by
        the user answering a question we asked."""
        import pathlib
        src = (pathlib.Path(__file__).resolve().parents[2]
               / "app/routers/chat.py").read_text()
        assert src.count('_offer.get("search_enabled")') >= 2


class TestTheConfirmTurnRecoversResearchIntent:
    """VERIFIED with the client team 2026-08-17, by them reading their
    own code rather than recalling it: **the client sends
    `search_enabled: false` on the confirm turn.** Their flag resets at
    the end of every completed send, deliberately, so an armed search
    never leaks into the next question. They are right not to change it
    — a flag that quietly persists is how someone spends a capped, paid
    allowance they did not ask to spend. We are the only party that
    knows the offer and the reply are the same piece of work, so the
    recovery belongs here.

    That makes this LOAD BEARING rather than belt and braces. If it
    regresses: a researched build silently produces a file with no
    research in it, AND the estimate is 85 seconds short on precisely
    the turns that run long. Failure shaped exactly like success, twice.

    Existing coverage was a store-level unit test plus two assertions
    about SOURCE TEXT (`src.index`, `src.count`). Those prove a line was
    written, never that the path runs — the exact habit that put five
    defects on a device one at a time. This drives the real turn and
    asserts on the request the provider actually receives, because that
    is what decides whether the search tool gets attached at all.
    """

    def _confirm(self, client, pro_user, *, offer_search: bool, **meta):
        """Arrange an offer that does or does not remember research,
        then answer it the way the client really answers it."""
        from app.services import generation_offers

        oid = generation_offers.create(
            pro_user["user_id"], "xlsx", "test scenarios",
            artifact_id="test_plan", search_enabled=offer_search)

        yes = {"confirm": True, "format": "xlsx", "style": None,
               "version": None}
        with patch("app.services.document_generation.interpret_offer_reply",
                   new_callable=AsyncMock, return_value=yes), \
             patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            resp = _send(client, pro_user, text="Yes", offer_id=oid,
                         reply_text="Yes",
                         # What iOS actually puts on the wire here.
                         search_enabled=False, **meta)
        return resp, route

    def test_the_build_still_searches_though_the_client_said_false(
            self, contract_lane, pro_user):
        """Since the research leg landed, the recovered flag shows up as
        a research CALL rather than a tool on the build call, so this
        asserts on the first leg. If the recovery regresses there is no
        research leg at all and the count drops to one, which is what
        makes this still catch it."""
        resp, route = self._confirm(contract_lane, pro_user,
                                    offer_search=True)
        assert resp.status_code == 200, resp.text[:400]
        assert route.await_count == 2, (
            f"expected a research leg then a build, got {route.await_count} "
            "call(s); the confirm turn dropped the research the ask opted "
            "into and the user gets a file with no research in it")
        researched = route.await_args_list[0].args[1]
        assert researched.get_meta("search_enabled") is True
        assert not researched.tools, (
            "the research leg was handed the build tool, so it can build "
            "immediately and the boundary is not real")

    def test_the_estimate_still_covers_the_research(
            self, contract_lane, pro_user):
        """The second half of the same regression, and the one a user
        actually watches: a short promise on a long turn."""
        resp, _ = self._confirm(contract_lane, pro_user, offer_search=True)
        started = next((d for n, d in _events(resp)
                        if n == "generation_started"), None)
        assert started, [n for n, _ in _events(resp)]
        assert "search_expected_seconds" in started, started

    def test_an_offer_without_research_is_not_given_any(
            self, contract_lane, pro_user):
        """The client's false must still MEAN false when there is no
        opted-in research to recover. Otherwise the recovery is just a
        cap bypass that happens to be shaped like a bugfix."""
        resp, route = self._confirm(contract_lane, pro_user,
                                    offer_search=False)
        assert resp.status_code == 200, resp.text[:400]
        sent = route.await_args.args[1]
        assert not sent.get_meta("search_enabled"), (
            "a plain ask was silently upgraded to a paid search")


class TestResearchPlusFileSkipsTheOffer:
    """Scott, 2026-08-16, after watching it fail three times in a row.

    He asked us to find information online and put it into a test plan.
    That is one request with two parts. We answered by deferring both
    and asking a question he had already answered, and the offer turn
    never calls the main model, so the search he opted into did not run.
    Meanwhile the client showed "Searching the web..." off the toggle,
    which said it had. Three sends, three offers, zero searches, counter
    unmoved at 3.

    Turning search on spends from a monthly cap, so it is deliberate
    rather than a default. Paired with a file request there is nothing
    left to confirm.
    """

    def _armed_tools(self, client, pro_user, **meta):
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(client, pro_user,
                  text="find what typically breaks these bots online and put "
                       "it into a complete test plan",
                  **meta)
        if not route.await_args:
            return None
        return getattr(route.await_args.args[1], "tools", None)

    def test_search_plus_file_builds_instead_of_asking(
            self, contract_lane, pro_user):
        tools = self._armed_tools(contract_lane, pro_user, search_enabled=True)
        assert tools, "a research-and-build ask still only got an offer"
        assert any(t.get("name") == "emit_artifact" for t in tools), tools

    def test_the_same_ask_without_search_still_offers_first(
            self, contract_lane, pro_user):
        """The opt-in is what removes the ambiguity. Without it this is
        an ordinary file ask and the confirmation still earns its place."""
        assert not self._armed_tools(contract_lane, pro_user)

    def test_the_build_is_not_charged_a_second_classifier_call(
            self, contract_lane, pro_user):
        """The research path already ran the classifier, so re-asking it
        which artifact this is would be paying twice for one answer."""
        from app.services import document_generation as dg

        intent = {"file_request": True, "format": "xlsx", "gist": "g",
                  "artifact": "test_plan", "artifact_confidence": "high"}
        with patch.object(dg, "classify_generation_intent",
                          new_callable=AsyncMock, return_value=intent) as clf, \
             patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            _send(contract_lane, pro_user,
                  text="find what breaks these bots online and put it into a "
                       "complete test plan",
                  search_enabled=True)
        assert clf.await_count <= 1, (
            f"classifier called {clf.await_count}x for one answer")


class TestTheEstimateCoversTheResearch:
    """Every per-artifact estimate we hold was measured on a build
    ALONE. A research-backed build searches first, inside the same turn,
    and that time is invisible to those numbers.

    Live 2026-08-16: a test plan promised 170s, took 258s, and the user
    watched the counter sail past the promise wondering what had broken.
    Same artifact and model without research: 169.7s and 183.0s. With
    four searches: 255.0s.
    """

    def _started(self, client, pro_user, **meta):
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            resp = _send(client, pro_user, stream=True, **meta)
        for name, data in _events(resp):
            if name == "generation_started":
                return data
        raise AssertionError(f"no generation_started: {resp.text[:300]}")

    def test_a_plain_build_promises_the_artifact_time(
            self, contract_lane, pro_user):
        assert self._started(contract_lane, pro_user)["expected_seconds"] == 170

    def test_a_researched_build_promises_the_research_too(
            self, contract_lane, pro_user):
        from app.routers.chat import SEARCH_PHASE_SECONDS
        got = self._started(contract_lane, pro_user,
                            search_enabled=True)["expected_seconds"]
        assert got == 170 + SEARCH_PHASE_SECONDS
        # And the promise now brackets what the real turn took.
        assert 240 <= got <= 280, got

    def test_the_allowance_is_grounded_in_measurement(self):
        """A number nobody measured is a number nobody can defend."""
        import pathlib
        from app.routers.chat import SEARCH_PHASE_SECONDS
        assert 60 <= SEARCH_PHASE_SECONDS <= 120
        src = (pathlib.Path(__file__).resolve().parents[2]
               / "app/routers/chat.py").read_text()
        i = src.index("SEARCH_PHASE_SECONDS = ")
        assert "Measured" in src[i - 700:i]


class TestResearchIsItsOwnLeg:
    """Research runs as its own call, finishes, and hands its findings
    to the build. Two calls we sequence rather than one call juggling
    two tools, so the search/build boundary is known because we drew it.

    The failures this guards are all silent. A build that keeps the
    search tool can research again and the boundary stops being true. A
    research leg whose searches nobody counts spends a capped paid
    allowance with the meter stopped. A findings block that never
    reaches the build means we paid for research and threw it away.
    """

    def _run(self, client, pro_user, research_text="FINDING: rate is 4.2%"):
        from app.models.chat import ChatResponse

        research = ChatResponse(
            text=research_text, input_tokens=10, output_tokens=20,
            model="claude-sonnet-4-6", provider="anthropic",
            usage={"web_search_requests": 3},
            raw_request_json="{}", raw_response_json="{}")

        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   side_effect=[research, _emit_artifact_response()]) as route:
            resp = _send(client, pro_user, stream=True,
                         text="find what typically breaks these bots online "
                              "and put it into a complete test plan",
                         search_enabled=True)
        return resp, route

    def test_research_runs_first_and_the_build_follows(
            self, contract_lane, pro_user):
        resp, route = self._run(contract_lane, pro_user)
        assert route.await_count == 2, route.await_count
        first, second = (c.args[1] for c in route.await_args_list)
        assert not first.tools, "research leg must not be able to build"
        assert any(t["name"] == "emit_artifact" for t in (second.tools or [])), (
            "the second leg is the build and must carry the build tool")

    def test_the_build_cannot_search_again(self, contract_lane, pro_user):
        """The boundary is only real if the build cannot cross back."""
        _, route = self._run(contract_lane, pro_user)
        build = route.await_args_list[1].args[1]
        assert not build.get_meta("search_enabled"), (
            "the build kept the search tool, so it can research mid-build "
            "and the phase we just reported stops being true")

    def test_the_findings_reach_the_build(self, contract_lane, pro_user):
        _, route = self._run(contract_lane, pro_user)
        build = route.await_args_list[1].args[1]
        assert "FINDING: rate is 4.2%" in (build.system_prompt or ""), (
            "we paid for research and did not hand it to the build")
        assert "cannot search again" in (build.system_prompt or ""), (
            "the build was not told the tool is gone, so a named gap "
            "becomes something it quietly fills instead")

    def test_the_research_legs_searches_still_count(
            self, contract_lane, pro_user):
        """The meter must not stop when the feature starts working."""
        resp, _ = self._run(contract_lane, pro_user)
        state = _result_of(resp).get("search_state") or {}
        assert state.get("was_used") is True, state

    def test_a_failed_research_leg_still_delivers_the_file(
            self, contract_lane, pro_user):
        """Falls open on purpose: a research failure must not cost the
        user the document they asked for."""
        from httpx import HTTPError
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   side_effect=[HTTPError("research died"),
                                _emit_artifact_response()]):
            resp = _send(contract_lane, pro_user, stream=True,
                         text="find what breaks these bots online and put "
                              "it into a complete test plan",
                         search_enabled=True)
        assert _files_from(resp), resp.text[:400]

    def test_a_plain_build_runs_exactly_one_leg(
            self, contract_lane, pro_user):
        """No search opted into means no boundary to report and nothing
        to change: one call, as before."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(contract_lane, pro_user, stream=True,
                  text="create a test plan spreadsheet from the meeting")
        assert route.await_count == 1, route.await_count


class TestThePhaseIsObservedNotTimed:
    """SS maps `phase: "searching"` client side already, so the day we
    emit it their card retitles itself with no client release. It must
    mean a search is in flight RIGHT NOW, never a guess from a clock.

    Every test here drives REAL progress ticks: the tick interval is
    patched down and the stubbed provider is made slow enough to cross
    it. Without that a stub returns before the first tick, the event
    list is empty, and "searching not in []" passes while the server
    shouts "searching" at every user. That is a test that cannot fail,
    which is worse than no test because it is why nobody looks.
    """

    def _phases(self, client, pro_user, *, legs, delay=0.05, text=None,
                **meta):
        async def _slow(*a, **kw):
            await asyncio.sleep(delay)
            return legs.pop(0)

        with patch("app.routers.chat._PROGRESS_TICK_SECONDS", 0.01), \
             patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock, side_effect=_slow):
            resp = _send(client, pro_user, stream=True,
                         text=text or ("find what breaks these bots online "
                                       "and put it into a complete test plan"),
                         **meta)
        phases = [d.get("phase") for n, d in _events(resp)
                  if n == "generation_progress"]
        assert phases, "no progress ticks fired; this test proves nothing"
        return phases

    def _research(self):
        from app.models.chat import ChatResponse
        return ChatResponse(
            text="found things", input_tokens=1, output_tokens=1,
            model="claude-sonnet-4-6", provider="anthropic",
            usage={"web_search_requests": 2},
            raw_request_json="{}", raw_response_json="{}")

    def test_a_plain_build_never_claims_to_be_searching(
            self, contract_lane, pro_user):
        phases = self._phases(
            contract_lane, pro_user, legs=[_emit_artifact_response()],
            # Format noun, so this arms a build on this turn rather than
            # minting an offer. Without it there is no build to report a
            # phase for and the assertion would pass on an empty list.
            text="create a test plan spreadsheet from the meeting")
        assert "searching" not in phases, phases
        assert set(phases) == {"working"}, phases

    def test_a_researched_build_reports_searching_while_it_searches(
            self, contract_lane, pro_user):
        """The positive case. Ticks fire during the research leg, and
        those must say searching rather than claiming to build a file
        that nothing has started building."""
        phases = self._phases(contract_lane, pro_user, delay=0.12,
                              legs=[self._research(),
                                    _emit_artifact_response()],
                              search_enabled=True)
        assert "searching" in phases, (
            f"the card spent the search claiming to build: {phases}")

    def test_it_stops_saying_searching_once_research_returns(
            self, contract_lane, pro_user):
        """And it has to STOP. A phase that latches is the same lie
        pointing the other way."""
        phases = self._phases(contract_lane, pro_user, delay=0.12,
                              legs=[self._research(),
                                    _emit_artifact_response()],
                              search_enabled=True)
        assert phases[-1] == "working", (
            f"still claiming to search after research returned: {phases}")


class TestTheEstimateSaysWhichPartIsWhich:
    """A single number explains that the wait is long but never why.
    The client can say "usually about 1:25 searching, then 2:50
    building" only if we break the total down.

    Agreed with the client team 2026-08-17: they render this as caption
    copy and NEVER as a phase claim, because a timer-driven retitle
    would be the same fabrication we refused to emit ourselves. The
    observed phase token supersedes it later, additively, with nothing
    to rip out.
    """

    def _started(self, client, pro_user, **meta):
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            resp = _send(client, pro_user, stream=True, **meta)
        for name, data in _events(resp):
            if name == "generation_started":
                return data
        raise AssertionError(f"no generation_started: {resp.text[:300]}")

    def test_a_researched_build_breaks_its_estimate_down(
            self, contract_lane, pro_user):
        from app.routers.chat import SEARCH_PHASE_SECONDS
        got = self._started(contract_lane, pro_user, search_enabled=True)
        assert got["search_expected_seconds"] == SEARCH_PHASE_SECONDS

    def test_the_parts_cannot_disagree_with_the_total(
            self, contract_lane, pro_user):
        """Build slice is the subtraction, so it is always exact."""
        got = self._started(contract_lane, pro_user, search_enabled=True)
        build = got["expected_seconds"] - got["search_expected_seconds"]
        assert build == 170, (
            f"build slice {build} is not the artifact's own estimate")

    def test_a_plain_build_sends_no_breakdown(
            self, contract_lane, pro_user):
        """Absence is the signal that there is one expectation to
        render. Sending a 0 would invite a "0:00 searching" caption."""
        got = self._started(contract_lane, pro_user)
        assert "search_expected_seconds" not in got, got
class TestTheCountingArtifactPullsEverything:
    """topic_tracker is the only contract that needs cross-meeting
    context, and its whole job is counting. The cap that truncated it
    was OURS: CQ's limit is caller-supplied and absent by default, so
    sending one is what made them truncate. Measured on the wire
    2026-08-17, Scott's quilt reports total_available 2136 against the
    500 we asked for.

    This drives the real turn. A service-level test that calls
    quilt_dossier(limit=None) directly proves the service honours None
    and would pass happily while the lane went on sending 500, which is
    exactly what happened to the first version of it.
    """

    def _dossier_call(self, client, pro_user):
        from app.services import context_quilt as cq

        intent = {"file_request": True, "format": "xlsx",
                  "gist": "what keeps coming up",
                  "artifact": "topic_tracker", "artifact_confidence": "high"}
        dossier = AsyncMock(return_value={"meetings": [
            {"patches": [{"patch_id": "p1", "created_at": "2026-08-01",
                          "text": "pricing came up"}]}]})
        with patch("app.services.document_generation."
                   "classify_generation_intent",
                   new_callable=AsyncMock, return_value=intent), \
             patch.object(cq, "quilt_dossier", dossier), \
             patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            _send(client, pro_user, stream=True,
                  text="build a topic tracker spreadsheet",
                  project_id="PROJ-1")
        return dossier

    def test_the_lane_asks_for_the_whole_quilt(self, contract_lane, pro_user):
        d = self._dossier_call(contract_lane, pro_user)
        assert d.await_count == 1, (
            f"dossier fetched {d.await_count}x for one counting build")
        assert d.await_args.kwargs.get("limit") is None, (
            "the counting artifact asked for a capped dossier, so its "
            "counts are computed from a truncated set and the cell is "
            "confidently wrong: "
            f"limit={d.await_args.kwargs.get('limit')}")


class TestATeaserTapBuildsTheAnswer:
    """Live failure 2026-08-17, from a screen recording.

    Scott asked "Are there any documents I need to update?" That is a
    question ABOUT documents. Both real guards said so: explicit_file_ask
    matched nothing and the classifier returned file_request=false,
    twice. But a cheap vocabulary matcher still sees "documents", so the
    TEASER fired: "want this as a real file?"

    He tapped it. Generation armed, the card said "Building your file,
    usually about 2:30", and eight seconds later the card was replaced by
    a paragraph. No file, no error, nothing said.

    The cause was what the tap sent. A teaser offers the ANSWER on
    screen, but the offer is minted before that answer exists, so it
    stored the QUESTION. The tap re-ran the question with a build armed,
    and the model correctly answered it again in prose.
    """

    def test_the_offer_remembers_the_answer_it_offered(self):
        from app.services import generation_offers as go
        oid = go.create("u-teaser", "xlsx", "gist", ask_content="any docs?")
        go.attach_answer("u-teaser", oid, "Yes: the Cigna scenarios doc.")
        assert go.take("u-teaser", oid)["answer_content"] == (
            "Yes: the Cigna scenarios doc.")

    def test_attaching_to_a_dead_offer_is_harmless(self):
        """The card is already gone; this must not raise into the turn."""
        from app.services import generation_offers as go
        go.attach_answer("u-teaser", "nosuchoffer", "text")

    def test_the_build_turn_is_told_to_build_the_answer(
            self, contract_lane, pro_user):
        from app.services import generation_offers as go

        oid = go.create(pro_user["user_id"], "xlsx", "documents to update",
                        ask_content="Are there any documents I need to update?")
        go.attach_answer(pro_user["user_id"], oid,
                         "The Cigna scenarios document needs P2 and P3 "
                         "demo scripts adding.")
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(contract_lane, pro_user, text="Yes", offer_id=oid,
                  generation_confirmed=True)

        sent = route.await_args.args[1].user_content
        assert "Cigna scenarios document" in sent, (
            "the answer never reached the build turn, so the model is "
            "being asked the question again and will answer it again")
        assert "Do not answer the question again" in sent
        # The question survives as framing, not as the thing to answer.
        assert "Are there any documents I need to update?" in sent

    def test_an_offer_with_no_answer_still_uses_the_ask(
            self, contract_lane, pro_user):
        """Real file asks mint offers before any answer exists and must
        keep working exactly as before."""
        from app.services import generation_offers as go
        oid = go.create(pro_user["user_id"], "xlsx", "test scenarios",
                        artifact_id="test_plan",
                        ask_content="build me a test plan")
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(contract_lane, pro_user, text="Yes", offer_id=oid,
                  generation_confirmed=True)
        sent = route.await_args.args[1].user_content
        assert "build me a test plan" in sent
        assert "CONTENT TO PUT IN THE FILE" not in sent


class TestAConfirmedBuildNeverFailsSilently:
    """The user tapped a button and watched a progress card count up
    against a served estimate. If no file arrives, saying nothing means
    the card is replaced by prose and it reads as the app losing the
    request."""

    def _no_file_turn(self, client, pro_user):
        from app.models.chat import ChatResponse
        from app.services import generation_offers as go

        oid = go.create(pro_user["user_id"], "xlsx", "docs",
                        ask_content="Are there any documents I need to update?")
        prose = ChatResponse(
            text="Based on the meeting, the Cigna scenarios document needs "
                 "updating.",
            input_tokens=10, output_tokens=20, model="claude-sonnet-4-6",
            provider="anthropic", usage={}, raw_request_json="{}",
            raw_response_json=json.dumps(
                {"id": "m", "stop_reason": "end_turn",
                 "content": [{"type": "text", "text": "prose only"}],
                 "usage": {}}))
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock, return_value=prose):
            return _send(client, pro_user, stream=True, text="Yes",
                         offer_id=oid, generation_confirmed=True)

    def test_it_reports_that_no_file_was_produced(
            self, contract_lane, pro_user):
        result = _result_of(self._no_file_turn(contract_lane, pro_user))
        assert not result.get("generated_files")
        assert result.get("build_outcome") == "no_file", (
            "a confirmed build produced nothing and said nothing; the "
            "progress card is replaced by prose and it reads as the "
            "request being lost")

    def test_the_answer_is_still_served(self, contract_lane, pro_user):
        """The prose is still worth reading. Report the outcome, do not
        throw away the work."""
        result = _result_of(self._no_file_turn(contract_lane, pro_user))
        assert "Cigna scenarios document" in result["text"]
        assert "No file was produced" in result["text"]

    def test_a_successful_build_says_nothing_about_it(
            self, contract_lane, pro_user):
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            resp = _send(contract_lane, pro_user, stream=True)
        result = _result_of(resp)
        assert result.get("generated_files")
        assert "build_outcome" not in result, result.get("build_outcome")
        assert "No file was produced" not in (result.get("text") or "")


class TestTheTeaserTurnRemembersWhatItOffered:
    """Drives the REAL teaser turn. The tests above attach the answer by
    hand, which proves the store and the confirm path work and would
    pass happily while the router never attached anything at all. That
    is the fourth time today a test has been written one level away from
    the thing that breaks."""

    def _teaser_turn(self, client, pro_user):
        from app.main import app

        cfg = app.state.remote_configs
        gen = (cfg.setdefault("client-config", {})
                  .setdefault("documents", {}).setdefault("generation", {}))
        before = gen.get("teaser_text")
        gen["teaser_text"] = "Want this as a real file?"

        from app.models.chat import ChatResponse
        answer = ChatResponse(
            text="Yes. The Cigna scenarios document needs P2 and P3 demo "
                 "scripts adding, and Scott owns that.",
            input_tokens=10, output_tokens=20, model="claude-sonnet-4-6",
            provider="anthropic", usage={}, raw_request_json="{}",
            raw_response_json="{}")
        # The classifier says NO. That is the whole point: a question
        # about documents is not a request for one.
        no_file = {"file_request": False, "format": None, "gist": "",
                   "artifact": None, "artifact_confidence": "high"}
        try:
            with patch("app.services.document_generation."
                       "classify_generation_intent",
                       new_callable=AsyncMock, return_value=no_file), \
                 patch("app.services.anthropic_or_fallback.route_with_fallback",
                       new_callable=AsyncMock, return_value=answer):
                return _send(client, pro_user,
                             text="Are there any documents I need to update?")
        finally:
            if before is None:
                gen.pop("teaser_text", None)
            else:
                gen["teaser_text"] = before

    def test_the_teaser_offer_carries_the_answer(self, contract_lane, pro_user):
        from app.services import generation_offers as go

        resp = self._teaser_turn(contract_lane, pro_user)
        cta = ((resp.json().get("feature_state") or {}).get("cta") or {})
        assert cta.get("kind") == "generation_teaser", resp.json()
        oid = (cta.get("details") or {}).get("offer_id")
        assert oid, cta

        stored = go.peek(pro_user["user_id"], oid)
        assert stored and stored.get("answer_content"), (
            "the teaser offered to turn the answer into a file and then "
            "did not remember the answer; the tap will re-run the "
            "question and produce prose again")
        assert "Cigna scenarios document" in stored["answer_content"]

    def test_a_template_offer_still_builds_from_the_meeting(
            self, contract_lane, pro_user):
        """A lane that already knows what it is building must not have
        its source material swapped for a chat answer. Caught in review
        by the Gantt pill-tap test, pinned here so the narrowing cannot
        be undone by accident."""
        from app.services import generation_offers as go
        oid = go.create(pro_user["user_id"], "xlsx", "plan",
                        template_id="gantt_detailed",
                        ask_content="build the project plan")
        go.attach_answer(pro_user["user_id"], oid, "some chat answer")
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(contract_lane, pro_user, text="Yes", offer_id=oid,
                  generation_confirmed=True)
        sent = route.await_args.args[1].user_content
        assert "build the project plan" in sent
        assert "some chat answer" not in sent


class TestAConfirmedBuildOutranksTheConversation:
    """Live 2026-08-17, second silent failure of the day.

    Scott asked for scenarios, got the offer, and replied "just answer
    here and chat" — a decline. But that reply itself mentioned a word
    document and a workbook, so the classifier read it as a NEW file
    request and minted an offer whose stored ask WAS the decline. He
    changed his mind and tapped yes. The build armed correctly and the
    model was handed "Current question: Just answer here and chat..."
    followed by our single line "The user confirmed the file build."

    It obeyed the human sentence over the boilerplate and opened with
    "I'll answer both pieces right here in chat." No file, 77 seconds,
    nothing said. A trailing sentence in the user turn cannot outrank an
    explicit instruction earlier in the same turn.
    """

    def _armed(self, client, pro_user, ask):
        from app.services import generation_offers as go
        oid = go.create(pro_user["user_id"], "xlsx", "scenarios",
                        ask_content=ask)
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(client, pro_user, text="Yes", offer_id=oid,
                  generation_confirmed=True)
        return route.await_args.args[1]

    def test_the_build_is_told_the_decision_is_already_made(
            self, contract_lane, pro_user):
        sent = self._armed(contract_lane, pro_user,
                           "Just answer here and chat, it is two things")
        assert "FILE BUILD CONFIRMED" in (sent.system_prompt or ""), (
            "nothing in the system prompt outranks the stale 'answer in "
            "chat' sitting in the ask")

    def test_it_names_answering_in_chat_as_the_failure(
            self, contract_lane, pro_user):
        """Naming the exact failure matters: the model satisfied the
        request by answering well, which felt like success to it."""
        sent = self._armed(contract_lane, pro_user, "just answer here")
        sp = sent.system_prompt or ""
        assert "STALE" in sp
        assert "failure of this turn" in sp

    def test_the_steering_carries_no_dash_punctuation(self):
        from app.services.document_generation import BUILD_COMMITMENT_STEERING
        for ch in ("—", "–"):
            assert ch not in BUILD_COMMITMENT_STEERING


class TestTheTapIsRecordedInTheConversation:
    """Scott, 2026-08-17: a tap is a DECISION and belongs in the record.

    A tap currently renders as the original question repeated, so the
    transcript shows a question that mysteriously produced a file with
    no sign anybody chose anything. We serve the words rather than let
    the client invent them.
    """

    def _confirmed(self, client, pro_user):
        from app.services import generation_offers as go
        oid = go.create(pro_user["user_id"], "xlsx", "scenarios",
                        artifact_id="test_plan",
                        ask_content="build the scenarios")
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            return _result_of(_send(client, pro_user, stream=True, text="Yes",
                                    offer_id=oid, generation_confirmed=True))

    def test_a_tap_carries_a_decision_record(self, contract_lane, pro_user):
        rec = self._confirmed(contract_lane, pro_user).get("decision_record")
        assert rec, "the tap left no trace in the conversation"
        assert rec["kind"] == "file_build_confirmed"
        assert rec["persist"] is True, (
            "a decision that does not persist is invisible the moment the "
            "user scrolls back, which is the whole ask")

    def test_it_says_a_file_is_being_made_not_the_question_again(
            self, contract_lane, pro_user):
        rec = self._confirmed(contract_lane, pro_user)["decision_record"]
        assert "file" in rec["user_line"].lower()
        assert "Building" in rec["assistant_line"]

    def test_an_ordinary_turn_records_no_decision(
            self, contract_lane, pro_user):
        """Only a real decision goes in the record. A turn nobody
        confirmed must not claim one was made."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            resp = _send(contract_lane, pro_user, stream=True)
        assert "decision_record" not in _result_of(resp)


def test_a_contract_turn_never_falls_through_to_the_sandbox(
        contract_lane, pro_user):
    """The behavioural half of the wiring file's ordering guard.

    The sandbox branch is the catch-all, so a contract resolving after
    it would silently build freeform instead: no column schema, no
    deterministic renderer, and a workbook that looks plausible. The
    wiring test pins the source order and cannot tell a live branch
    from a dead one. This pins what the model is actually sent.
    """
    with patch("app.services.anthropic_or_fallback.route_with_fallback",
               new_callable=AsyncMock,
               return_value=_emit_artifact_response()) as route:
        _send(contract_lane, pro_user)
    sent = route.await_args.args[1]
    assert any(t["name"] == "emit_artifact" for t in (sent.tools or [])), (
        "no tool schema: the contract fell through to the sandbox")
    assert sent.generation is False, (
        "the contract turn set the sandbox flag, so it took the "
        "catch-all branch and the columns are the model's again")


class TestADeclineIsNotANewFileRequest:
    """Live 2026-08-17, and Scott's own description of it: he told us he
    did not want a file, and was then asked if he wanted a file.

    He declined with "just answer here and chat, but remember it's two
    different things, one's going to be a script that goes into a word
    document and the other is information for the new sheets". That is a
    decline. But it names a word document, a workbook and sheets,
    because people say what they are declining, so the classifier read
    the DECLINE as a fresh file request and offered again in the same
    turn. The offer it minted stored the decline as its ask, so when he
    later changed his mind the build was handed "just answer here and
    chat" as its instruction and answered in prose for 77 seconds.

    A decline that names file words is the NORMAL shape of a decline,
    not a rare phrasing.
    """

    def _reply_to_offer(self, client, pro_user, *, confirm, text):
        from app.services import generation_offers as go
        oid = go.create(pro_user["user_id"], "xlsx", "test scenarios",
                        artifact_id="test_plan",
                        ask_content="create the scenarios")
        judged = {"confirm": confirm, "format": "xlsx" if confirm else None,
                  "style": None, "version": None}
        clf = AsyncMock(return_value={
            "file_request": True, "format": "xlsx", "gist": "scenarios",
            "artifact": "test_plan", "artifact_confidence": "high"})
        with patch("app.services.document_generation.interpret_offer_reply",
                   new_callable=AsyncMock, return_value=judged), \
             patch("app.services.document_generation."
                   "classify_generation_intent", clf), \
             patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            resp = _send(client, pro_user, text=text, offer_id=oid,
                         reply_text=text)
        return resp, clf

    def test_a_decline_does_not_mint_another_offer(
            self, contract_lane, pro_user):
        resp, _ = self._reply_to_offer(
            contract_lane, pro_user, confirm=False,
            text="Just answer here and chat, but remember it is two "
                 "different things, one goes in a word document and the "
                 "other is new sheets for the existing workbook")
        body = resp.json()
        cta = ((body.get("feature_state") or {}).get("cta") or {})
        assert cta.get("kind") not in ("generation_offer", "generation_teaser"), (
            "we asked again in the same breath he said no; that is the "
            "loop he reported")

    def test_a_decline_is_not_even_re_classified(
            self, contract_lane, pro_user):
        """Cheaper and stricter: we should not spend a model call asking
        whether the message that just said no is a file request."""
        _, clf = self._reply_to_offer(
            contract_lane, pro_user, confirm=False,
            text="no thanks, just put it in the chat")
        assert clf.await_count == 0, (
            f"classifier ran {clf.await_count}x on a declined offer")

    def test_a_confirm_is_untouched(self, contract_lane, pro_user):
        """The yes road must keep working exactly as before."""
        resp, _ = self._reply_to_offer(
            contract_lane, pro_user, confirm=True, text="yes please")
        assert _files_from(resp), resp.text[:400]

    def test_a_later_independent_ask_still_offers(
            self, contract_lane, pro_user):
        """Suppression is for the declining TURN only. A fresh ask with
        no offer echo is a new conversation and still gets the lane."""
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()) as route:
            _send(contract_lane, pro_user,
                  text="create a test plan spreadsheet from the meeting")
        assert route.await_args, "a later ask was suppressed too"
        assert any(t["name"] == "emit_artifact"
                   for t in (route.await_args.args[1].tools or []))


class TestTheNoFileSentenceIsLocalizable:
    """The sentence is PERSISTED into the stored answer, which is what
    makes it durable and also what makes hardcoding it a defect: the
    client stores question and answer, so it survives scroll, relaunch,
    sync and reopening the meeting months later, and a response field
    does not. Written in English in the router it would have sat in
    English forever inside a French transcript, past the reach of both
    our own text hygiene and the client's localization.

    Flagged by the client team while drawing the boundary on the
    text-versus-field rule: text is also the thing they cannot restyle,
    translate or suppress once it is stored.
    """

    def _no_file_turn(self, client, pro_user, headers_extra=None):
        from app.models.chat import ChatResponse
        from app.services import generation_offers as go

        oid = go.create(pro_user["user_id"], "xlsx", "docs",
                        ask_content="are there any documents to update?")
        prose = ChatResponse(
            text="The Cigna scenarios document needs updating.",
            input_tokens=10, output_tokens=20, model="claude-sonnet-4-6",
            provider="anthropic", usage={}, raw_request_json="{}",
            raw_response_json=json.dumps(
                {"id": "m", "stop_reason": "end_turn",
                 "content": [{"type": "text", "text": "prose"}], "usage": {}}))
        headers = dict(pro_user["headers"])
        headers.update(headers_extra or {})
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock, return_value=prose):
            return client.post("/v1/chat", json={
                "provider": "auto", "model": "auto", "user_content": "Yes",
                "system_prompt": "You are helping with a meeting.",
                "stream": True,
                "metadata": {"prompt_mode": "PostMeetingChat",
                             "call_type": "meeting_chat", "offer_id": oid,
                             "generation_confirmed": True},
            }, headers=headers)

    def test_the_sentence_comes_from_served_config(
            self, contract_lane, pro_user):
        """Not a string literal in the router. Change the served value
        and the served value is what the user reads."""
        from app.main import app

        cfg = app.state.remote_configs
        gen = (cfg.setdefault("client-config", {})
                  .setdefault("documents", {}).setdefault("generation", {}))
        conf = gen.setdefault("confirmation", {})
        before = conf.get("no_file_text")
        conf["no_file_text"] = "SERVED SENTINEL TEXT"
        try:
            text = _result_of(self._no_file_turn(contract_lane, pro_user))["text"]
        finally:
            if before is None:
                conf.pop("no_file_text", None)
            else:
                conf["no_file_text"] = before
        assert "SERVED SENTINEL TEXT" in text, (
            "the router is still writing its own English sentence, which "
            "would persist untranslated into a stored transcript")

    def test_a_french_client_gets_the_french_variant(
            self, contract_lane, pro_user):
        """The whole point. Accept-Language picks the localized
        client-config exactly as it does for every other served string."""
        from app.main import app

        cfg = app.state.remote_configs
        fr = cfg.setdefault("client-config.fr", {})
        (fr.setdefault("documents", {}).setdefault("generation", {})
           .setdefault("confirmation", {}))["no_file_text"] = (
               "Aucun fichier n'a ete produit pour celui-ci.")
        try:
            text = _result_of(self._no_file_turn(
                contract_lane, pro_user,
                {"Accept-Language": "fr-FR,fr;q=0.9"}))["text"]
        finally:
            cfg.pop("client-config.fr", None)
        assert "Aucun fichier" in text, text[-200:]
        assert "No file was produced" not in text, (
            "a French client was served the English sentence, and it "
            "persists into their transcript that way")

    def test_an_empty_served_value_appends_nothing(
            self, contract_lane, pro_user):
        """Suppressible. An operator who blanks the string gets silence,
        not a fallback English sentence they thought they had removed."""
        from app.main import app

        cfg = app.state.remote_configs
        conf = (cfg.setdefault("client-config", {})
                   .setdefault("documents", {}).setdefault("generation", {})
                   .setdefault("confirmation", {}))
        before = conf.get("no_file_text")
        conf["no_file_text"] = ""
        try:
            result = _result_of(self._no_file_turn(contract_lane, pro_user))
        finally:
            if before is None:
                conf.pop("no_file_text", None)
            else:
                conf["no_file_text"] = before
        assert result.get("build_outcome") == "no_file", (
            "the machine-readable outcome must survive even when the "
            "human sentence is suppressed")
        assert "No file was produced" not in result["text"]


class TestTheOfferPromisesWhatTheCardWillShow:
    """The offer carried a flat `expected_seconds` of 150 for every
    artifact, while the progress card AFTER the tap used the real
    per-contract number, which runs from 35 seconds for an open
    questions log to 255 for a researched test plan.

    So the two disagreed about the same build, and the offer was the one
    the user made the decision on. Found 2026-08-18 while checking the
    client team's claim that the number was already in the payload: it
    was, and it was wrong, so a render built against it would have
    shipped 150 seconds into the sentence and looked like a fix.
    """

    def _offer(self, client, pro_user, **meta):
        with patch("app.services.anthropic_or_fallback.route_with_fallback",
                   new_callable=AsyncMock,
                   return_value=_emit_artifact_response()):
            resp = _send(client, pro_user,
                         text="create a test plan from what we discussed",
                         **meta)
        cta = ((resp.json().get("feature_state") or {}).get("cta") or {})
        assert cta.get("kind") == "generation_offer", resp.json()
        return cta.get("details") or {}

    def test_the_offer_quotes_the_artifact_estimate(
            self, contract_lane, pro_user):
        from app.services.artifact_types import CONTRACTS
        got = self._offer(contract_lane, pro_user)["expected_seconds"]
        assert got == CONTRACTS["test_plan"].expected_seconds, (
            f"offer promised {got}s; the card will show "
            f"{CONTRACTS['test_plan'].expected_seconds}s for the same build")

    def test_it_is_not_the_flat_default(self, contract_lane, pro_user):
        """The specific defect. 150 was right for nothing in particular."""
        from app.services.document_generation import _CONFIRMATION_DEFAULTS
        got = self._offer(contract_lane, pro_user)["expected_seconds"]
        assert got != _CONFIRMATION_DEFAULTS["expected_seconds"], got

    def test_a_contract_offer_never_carries_a_research_allowance(
            self, contract_lane, pro_user):
        """A contract ask WITH search arms the build on that turn rather
        than offering (#700), so an offer never needs the allowance.
        Pinned because the obvious symmetry with the card is wrong here,
        and adding it would be an untestable branch."""
        from app.services.artifact_types import CONTRACTS
        got = self._offer(contract_lane, pro_user)["expected_seconds"]
        assert got == CONTRACTS["test_plan"].expected_seconds

    def test_the_stored_artifact_and_the_promise_agree(
            self, contract_lane, pro_user):
        """Resolved once. Read twice they could differ, and the user
        would be told about one build and handed another."""
        from app.services import generation_offers as go
        from app.services.artifact_types import CONTRACTS
        details = self._offer(contract_lane, pro_user)
        stored = go.peek(pro_user["user_id"], details["offer_id"])
        assert stored["artifact_id"] == "test_plan"
        assert details["expected_seconds"] == (
            CONTRACTS[stored["artifact_id"]].expected_seconds)
