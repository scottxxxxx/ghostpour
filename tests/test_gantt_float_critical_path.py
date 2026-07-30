"""Total float and the critical path, as live formulas.

CPM backward pass on the detailed Gantt: two hidden helper columns carry the
late-start chain, the visible Float column is the difference from the task's
start, and zero float highlights as the critical path.

Live rather than baked in at generation on purpose. Static float sitting next
to live date formulas would repeat exactly the mistake working-day scheduling
fixed, a sheet that contradicts itself the moment someone edits.

The value assertions run through LibreOffice, which is the only way to
execute the formulas we ship. They skip where soffice is absent (CI), so the
shape assertions above them carry the load there.
"""

from __future__ import annotations

import csv
import datetime
import io
import shutil
import subprocess

import openpyxl
import pytest

from app.services.doc_templates import render_gantt, render_gantt_detailed

D = datetime.date.fromisoformat

# A -> B -> D is the long chain. C hangs off A and carries slack.
_PLAN = {
    "project": "CPM",
    "tasks": [
        {"id": 1, "name": "Phase", "type": "phase", "parent_id": None,
         "owner": None, "status": "in_progress", "start": "2026-07-06",
         "end": "2026-07-31", "depends_on": []},
        {"id": 2, "name": "A", "type": "task", "parent_id": 1, "owner": None,
         "status": "complete", "start": "2026-07-06", "end": "2026-07-10",
         "depends_on": []},
        {"id": 3, "name": "B", "type": "task", "parent_id": 1, "owner": None,
         "status": "in_progress", "start": "2026-07-13", "end": "2026-07-24",
         "depends_on": [2]},
        {"id": 4, "name": "C", "type": "task", "parent_id": 1, "owner": None,
         "status": "not_started", "start": "2026-07-13", "end": "2026-07-15",
         "depends_on": [2]},
        {"id": 5, "name": "D", "type": "milestone", "parent_id": 1,
         "owner": None, "status": "not_started", "start": "2026-07-27",
         "end": "2026-07-27", "depends_on": [3]},
    ],
}


def _sheet(renderer=render_gantt_detailed, plan=None):
    blob = renderer(plan or _PLAN, today=D("2026-07-20"))
    return openpyxl.load_workbook(io.BytesIO(blob))["Gantt View"]


def _rows(gv):
    out = {}
    for r in range(1, gv.max_row + 1):
        v = gv.cell(r, 2).value
        if isinstance(v, str) and v.strip():
            out[v.strip()] = r
    return out


def _row_for(gv, name):
    """Row of a task by name. Milestones carry a flag prefix in column B."""
    for label, r in _rows(gv).items():
        if label == name or label.endswith(" " + name):
            return r
    raise AssertionError(f"no row for {name!r}: {sorted(_rows(gv))}")


# --- shape ----------------------------------------------------------------

def test_float_column_is_detailed_only():
    gv = _sheet()
    assert "Float" in str(gv.cell(3, 12).value), "column L is the Float header"
    # the simple style keeps its old layout: the day grid starts at K
    simple = _sheet(render_gantt)
    assert simple.cell(3, 11).value is None or "Float" not in str(
        simple.cell(3, 11).value)


def test_float_is_a_formula_off_the_hidden_late_start():
    gv = _sheet()
    r = _row_for(gv, "C")
    f = gv.cell(r, 12).value
    assert isinstance(f, str) and f.startswith("=NETWORKDAYS("), f
    assert f"E{r}" in f, "float measures from the task's own start"


def _pf_cell(gv):
    """(address, column) of the hidden project-finish helper."""
    for c in range(13, gv.max_column + 1):
        v = gv.cell(1, c).value
        if isinstance(v, str) and v.startswith("=MAX(F"):
            return f"${gv.cell(1, c).column_letter}$1", c
    raise AssertionError("no project-finish helper found")


def test_terminal_task_hangs_off_the_project_finish():
    """A task nothing depends on is only held by the project end, so its
    late start is that date walked back by its own duration."""
    gv = _sheet()
    pf, pf_col = _pf_cell(gv)
    ls_col = pf_col - 1
    ls = gv.cell(_row_for(gv, "D"), ls_col).value
    assert isinstance(ls, str) and pf in ls, (ls, pf)
    assert "NETWORKDAYS" in ls, "duration must read live off the row's dates"
    # a task WITH a successor is held by that successor instead
    ls_b = gv.cell(_row_for(gv, "B"), ls_col).value
    assert pf not in ls_b and f"${gv.cell(1, ls_col).column_letter}$" in ls_b


def test_helper_columns_are_hidden():
    gv = _sheet()
    hidden = [k for k, v in gv.column_dimensions.items() if v.hidden]
    assert len(hidden) >= 2, f"late-start and project-finish must be hidden: {hidden}"


def test_zero_float_highlights_as_critical():
    gv = _sheet()
    ranges = [str(cf.sqref) for cf in gv.conditional_formatting]
    assert any(r.startswith("L") for r in ranges), (
        f"the Float column needs a critical-path rule: {ranges}"
    )


def test_legend_explains_what_float_means():
    gv = _sheet()
    text = " ".join(str(gv.cell(r, 2).value or "") for r in range(1, 14))
    assert "critical path" in text
    assert "recalculates" in text, "say plainly that it is live"


def test_cyclic_graph_emits_no_float_rather_than_circular_references():
    plan = {"project": "Cycle", "tasks": [
        {"id": 1, "name": "P", "type": "phase", "parent_id": None,
         "owner": None, "status": "in_progress", "start": "2026-07-06",
         "end": "2026-07-17", "depends_on": []},
        {"id": 2, "name": "X", "type": "task", "parent_id": 1, "owner": None,
         "status": "not_started", "start": "2026-07-06", "end": "2026-07-10",
         "depends_on": [3]},
        {"id": 3, "name": "Y", "type": "task", "parent_id": 1, "owner": None,
         "status": "not_started", "start": "2026-07-13", "end": "2026-07-17",
         "depends_on": [2]},
    ]}
    gv = _sheet(plan=plan)
    for name in ("X", "Y"):
        assert gv.cell(_row_for(gv, name), 12).value is None


def test_phase_and_project_rows_carry_no_float():
    gv = _sheet()
    for name in ("Phase", "CPM"):
        for key, r in _rows(gv).items():
            if key.endswith(name):
                assert gv.cell(r, 12).value is None, f"{key} is a rollup"


# --- behavior, via a real formula engine ----------------------------------

_SOFFICE = shutil.which("soffice")


@pytest.mark.skipif(_SOFFICE is None, reason="needs LibreOffice to recalculate")
def test_recalculated_float_matches_hand_computed_cpm(tmp_path):
    """A->B->D is critical (zero float). C ends 8 working days before the
    project does and nothing depends on it, so it can slip 8 days."""
    xlsx = tmp_path / "cpm.xlsx"
    xlsx.write_bytes(render_gantt_detailed(_PLAN, today=D("2026-07-20")))
    subprocess.run(
        [_SOFFICE, "--headless", "--convert-to", "csv",
         str(xlsx), "--outdir", str(tmp_path)],
        check=True, capture_output=True, timeout=240,
    )
    rows = list(csv.reader((tmp_path / "cpm.csv").open()))
    got = {}
    for r in rows:
        if len(r) < 12:
            continue
        name = (r[1] or "").strip()
        if name in ("A", "B", "C", "D") or name.endswith("D"):
            got[name[-1]] = r[11].strip()
    assert got.get("A") == "0", got
    assert got.get("B") == "0", got
    assert got.get("C") == "8", got
    assert got.get("D") == "0", got


@pytest.mark.skipif(_SOFFICE is None, reason="needs LibreOffice to recalculate")
def test_recalculated_dates_all_land_on_working_days(tmp_path):
    xlsx = tmp_path / "wd.xlsx"
    xlsx.write_bytes(render_gantt_detailed(_PLAN, today=D("2026-07-20")))
    subprocess.run(
        [_SOFFICE, "--headless", "--convert-to", "csv",
         str(xlsx), "--outdir", str(tmp_path)],
        check=True, capture_output=True, timeout=240,
    )
    seen = 0
    for r in list(csv.reader((tmp_path / "wd.csv").open())):
        if len(r) < 6:
            continue
        for cell in (r[4], r[5]):
            try:
                d = D(cell.strip())
            except ValueError:
                continue
            assert d.weekday() < 5, f"{cell} is a weekend"
            seen += 1
    assert seen >= 8, "expected several resolved dates to check"


def _recalc_floats(path, tmp_path):
    """Run a workbook through LibreOffice and read back (end, float) per task."""
    subprocess.run(
        [_SOFFICE, "--headless", "--convert-to", "csv",
         str(path), "--outdir", str(tmp_path)],
        check=True, capture_output=True, timeout=240,
    )
    csv_path = tmp_path / (path.stem + ".csv")
    out = {}
    for r in csv.reader(csv_path.open()):
        name = (r[1] or "").strip() if len(r) > 1 else ""
        if len(r) > 11 and name and len(name) < 4:
            key = name[-1]
            if key in "ABCD":
                out[key] = (r[5][:10], r[11].strip())
    return out


@pytest.mark.skipif(_SOFFICE is None, reason="needs LibreOffice to recalculate")
def test_float_recalculates_when_a_user_edits_a_duration(tmp_path):
    """The whole reason float is formulas rather than numbers.

    Durations read live off each row's own dates. Baking the extracted
    length in as a constant looked fine until you did the thing people
    actually do: overwrite a task's End by hand. Its slack was visibly
    spent while its float sat unchanged.

    C carries 8 working days of float. Push its End 4 working days later
    and 4 should remain, with the A -> B -> D chain still at zero.
    """
    import openpyxl as _o

    src = tmp_path / "before.xlsx"
    src.write_bytes(render_gantt_detailed(_PLAN, today=D("2026-07-20")))
    before = _recalc_floats(src, tmp_path)
    assert before["C"][1] == "8", before

    wb = _o.load_workbook(src)
    gv = wb["Gantt View"]
    edited = 0
    for r in range(1, gv.max_row + 1):
        v = gv.cell(r, 2).value
        if isinstance(v, str) and v.strip() == "C":
            gv.cell(r, 6).value = datetime.datetime(2026, 7, 21)
            gv.cell(r, 6).number_format = "yyyy-mm-dd"
            edited += 1
    assert edited == 1
    dst = tmp_path / "after.xlsx"
    wb.save(dst)

    after = _recalc_floats(dst, tmp_path)
    assert after["C"] == ("2026-07-21", "4"), after
    for k in ("A", "B", "D"):
        assert after[k][1] == "0", (k, after)
