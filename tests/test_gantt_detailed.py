"""Detailed gantt v1 (2026-07-21): simple-vs-detailed style choice.

The offer asks the style question until the project has a saved answer;
the reply word resolves at the template lane (single point for typed
replies AND pill taps) and persists per project. The detailed renderer
adds Progress / Workload / Receipts sheets to the identical Gantt View;
percent complete and effort are strictly blank-when-not-stated. Every
template build also writes a plan snapshot (v2 slip groundwork) and logs
its own doc_extract_* call_type.
"""

from __future__ import annotations

import re

import pytest

from tests.test_document_generation import _enable_confirmed_generation

_DPLAN = {
    "project": "Field Kit",
    "meeting_date": "2026-07-20",
    "tasks": [
        {"id": 1, "name": "Release 1.2", "type": "phase", "parent_id": None,
         "owner": None, "status": "in_progress", "start": "2026-07-06",
         "end": "2026-07-27", "depends_on": []},
        {"id": 2, "name": "Payments integration", "type": "task", "parent_id": 1,
         "owner": "Maya Chen", "status": "in_progress", "start": "2026-07-06",
         "end": "2026-07-24", "depends_on": [], "percent_complete": 70,
         "effort": None,
         "evidence": [{"field": "percent_complete",
                       "quote": "I'd say payments is 70 percent",
                       "speaker": "Maya"}]},
        {"id": 3, "name": "Offline sync", "type": "task", "parent_id": 1,
         "owner": "Jordan Lee", "status": "in_progress", "start": "2026-07-06",
         "end": "2026-07-22", "depends_on": [], "percent_complete": None,
         "effort": "3 days",
         "evidence": [{"field": "effort",
                       "quote": "probably three days of actual work",
                       "speaker": "Jordan"}]},
        {"id": 4, "name": "Crash SDK swap", "type": "task", "parent_id": 1,
         "owner": "Jordan Lee", "status": "in_progress", "start": "2026-07-13",
         "end": "2026-07-21", "depends_on": []},
        {"id": 5, "name": "Beta build", "type": "milestone", "parent_id": 1,
         "owner": None, "status": "not_started", "start": "2026-07-27",
         "end": "2026-07-27", "depends_on": [4]},
    ],
}


def test_registry_detailed_entry_never_auto_matches():
    from app.services.doc_templates import (
        STYLE_TO_TEMPLATE,
        TEMPLATES,
        match_template,
    )
    assert "gantt_detailed" in TEMPLATES
    assert TEMPLATES["gantt_detailed"]["hints"] == ()
    # the family match is always the simple entry; style swaps at arm time
    assert match_template("can you build a gantt chart of our plan?") == "gantt_smartsheet"
    assert STYLE_TO_TEMPLATE == {"simple": "gantt_smartsheet",
                                 "detailed": "gantt_detailed"}


def test_render_detailed_adds_sheets_and_stays_honest():
    import datetime
    import io

    import openpyxl

    from app.services.doc_templates import render_gantt, render_gantt_detailed
    blob = render_gantt_detailed(_DPLAN, today=datetime.date(2026, 7, 21))
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert wb.sheetnames == ["Gantt View", "Progress", "Workload", "Slip",
                             "Receipts"]
    # no history: slip sheet states that tracking starts now
    slip_texts = " ".join(str(c.value) for row in wb["Slip"].iter_rows()
                          for c in row if c.value)
    assert "History starts with this version" in slip_texts

    # simple keeps its exact layout; detailed shifts the day grid right
    # to make room for the on-view % Done / Effort columns
    simple = openpyxl.load_workbook(io.BytesIO(
        render_gantt(_DPLAN, today=datetime.date(2026, 7, 21))))
    assert simple["Gantt View"].freeze_panes == "J4"
    gv = wb["Gantt View"]
    assert gv.freeze_panes == "L4"
    # progress is ON the timeline view: stated percent + effort as columns
    gv_rows = {str(gv.cell(r, 2).value or "").strip(): r
               for r in range(1, 40) if gv.cell(r, 2).value}
    payments_r = gv_rows["Payments integration"]
    sync_r = gv_rows["Offline sync"]
    assert gv.cell(payments_r, 10).value == 0.7
    assert gv.cell(payments_r, 10).number_format == "0%"
    assert gv.cell(payments_r, 11).value is None      # effort not stated
    assert gv.cell(sync_r, 11).value == "3 days"
    assert gv.cell(sync_r, 10).value is None          # fixture states no %
    crash_r = gv_rows["Crash SDK swap"]
    assert gv.cell(crash_r, 10).value is None         # % not stated -> blank
    # live completed-portion overlay keyed to the % cell
    gv_formulas = " | ".join(
        f for rng in gv.conditional_formatting for rule in rng.rules
        for f in (rule.formula or []))
    assert "*$J" in gv_formulas

    prog = wb["Progress"]
    rows = {prog.cell(r, 1).value: r for r in range(5, 9)}
    # stated percent lands, formatted as a percent
    pr = rows["Payments integration"]
    assert prog.cell(pr, 5).value == 0.7
    assert prog.cell(pr, 5).number_format == "0%"
    # blank-when-not-stated: no percent for sync, no effort for payments
    sr = rows["Offline sync"]
    assert prog.cell(sr, 5).value is None
    assert prog.cell(sr, 6).value == "3 days"
    assert prog.cell(pr, 6).value is None
    # receipts refs cite into the Receipts sheet
    assert prog.cell(pr, 7).value == "R1"
    assert prog.cell(sr, 7).value == "R2"

    wl = wb["Workload"]
    formulas = [str(c.value) for row in wl.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    assert formulas and all("COUNTIFS(Progress!" in f for f in formulas)
    # both Jordan tasks and Maya's task produce owner rows
    owners = [wl.cell(r, 1).value for r in (5, 6)]
    assert owners == ["Jordan Lee", "Maya Chen"]
    rules = [r for rng in wl.conditional_formatting for r in rng.rules]
    assert any(r.operator == "greaterThanOrEqual" for r in rules if hasattr(r, "operator"))
    assert any("TODAY()" in f for r in rules for f in (r.formula or []))

    rc = wb["Receipts"]
    assert rc.cell(5, 1).value == "R1"
    assert "70 percent" in str(rc.cell(5, 5).value)
    assert rc.cell(5, 4).value == "Maya"
    assert rc.cell(6, 3).value == "effort"

    # determinism: same plan, same bytes
    assert render_gantt_detailed(_DPLAN, today=datetime.date(2026, 7, 21)) == blob


_HISTORY = [
    {"as_of": "2026-06-22", "created_at": "2026-06-22T18:00:00+00:00",
     "tasks": [{"id": 9, "name": "payments integration", "type": "task",
                "parent_id": 1, "owner": "Maya Chen", "status": "in_progress",
                "start": "2026-07-06", "end": "2026-07-10", "depends_on": []}]},
    {"as_of": "2026-07-06", "created_at": "2026-07-06T18:00:00+00:00",
     "tasks": [{"id": 4, "name": "Payments Integration", "type": "task",
                "parent_id": 1, "owner": "Maya Chen", "status": "in_progress",
                "start": "2026-07-06", "end": "2026-07-17", "depends_on": []}]},
]


def test_milestone_rows_carry_ignored_errors_mark():
    """Green-triangle regression (Scott 2026-07-21): the milestone glyph
    formula differs from its bar-row neighbors, so Excel's inconsistent-
    formula checker flags the whole row unless the file itself declares
    the day grid ignored. Both styles get the mark."""
    import datetime
    import io
    import zipfile

    from app.services.doc_templates import render_gantt, render_gantt_detailed
    for blob in (render_gantt(_DPLAN, today=datetime.date(2026, 7, 21)),
                 render_gantt_detailed(_DPLAN, today=datetime.date(2026, 7, 21))):
        xml = zipfile.ZipFile(io.BytesIO(blob)).read(
            "xl/worksheets/sheet1.xml").decode()
        assert "<ignoredErrors><ignoredError sqref=" in xml
        assert 'formula="1"' in xml
        # only the Gantt View sheet carries it
        z = zipfile.ZipFile(io.BytesIO(blob))
        for n in z.namelist():
            if n.startswith("xl/worksheets/") and n != "xl/worksheets/sheet1.xml":
                assert b"ignoredErrors" not in z.read(n)


@pytest.mark.asyncio
async def test_history_keeps_one_version_per_as_of(client, tmp_db_path):
    """Regenerating after the same meeting replaces that day's version;
    it must never read as schedule movement (live 2026-07-21: duplicate
    same-day snapshots + extraction jitter showed 4 phantom moves)."""
    import json as _json
    import sqlite3

    import aiosqlite

    from app.services import plan_snapshots
    con = sqlite3.connect(tmp_db_path)
    rows = [
        ("s1", "2026-06-22", "2026-06-22T10:00:00+00:00", "2026-07-10"),
        ("s2", "2026-07-20", "2026-07-20T10:00:00+00:00", "2026-07-17"),
        ("s3", "2026-07-20", "2026-07-20T11:00:00+00:00", "2026-07-24"),
    ]
    for sid, as_of, created, end in rows:
        con.execute(
            "INSERT INTO plan_snapshots VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (sid, "u-hist", "shouldersurf", "p-hist", "gantt_detailed",
             "X", as_of,
             _json.dumps([{"name": "Payments", "type": "task",
                           "end": end}]), created))
    con.commit()
    con.close()
    async with aiosqlite.connect(tmp_db_path) as db:
        hist = await plan_snapshots.history(db, user_id="u-hist",
                                            project_id="p-hist")
    assert [v["as_of"] for v in hist] == ["2026-06-22", "2026-07-20"]
    # the later build of Jul 20 won
    assert hist[1]["tasks"][0]["end"] == "2026-07-24"


def test_compute_slip_unit():
    from app.services.doc_templates import _compute_slip
    rows = {r["task"]["name"]: r for r in _compute_slip(_DPLAN["tasks"], _HISTORY)}
    p = rows["Payments integration"]   # matched despite case drift + id churn
    assert str(p["baseline"]) == "2026-07-10"
    assert p["baseline_as_of"] == "2026-06-22"
    assert str(p["current"]) == "2026-07-24"
    assert p["moves"] == 2
    assert p["first_tracked"] is False
    s = rows["Offline sync"]           # never in history: first tracked now
    assert s["first_tracked"] is True and s["moves"] == 0
    assert s["baseline"] == s["current"]


def test_render_detailed_slip_sheet_from_history():
    import datetime
    import io

    import openpyxl

    from app.services.doc_templates import render_gantt_detailed
    blob = render_gantt_detailed(_DPLAN, today=datetime.date(2026, 7, 21),
                                 history=_HISTORY)
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    sl = wb["Slip"]
    rows = {str(sl.cell(r, 1).value): r for r in range(5, 20)
            if sl.cell(r, 1).value}
    pr = rows["Payments integration"]
    assert sl.cell(pr, 3).value.date() == datetime.date(2026, 7, 10)
    assert sl.cell(pr, 4).value == "2026-06-22"
    assert sl.cell(pr, 5).value.date() == datetime.date(2026, 7, 24)
    assert sl.cell(pr, 6).value == 2
    assert sl.cell(pr, 7).value == f"=E{pr}-C{pr}"   # live slip-days formula
    assert "as of Jul 06" in str(sl.cell(pr, 8).value)
    sr = rows["Offline sync"]
    assert sl.cell(sr, 4).value == "first tracked now"
    assert sl.cell(sr, 6).value == 0
    # determinism holds with history in play
    assert render_gantt_detailed(_DPLAN, today=datetime.date(2026, 7, 21),
                                 history=_HISTORY) == blob


def test_e2e_slip_reads_project_snapshots(client, free_user, mock_provider,
                                          tmp_db_path, monkeypatch):
    import datetime
    import io
    import json as _json
    import sqlite3
    from unittest.mock import AsyncMock

    import openpyxl

    import app.services.document_generation as dg
    from tests.conftest import chat_request

    _enable_confirmed_generation(client)
    con = sqlite3.connect(tmp_db_path)
    for i, ver in enumerate(_HISTORY):
        con.execute(
            "INSERT INTO plan_snapshots VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (f"snap-{i}", free_user["user_id"], "shouldersurf", "proj-slip",
             "gantt_detailed", "Field Kit", ver["as_of"],
             _json.dumps(ver["tasks"]), ver["created_at"]))
    con.commit()
    con.close()

    cta = _offer_gantt(client, free_user, monkeypatch, project_id="proj-slip")
    monkeypatch.setattr(dg, "interpret_offer_reply", AsyncMock(
        return_value={"confirm": True, "format": "xlsx", "style": "detailed"}))
    mock_provider.canned_response.text = _json.dumps(_DPLAN)
    mock_provider.return_value = mock_provider.canned_response
    r = client.post("/v1/chat", json={
        **chat_request(prompt_mode="ProjectChat", call_type="query",
                       user_content="detailed please"),
        "project_id": "proj-slip",
        "metadata": {"offer_id": cta["details"]["offer_id"],
                     "generation_id": "gen-slip-1"},
    }, headers=free_user["headers"])
    assert r.status_code == 200
    con = sqlite3.connect(tmp_db_path)
    path = con.execute(
        "SELECT storage_path FROM generated_files WHERE user_id=?"
        " ORDER BY created_at DESC LIMIT 1",
        (free_user["user_id"],)).fetchone()[0]
    con.close()
    wb = openpyxl.load_workbook(io.BytesIO(open(path, "rb").read()))
    sl = wb["Slip"]
    rows = {str(sl.cell(r, 1).value): r for r in range(5, 20)
            if sl.cell(r, 1).value}
    pr = rows["Payments integration"]
    assert sl.cell(pr, 6).value == 2                  # two moves, from history
    assert sl.cell(pr, 3).value.date() == datetime.date(2026, 7, 10)


@pytest.mark.asyncio
async def test_interpreter_returns_style_word():
    from unittest.mock import AsyncMock, MagicMock

    from app.services.document_generation import interpret_offer_reply
    router = MagicMock()
    router.route = AsyncMock(return_value=MagicMock(
        text='{"confirm": true, "format": null, "style": "detailed"}'))
    out = await interpret_offer_reply(
        router, {"format": "xlsx", "gist": "plan"}, "detailed please",
        verbatim=True)
    assert out == {"confirm": True, "format": "xlsx", "style": "detailed"}
    # junk style values fail closed to None
    router.route = AsyncMock(return_value=MagicMock(
        text='{"confirm": true, "format": null, "style": "fancy"}'))
    out = await interpret_offer_reply(
        router, {"format": "xlsx", "gist": "plan"}, "yes", verbatim=True)
    assert out["style"] is None


def _offer_gantt(client, free_user, monkeypatch, project_id=None):
    from unittest.mock import AsyncMock

    import app.services.document_generation as dg
    from tests.conftest import chat_request
    monkeypatch.setattr(dg, "classify_generation_intent", AsyncMock(
        return_value={"file_request": True, "format": "xlsx", "gist": "of the plan"}))
    body = chat_request(prompt_mode="ProjectChat", call_type="query",
                        user_content="build a gantt chart of our project plan")
    if project_id:
        body["project_id"] = project_id
    r = client.post("/v1/chat", json=body, headers=free_user["headers"])
    return r.json()["feature_state"]["cta"]


def test_reply_style_detailed_builds_persists_meters_and_snapshots(
        client, free_user, mock_provider, tmp_db_path, monkeypatch):
    import json as _json
    import sqlite3
    from unittest.mock import AsyncMock

    import app.services.document_generation as dg
    from tests.conftest import chat_request

    _enable_confirmed_generation(client)
    cta = _offer_gantt(client, free_user, monkeypatch, project_id="proj-fk")
    assert "simple or detailed" in cta["text"]
    assert cta["details"]["expected_seconds"] == 45   # simple default pre-choice
    oid = cta["details"]["offer_id"]

    monkeypatch.setattr(dg, "interpret_offer_reply", AsyncMock(
        return_value={"confirm": True, "format": "xlsx", "style": "detailed"}))
    mock_provider.canned_response.text = _json.dumps(_DPLAN)
    mock_provider.return_value = mock_provider.canned_response
    r2 = client.post("/v1/chat", json={
        **chat_request(prompt_mode="ProjectChat", call_type="query",
                       user_content="detailed please"),
        "project_id": "proj-fk",
        "metadata": {"offer_id": oid, "generation_id": "gen-det-1"},
    }, headers=free_user["headers"])
    assert r2.status_code == 200
    result = _json.loads(
        r2.text.split("event: generation_result\ndata: ")[1].split("\n")[0])
    assert re.fullmatch(r"([A-Za-z0-9_]+_)?Gantt_Detailed_\d{6}\.xlsx",
                        result["generated_files"][0]["name"])

    con = sqlite3.connect(tmp_db_path)
    # style persisted per project
    assert con.execute(
        "SELECT value FROM project_prefs WHERE user_id=? AND project_id=?"
        " AND key='gantt_style'",
        (free_user["user_id"], "proj-fk")).fetchone()[0] == "detailed"
    # snapshot written with the resolved template
    snap = con.execute(
        "SELECT template_id, project_id, tasks_json FROM plan_snapshots"
        " WHERE user_id=?", (free_user["user_id"],)).fetchone()
    assert snap[0] == "gantt_detailed" and snap[1] == "proj-fk"
    assert "Payments integration" in snap[2]
    # extraction leg logged under its own call_type
    ct = con.execute(
        "SELECT call_type FROM usage_log WHERE user_id=?"
        " ORDER BY rowid DESC LIMIT 1", (free_user["user_id"],)).fetchone()[0]
    assert ct == "doc_extract_gantt_detailed"
    con.close()

    # next offer for the SAME project skips the question and quotes the
    # saved detailed style
    cta2 = _offer_gantt(client, free_user, monkeypatch, project_id="proj-fk")
    assert "detailed style this project uses" in cta2["text"]
    assert "simple or detailed" not in cta2["text"]
    assert cta2["details"]["expected_seconds"] == 60


def test_pill_tap_honors_saved_detailed_pref(
        client, free_user, mock_provider, tmp_db_path, monkeypatch):
    import json as _json
    import sqlite3
    from datetime import datetime, timezone

    from tests.conftest import chat_request

    _enable_confirmed_generation(client)
    con = sqlite3.connect(tmp_db_path)
    con.execute(
        "INSERT OR REPLACE INTO project_prefs VALUES (?, ?, ?, ?, ?)",
        (free_user["user_id"], "proj-pill", "gantt_style", "detailed",
         datetime.now(timezone.utc).isoformat()))
    con.commit()
    con.close()

    cta = _offer_gantt(client, free_user, monkeypatch, project_id="proj-pill")
    oid = cta["details"]["offer_id"]
    mock_provider.canned_response.text = _json.dumps(_DPLAN)
    mock_provider.return_value = mock_provider.canned_response
    r = client.post("/v1/chat", json={
        **chat_request(prompt_mode="ProjectChat", call_type="query",
                       user_content="yes"),
        "project_id": "proj-pill",
        "metadata": {"offer_id": oid, "generation_confirmed": True,
                     "generation_id": "gen-pill-1"},
    }, headers=free_user["headers"])
    assert r.status_code == 200
    result = _json.loads(
        r.text.split("event: generation_result\ndata: ")[1].split("\n")[0])
    assert "Gantt_Detailed" in result["generated_files"][0]["name"]


def test_reply_style_simple_persists_and_builds_simple(
        client, free_user, mock_provider, tmp_db_path, monkeypatch):
    import json as _json
    import sqlite3
    from unittest.mock import AsyncMock

    import app.services.document_generation as dg
    from tests.conftest import chat_request

    _enable_confirmed_generation(client)
    cta = _offer_gantt(client, free_user, monkeypatch, project_id="proj-sm")
    oid = cta["details"]["offer_id"]
    monkeypatch.setattr(dg, "interpret_offer_reply", AsyncMock(
        return_value={"confirm": True, "format": "xlsx", "style": "simple"}))
    mock_provider.canned_response.text = _json.dumps(_DPLAN)
    mock_provider.return_value = mock_provider.canned_response
    r = client.post("/v1/chat", json={
        **chat_request(prompt_mode="ProjectChat", call_type="query",
                       user_content="the simple one"),
        "project_id": "proj-sm",
        "metadata": {"offer_id": oid, "generation_id": "gen-sm-1"},
    }, headers=free_user["headers"])
    result = _json.loads(
        r.text.split("event: generation_result\ndata: ")[1].split("\n")[0])
    assert "Gantt_Detailed" not in result["generated_files"][0]["name"]

    con = sqlite3.connect(tmp_db_path)
    assert con.execute(
        "SELECT value FROM project_prefs WHERE project_id='proj-sm'"
        " AND key='gantt_style'").fetchone()[0] == "simple"
    # snapshots write on the simple lane too (slip history is style-blind)
    assert con.execute(
        "SELECT template_id FROM plan_snapshots WHERE project_id='proj-sm'"
    ).fetchone()[0] == "gantt_smartsheet"
    ct = con.execute(
        "SELECT call_type FROM usage_log WHERE user_id=?"
        " ORDER BY rowid DESC LIMIT 1", (free_user["user_id"],)).fetchone()[0]
    assert ct == "doc_extract_gantt_smartsheet"
    con.close()
