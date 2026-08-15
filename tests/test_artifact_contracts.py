"""A contracted artifact cannot ship without its load-bearing columns.

Measured 2026-08-14, three identical freeform runs per model: Sonnet 5
included an expected-result column ONCE, Opus 4.8 twice. A scenario with
no expected result is not testable, and nothing about the resulting file
looks wrong. Re-running the same three trials with the contract in place:
zero missing required fields, zero placeholder test data, zero duplicate
IDs, and the same four tabs every time.

The mechanism is that the contract IS the tool's input schema, so a
missing `expected` is rejected at the API boundary and retried, rather
than rendering into a workbook that looks complete.
"""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.services.artifact_types import (
    CONTRACTS,
    TEST_PLAN,
    contract_tool_schema,
    plan_from_contract,
)
from app.services.workbook_plan import render_workbook


def _emitted(**over):
    base = {
        "filename": "test_scenarios_081426.xlsx",
        "title": "Cigna demo test scenarios",
        "sheets": [{
            "name": "Refill",
            "rows": [{
                "id": "RFL-01", "scenario": "Standard refill",
                "type": "Happy path", "preconditions": "Active member",
                "test_data": "Rx #8834524, last filled 07/15/2025",
                "expected": "Refill approved, ship date returned.",
                "notes": "Baseline path.",
            }],
        }],
    }
    base.update(over)
    return base


def test_the_model_cannot_choose_the_columns() -> None:
    """Columns are ours. The emission schema has no column field at all."""
    schema = contract_tool_schema(TEST_PLAN)
    sheet = schema["properties"]["sheets"]["items"]
    assert "columns" not in sheet["properties"]
    assert set(sheet["required"]) == {"name", "rows"}


def test_load_bearing_fields_are_required_by_the_schema() -> None:
    row = contract_tool_schema(TEST_PLAN)["properties"]["sheets"]["items"][
        "properties"]["rows"]["items"]
    # These are the ones whose absence makes the artifact useless rather
    # than merely thinner.
    for key in ("id", "scenario", "type", "test_data", "expected"):
        assert key in row["required"], f"{key} must be required"
    # Genuinely optional ones must NOT be required, or the model pads.
    assert "notes" not in row["required"]
    assert "preconditions" not in row["required"]


def test_stray_fields_are_rejected_rather_than_silently_dropped() -> None:
    row = contract_tool_schema(TEST_PLAN)["properties"]["sheets"]["items"][
        "properties"]["rows"]["items"]
    assert row["additionalProperties"] is False


def test_every_column_carries_a_description_because_it_is_the_prompt() -> None:
    row = contract_tool_schema(TEST_PLAN)["properties"]["sheets"]["items"][
        "properties"]["rows"]["items"]
    for key, spec in row["properties"].items():
        assert spec.get("description"), f"{key} has no description"
    # The data column has to forbid placeholders explicitly: Haiku 4.5
    # emitted MEM-001-ACTIVE style labels for 7 of 25 scenarios.
    assert "placeholder" in row["properties"]["test_data"]["description"]


def test_contracted_emission_renders_with_the_contract_columns() -> None:
    plan = plan_from_contract(TEST_PLAN, _emitted())
    ws = openpyxl.load_workbook(io.BytesIO(render_workbook(plan)))["Refill"]
    assert [ws.cell(2, i).value for i in range(1, 8)] == [
        "Scenario ID", "Scenario", "Type", "Preconditions",
        "Sample Test Data", "Expected Behavior", "Notes for Dev"]
    assert ws["A3"].value == "RFL-01"
    assert ws["F3"].value.startswith("Refill approved")


def test_column_order_is_fixed_regardless_of_key_order_in_the_row() -> None:
    """The model emits a JSON object; key order there must not reach the
    file, or two runs produce differently ordered sheets."""
    shuffled = _emitted()
    shuffled["sheets"][0]["rows"][0] = dict(
        reversed(list(shuffled["sheets"][0]["rows"][0].items())))
    plan = plan_from_contract(TEST_PLAN, shuffled)
    ws = openpyxl.load_workbook(io.BytesIO(render_workbook(plan)))["Refill"]
    assert ws["A2"].value == "Scenario ID"
    assert ws["A3"].value == "RFL-01"


def test_an_omitted_optional_field_renders_blank_not_shifted() -> None:
    e = _emitted()
    del e["sheets"][0]["rows"][0]["notes"]
    plan = plan_from_contract(TEST_PLAN, e)
    ws = openpyxl.load_workbook(io.BytesIO(render_workbook(plan)))["Refill"]
    assert ws["F3"].value.startswith("Refill approved")
    assert ws["G3"].value in (None, "")


def test_filename_and_title_fall_back_when_the_model_omits_them() -> None:
    plan = plan_from_contract(TEST_PLAN, {"sheets": [{"name": "S",
                                                      "rows": []}]})
    assert plan["filename"].endswith(".xlsx")
    assert TEST_PLAN.filename_hint in plan["filename"]
    assert plan["title"] == TEST_PLAN.label


def test_registry_exposes_contracts_by_name() -> None:
    assert CONTRACTS["test_plan"] is TEST_PLAN
    for name, contract in CONTRACTS.items():
        assert contract.name == name
        assert contract.required, "a contract with nothing required is a lie"
        assert contract.sheet_rule


@pytest.mark.parametrize("contract", list(CONTRACTS.values()))
def test_each_contract_produces_a_usable_tool_schema(contract) -> None:
    schema = contract_tool_schema(contract)
    assert schema["type"] == "object"
    assert "sheets" in schema["required"]
    row = schema["properties"]["sheets"]["items"]["properties"]["rows"][
        "items"]
    assert set(row["required"]) <= set(row["properties"])


def test_numeric_fields_are_typed_as_numbers_not_strings() -> None:
    """A score that arrives as "4" sorts as text and sums to zero."""
    from app.services.artifact_types import RISK_REGISTER
    row = contract_tool_schema(RISK_REGISTER)["properties"]["sheets"][
        "items"]["properties"]["rows"]["items"]
    assert row["properties"]["likelihood"]["type"] == "integer"
    assert row["properties"]["impact"]["type"] == "integer"


def test_optional_fields_accept_null_so_the_model_stops_writing_na() -> None:
    from app.services.artifact_types import ACTION_REGISTER
    row = contract_tool_schema(ACTION_REGISTER)["properties"]["sheets"][
        "items"]["properties"]["rows"]["items"]
    assert row["properties"]["blocker"]["type"] == ["string", "null"]
    assert row["properties"]["owner"]["type"] == "string"  # required


def test_a_computed_column_is_never_asked_of_the_model() -> None:
    """Severity is likelihood times impact. Asking for it invites a
    number that disagrees with its own inputs."""
    from app.services.artifact_types import RISK_REGISTER
    row = contract_tool_schema(RISK_REGISTER)["properties"]["sheets"][
        "items"]["properties"]["rows"]["items"]
    assert "severity" not in row["properties"]
    assert "severity" not in row["required"]

    plan = plan_from_contract(RISK_REGISTER, {"sheets": [{
        "name": "Risks",
        "rows": [{"id": "R-01", "risk": "Vendor slips",
                  "likelihood": 4, "impact": 5, "owner": "Mike",
                  "mitigation": "Weekly checkpoint"}]}]})
    ws = openpyxl.load_workbook(io.BytesIO(render_workbook(plan)))["Risks"]
    sev = [c.value for c in ws[3]
           if isinstance(c.value, str) and c.value.startswith("=")]
    assert sev == ['=IFERROR(D3*E3,"")']


def test_comparison_options_become_columns_in_the_order_given() -> None:
    from app.services.artifact_types import OPTION_COMPARISON
    plan = plan_from_contract(OPTION_COMPARISON, {
        "options": ["Vendor A", "Vendor B", "Build it"],
        "sheets": [{"name": "Comparison", "rows": [
            {"criterion": "SSO support", "weight": 5,
             "values": ["Native", "Add-on", "We build it"]}]}]})
    ws = openpyxl.load_workbook(io.BytesIO(render_workbook(plan)))[
        "Comparison"]
    assert [ws.cell(2, i).value for i in (4, 5, 6)] == [
        "Vendor A", "Vendor B", "Build it"]
    assert [ws.cell(3, i).value for i in (4, 5, 6)] == [
        "Native", "Add-on", "We build it"]
    # The criterion label has to stay on screen on a wide grid.
    assert ws.freeze_panes == "B3"


def test_a_short_values_array_leaves_a_gap_rather_than_shifting() -> None:
    """Positional data is only safe if a missing entry cannot slide the
    next one under the wrong option."""
    from app.services.artifact_types import OPTION_COMPARISON
    plan = plan_from_contract(OPTION_COMPARISON, {
        "options": ["A", "B", "C"],
        "sheets": [{"name": "Comparison", "rows": [
            {"criterion": "Cost", "values": ["cheap"]}]}]})
    ws = openpyxl.load_workbook(io.BytesIO(render_workbook(plan)))[
        "Comparison"]
    assert ws.cell(3, 4).value == "cheap"
    assert ws.cell(3, 5).value in (None, "")
    assert ws.cell(3, 6).value in (None, "")


def test_budget_gets_a_live_totals_row() -> None:
    from app.services.artifact_types import BUDGET
    plan = plan_from_contract(BUDGET, {"sheets": [{"name": "Estimate",
        "rows": [
            {"id": "B-01", "line_item": "Contractor", "basis": "Mike said 2 devs",
             "low": 40000, "high": 60000, "confidence": "Stated"},
            {"id": "B-02", "line_item": "Licenses", "basis": "Not stated in meeting",
             "low": None, "high": None, "confidence": "Guess"}]}]})
    ws = openpyxl.load_workbook(io.BytesIO(render_workbook(plan)))["Estimate"]
    assert ws.cell(5, 1).value == "Total"
    formulas = [c.value for c in ws[5]
                if isinstance(c.value, str) and c.value.startswith("=")]
    assert '=IFERROR(SUM(E3:E4),"")' in formulas
    assert '=IFERROR(SUM(F3:F4),"")' in formulas


def test_the_budget_basis_column_forbids_inventing_figures() -> None:
    """The one artifact where a confident wrong number does real damage."""
    from app.services.artifact_types import BUDGET
    row = contract_tool_schema(BUDGET)["properties"]["sheets"]["items"][
        "properties"]["rows"]["items"]
    basis = row["properties"]["basis"]["description"].lower()
    assert "invented" in basis or "not stated" in basis
    assert "basis" in row["required"]


def test_every_contract_names_its_source_of_truth_or_says_why_not() -> None:
    """A meeting artifact a reader cannot trace back is an assertion."""
    traceable = {"action_register", "decision_log", "risk_register",
                 "open_questions", "requirements"}
    for name in traceable:
        keys = {c.key for c in CONTRACTS[name].columns}
        assert "source" in keys or "source_quote" in keys, name
